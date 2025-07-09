import requests
import numpy as np
import pandas as pd
import streamlit as st
import altair as alt
import time
import io
from openpyxl import load_workbook
from copy import copy
from dotenv import load_dotenv
import base64
from utils.ParamHandler import ParamHandler



# Load environment variables
load_dotenv()

# =======================================================================================================
# ============================================ DEFAULT VALUE ============================================
# =======================================================================================================

st.set_page_config(page_title="AS 1170.2:2021", page_icon="༄", layout="wide")  # wide

st.markdown("""
    <style>
        header[data-testid="stHeader"] {
            display: none !important;
        }
        .main > div {
            padding-top: 0rem;
        }
        .block-container {
            padding-top: 0rem;
            padding-bottom: 0rem;
        }
        [data-testid="stToolbar"] {
            display: none !important;
        }
        [data-testid="stDecoration"] {
            display: none !important;
        }
        #MainMenu {
            display: none !important;
        }
        [data-testid="stStatusWidget"] {
            display: none !important;
        }
        footer {
            display: none !important;
        }
    </style>
""", unsafe_allow_html=True)
# --- URL 파라미터 읽기 ---
_base_url = ParamHandler.base_url
_mapi_key = ParamHandler.mapi_key
base_url = f"{_base_url}/gen"
# =======================================================================================================
# ============================================== Data Base ==============================================
# =======================================================================================================
# STOR / STLD 데이터 유효성 체크 FLAG
st.session_state["stor_valid_flag"] = False
st.session_state["stld_valid_flag"] = False
# 변수 입력을 col2에 받기 위한 초기 상태값
if "mode" not in st.session_state:
    st.session_state["mfactor"] = None
if "result" not in st.session_state:
    st.session_state["calculate"] = None

# 화면 축소 비율
zoom_ratio = 0.75

# All Cardinal Directions
au_cardinal_directions = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]
au_cardinal_angle = [0, 45, 90, 135, 180, 225, 270, 315]
au_building_angle = [0, 90, 180, 270]

# terrain category
au_terrain_category = {
    "Category" : ["TC1", "TC2", "TC2.5", "TC3", "TC4", "Select Multiple"],
    "Explain" : [
        "Very exposed open terrain with very few or no obstructions, and all water surfaces (e.g. flat, treeless, poorly grassed plains; open ocean, rivers, canals, bays and lakes).",
        "Open terrain, including grassland, with well-scattered obstructions having heights generally from 1.5 m to 5 m, with no more than two obstructions per hectare (e.g. farmland and cleared subdivisions with isolated trees and uncut grass).",
        "Terrain with some trees or isolated obstructions, terrain in developing outer urban areas with scattered houses, or large acreage developments with more than two and less than 10 buildings per hectare.",
        "Terrain with numerous closely spaced obstructions having heights generally from 3 m to 10 m. The minimum density of obstructions shall be at least the equivalent of 10 house-size obstructions per hectare (e.g. suburban housing, light industrial estates or dense forests).",
        "Terrain with numerous large, high (10 m to 30 m tall) and closely-spaced constructions, such as large city centres and well-developed industrial complexes.",
        "Averaging of terrain categories and terrain-height multipliers."
    ]
}
au_terrain_category_modal = {
    "Category" : ["TC1", "TC2", "TC2.5", "TC3", "TC4"],
    "Explain" : [
        "Very exposed open terrain with very few or no obstructions, and all water surfaces (e.g. flat, treeless, poorly grassed plains; open ocean, rivers, canals, bays and lakes).",
        "Open terrain, including grassland, with well-scattered obstructions having heights generally from 1.5 m to 5 m, with no more than two obstructions per hectare (e.g. farmland and cleared subdivisions with isolated trees and uncut grass).",
        "Terrain with some trees or isolated obstructions, terrain in developing outer urban areas with scattered houses, or large acreage developments with more than two and less than 10 buildings per hectare.",
        "Terrain with numerous closely spaced obstructions having heights generally from 3 m to 10 m. The minimum density of obstructions shall be at least the equivalent of 10 house-size obstructions per hectare (e.g. suburban housing, light industrial estates or dense forests).",
        "Terrain with numerous large, high (10 m to 30 m tall) and closely-spaced constructions, such as large city centres and well-developed industrial complexes."
    ]
}

# Select Wind Region
au_wind_region = {
    "Region" : ["A0", "A1", "A2", "A3", "A4", "A5", "B1", "B2", "C", "D"]
}

# Climage chage multiplier Mc
# 뉴질랜드는 해당플러그인에서 적용 제외
au_mc = {
    "Region" : au_wind_region["Region"],
    "Mc" : [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.05, 1.05, 1.05]
}

# Md (Directional Factor)
au_md = {
    "Cardinal Directions" : ["N", "NE", "E", "SE", "S", "SW", "W", "NW"],
    "A0" : [0.90, 0.85, 0.85, 0.90, 0.90, 0.95, 1.00, 0.95],
    "A1" : [0.90, 0.85, 0.85, 0.80, 0.80, 0.95, 1.00, 0.95],
    "A2" : [0.85, 0.75, 0.85, 0.95, 0.95, 0.95, 1.00, 0.95],
    "A3" : [0.90, 0.75, 0.75, 0.90, 0.90, 0.95, 1.00, 0.95],
    "A4" : [0.85, 0.75, 0.75, 0.80, 0.80, 0.90, 1.00, 1.00],
    "A5" : [0.95, 0.80, 0.80, 0.80, 0.80, 0.95, 1.00, 0.95],
    "B1" : [0.75, 0.75, 0.85, 0.90, 0.95, 0.95, 0.95, 0.90],
    "B2" : [0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90],
    "C" : [0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90],
    "D" : [0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90]
}

# Define terrain category data (for mzcat)
au_mzcat_db = {
    "Heights" : [3, 5, 10, 15, 20, 30, 40, 50, 75, 100, 150, 200],
    "TC1" : [0.97, 1.01, 1.08, 1.12, 1.14, 1.18, 1.21, 1.23, 1.27, 1.31, 1.36, 1.39],
    "TC2" : [0.91, 0.91, 1.00, 1.05, 1.08, 1.12, 1.16, 1.18, 1.22, 1.24, 1.27, 1.29],
    "TC2.5" : [0.87, 0.87, 0.92, 0.97, 1.01, 1.06, 1.10, 1.13, 1.17, 1.20, 1.24, 1.27],
    "TC3" : [0.83, 0.83, 0.83, 0.89, 0.94, 1.00, 1.04, 1.07, 1.12, 1.16, 1.21, 1.24],
    "TC4" : [0.75, 0.75, 0.75, 0.75, 0.75, 0.80, 0.85, 0.90, 0.98, 1.03, 1.11, 1.16]
}

# Shielding Type (default : "No shielding")
au_shielding_category = {
    "shielding_catetory" : ["No shielding", "Shielding"]
}

# Hill Tpe (default : "DO not consider")
au_hill_type = {
    "hill_type" : ["Do not consider", "Ridge", "Escarpment"]
}

au_rho_air = 1.200 # kg/m3

# Cpe_windward
au_cpe_ww = {
    "Cpe_windward" : [0.8, 0.7]
}

# Cpe_leeward
au_cpe_lw = {
    "d/b" : [1, 2, 4],
    "Cpe" : [-0.5, -0.3, -0.2]
}

# Wind Direction 
au_wind_direction = {
    "wind_direction" : ["x_dir", "y_dir"]
}

# # Kc (combination factor)
# au_kc = {
#     "kc" : [1.0, 0.9, 0.8]
# }

au_structure_position = ["Upwind", "Downwind"]

# =======================================================================================================
# ============================================= Definitions =============================================
# =======================================================================================================
# Extract Story Data
def get_stor_data(body) :
    header = {"MAPI-Key":_mapi_key}
    try:
        response = requests.post(url=base_url + "/ope/storyprop", headers=header, json=body)
        response.raise_for_status()  # HTTP 오류 확인
        get_story_data_origin = response.json()
        st.session_state["stor_valid_flag"] = True
    except requests.exceptions.RequestException as e:
        get_story_data_origin = {
            "STORYPROP": {
                "FORCE": "KN",
                "LENGTH": "M",
                "HEAD": [
                    "Story",
                    "Weight",
                    "Elev.",
                    "Loaded H",
                    "Loaded Bx",
                    "Loaded By"
                ],
                "DATA": [
                    {
                        "STORY": "ROOF",
                        "WEIGHT": "46482.2030",
                        "ELEV": "7.5000",
                        "LOADED_H": "2.2500",
                        "LOADED_BX": "29.1000",
                        "LOADED_BY": "36.0000"
                    },
                    {
                        "STORY": "1F",
                        "WEIGHT": "6694.3802",
                        "ELEV": "3.0000",
                        "LOADED_H": "2.2500",
                        "LOADED_BX": "29.1000",
                        "LOADED_BY": "36.0000"
                    },
                    {
                        "STORY": "G.L.",
                        "WEIGHT": "0.0000",
                        "ELEV": "0.0000",
                        "LOADED_H": "0.0000",
                        "LOADED_BX": "29.1000",
                        "LOADED_BY": "36.0000"
                    }
                ]
            }
        }
        st.session_state["stor_valid_flag"] = None
    except ValueError:
        st.error("Please check the structure of the Story data")
        st.stop()

    story_name = []
    story_elev = []
    story_weight = []
    loaded_h = []
    bx = []
    by = []

    for stor_cnt in range(len(get_story_data_origin["STORYPROP"]["DATA"])) :
        # Get Story Data
        story_name.append(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["STORY"])
        story_elev.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["ELEV"]))
        story_weight.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["WEIGHT"]))
        loaded_h.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["LOADED_H"]))
        bx.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["LOADED_BX"]))
        by.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["LOADED_BY"]))

    return {
        "story_name" : story_name,
        "story_elev" : story_elev,
        "story_weight" : story_weight,
        "loaded_h" : loaded_h,
        "loaded_bx" : bx,
        "loaded_by" : by
    }
# # calculate sample
# story_data = get_stor_data()
# print(story_data)

def get_stld_data() :
    header = {"MAPI-Key": _mapi_key}
    try:
        response = requests.get(url=base_url + "/db/stld", headers=header)
        response.raise_for_status()  # HTTP 오류 확인
        stld_data = response.json()
        if stld_data["STLD"] == {} :
            st.session_state["stld_valid_flag"] = False
        else :
            st.session_state["stld_valid_flag"] = True
    except requests.exceptions.RequestException as e:
        st.error(f"서버 요청에 실패했습니다: {e}")
        st.stop()
    except ValueError:
        st.error("응답을 JSON으로 파싱하는 데 실패했습니다.")
        st.stop()

    stld_idx = []
    stld_name = []
    stld_type = []

    stld_idx = list(stld_data["STLD"].keys())
    for index in stld_idx :
        stld_name.append(stld_data["STLD"][str(index)]["NAME"])
        stld_type.append(stld_data["STLD"][str(index)]["TYPE"])

    return {
        "INDEX" : stld_idx,
        "NAME" : stld_name,
        "TYPE" : stld_type
    }

# Define Mz,cat for AU
def au_mzcat(wind_region, input_terrain_category = None, input_terrain_length = None) :
    mzcat = {}
    story_elev = st.session_state["story_data"]["story_elev"]
    # Terrain Category 지정 안됐을 때

    if wind_region == "A0" :
        mzcat_calc = []
        for z in story_elev :
            if z <= 100 :
                mzcat_calc.append(2.0)
            else :
                mzcat_calc.append(1.24)
    else :
        cat_len = len(input_terrain_category)
        # Terrain Category 1개일 때
        if cat_len == 0 :
            st.error('Check the input(Mz,cat) data', icon="🚨")
        elif cat_len == 1 :
            mzcat_calc = []
            for cnt in range(len(story_elev)) :
                # Input mzcat
                mzcat_calc.append(np.interp(st.session_state["story_data"]["story_elev"][cnt], au_mzcat_db["Heights"], au_mzcat_db[input_terrain_category[0]]))

        # Terrain Category 2개 이상일 때
        elif cat_len >= 2 :
            xa_list = []
            mzcat_calc = []
            st.session_state["length_modified_flag"] = False

            # 25m 이하일 경우, 높이 별 xa값 고정 (z = h)
            if story_elev[0] <= 25  :
                z = story_elev[0]
                xa = max(500, 40*z)
                mzcat_temp = []

                # 추후 xa값 df로 추출할 수 있도록 기반 마련
                for index, z in enumerate(story_elev) :
                    xa_list.append(xa)

                # 가중평균하기 위한 mzcat list 작성
                if sum(input_terrain_length) == xa :
                    mzcat_temp.append(np.interp(z, au_mzcat_db["Heights"], au_mzcat_db[input_terrain_category[index]]))
                    mzcat_temp = mzcat_temp * len(story_elev)
                elif sum(input_terrain_length) < xa :
                    st.session_state["length_modified_flag"] = True
                    input_terrain_length[-1] = input_terrain_length[-1] + (xa - sum(input_terrain_length))
                    for category in input_terrain_category :
                        mzcat_temp.append(np.interp(z, au_mzcat_db["Heights"], au_mzcat_db[category]))
                    mzcat_calc.append(sum([mzcat * xt for mzcat, xt in zip(mzcat_temp, input_terrain_length)]) / sum(input_terrain_length))
                else :
                    st.session_state["length_modified_flag"] = True
                    length_sum = 0
                    for length_idx, length in enumerate(input_terrain_length) :
                        length_sum += length
                        if length_sum >= xa :
                            input_terrain_category = input_terrain_category[:length_idx+1]
                            input_terrain_length = input_terrain_length[:length_idx+1]
                            input_terrain_length[-1] = input_terrain_length[-1] - (sum(input_terrain_length) - xa)
                            break
                    for category in input_terrain_category :
                        mzcat_temp.append(np.interp(z, au_mzcat_db["Heights"], au_mzcat_db[category]))
                    mzcat_calc.append(sum([mzcat * xt for mzcat, xt in zip(mzcat_temp, input_terrain_length)]) / sum(input_terrain_length))
                            
            # 25m 초과일 경우, 높이 별 xa값 다른 부분 고려하여 mz,cat 계산
            else :
                # 추후 xa값 df로 추출할 수 있도록 기반 마련
                for z in story_elev :
                    xa = max(500, 40*z)
                    xa_list.append(xa)
                    mzcat_temp = []
                    
                    # 층별로 z값에 따라 다른 xa값을 고려하여 mzcat 계산
                    if sum(input_terrain_length) == xa :
                        pass
                    
                    elif sum(input_terrain_length) < xa :
                        st.session_state["length_modified_flag"] = True
                        input_terrain_length[-1] = input_terrain_length[-1] + (xa - sum(input_terrain_length))

                    else :
                        st.session_state["length_modified_flag"] = True
                        length_sum = 0
                        for length_idx, length in enumerate(input_terrain_length) :
                            length_sum += length
                            if length_sum >= xa :
                                input_terrain_category = input_terrain_category[:length_idx+1]
                                input_terrain_length = input_terrain_length[:length_idx+1]
                                input_terrain_length[-1] = input_terrain_length[-1] - (sum(input_terrain_length) - xa)
                                break

                    for category in input_terrain_category :
                        mzcat_temp.append(np.interp(z, au_mzcat_db["Heights"], au_mzcat_db[category]))
                    mzcat_calc.append(sum([mzcat * xt for mzcat, xt in zip(mzcat_temp, input_terrain_length)]) / xa)


        else :
            raise ValueError("Check the terrain category input data")
    mzcat["Mz,cat"] = mzcat_calc
    return mzcat

# Define Ms
def au_ms(height, hs, bs, ns) :
    # calculate shielding parameter, s
    db_shielding_parameter = {
        "s" : [1.5, 3, 6, 12],
        "Ms" : [0.7, 0.8, 0.9, 1.0]
    }
    # Calculate Ms
    # Calculate l_s
    ls = height * (10/ns + 5)
    # Calculate s
    s = ls / ((hs * bs)**0.5)
    # Calculate Ms
    ms = np.interp(s, db_shielding_parameter["s"], db_shielding_parameter["Ms"])
    return {
        "Ms" : ms
    }
# # calculate sample
# ms_dict = au_ms(48, "Shielding", 10, 20, 5)
# print(ms_dict)

# Define Mt
def au_mt(building_height, input_hill_type, hill_height=None, Lu=None, x=None, E=0.0) :
    mh_ww_list = []
    mh_lw_list = []

    hill_shape = input_hill_type
    h_hill = hill_height
    lu = Lu
    dist = x

    db_mh_table = {
        "up_slope" : [0.05, 0.05, 0.1, 0.2, 0.3, 0.45],
        "mh" : [1.0, 1.08, 1.16, 1.32, 1.48, 1.71]
    }
    
    if input_hill_type == "Do not consider" :
        mh_ww = mh_lw = 1.0
        for z in st.session_state["story_data"]["story_elev"] :
            mh_ww_list.append(mh_ww)
            mh_lw_list.append(mh_lw)
    elif input_hill_type != "Do not consider" :
        if hill_height == None or Lu == None or x == None :
            raise ValueError("Check the hill type input data")
        else :
            slope = h_hill/(2*lu)

            # Calculate L1
            l1 = max(0.36*lu, 0.4*h_hill)

            # Calculate L2
            if hill_shape == "Escarpment" :
                l2_up = 4 * l1
                l2_dn = 10 * l1
            else :
                l2_up = l2_dn = 4 * l1

            # Calculate Mh (up / down)
            if h_hill < min(0.4*building_height, 5) :
                story_count = len(st.session_state["story_data"]["story_elev"])
                mh_ww_list = [1.0] * story_count
                mh_lw_list = [1.0] * story_count
            else :
                for z in st.session_state["story_data"]["story_elev"] :
                    # Calculate Mh (up / down)
                    if z == 0 and x == 0 : # Table 4.3 에 따라 Mh 산정
                        if slope < 0.05 :
                            mh_ww = mh_lw = db_mh_table["mh"][0]
                        else :
                            mh_ww = np.interp(slope, db_mh_table["up_slope"][1:], db_mh_table["mh"][1:])
                            mh_lw = mh_ww
                    else : # 4.4.2 a) ~ c)에 따라 계산
                        if slope < 0.05 :
                            mh_ww = mh_lw = 1
                        elif slope <= 0.45 :
                            mh_ww = 1 + (h_hill / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_up)
                            mh_lw = 1 + (h_hill / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_dn)
                        elif slope > 0.45 :
                            if dist < h_hill/4 and z <= h_hill/10 :
                                mh_ww = 1 + 0.71 * (1 - abs(dist)/l2_up)
                                mh_lw = 1 + 0.71 * (1 - abs(dist)/l2_dn)
                            else :
                                mh_ww = 1 + (h_hill / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_up)
                                mh_lw = 1 + (h_hill / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_dn)
                    mh_ww_list.append(mh_ww)
                    mh_lw_list.append(mh_lw)

    # Calculate Mt                    
    if st.session_state["wind_region"] == "A4" :
        mt_ww_list = [x * (1+0.00015*E) for x in mh_ww_list]
        mt_lw_list = [x * (1+0.00015*E) for x in mh_lw_list]
    elif st.session_state["wind_region"] == "A0" :
        mt_ww_list = [0.5 + 0.5 * x for x in mh_ww_list]
        mt_lw_list = [0.5 + 0.5 * x for x in mh_lw_list]
    else :
        mt_ww_list = mh_ww_list
        mt_lw_list = mh_lw_list

    return {
        "Mt(ww)" : mt_ww_list,
        "Mt(lw)" : mt_lw_list
    }

# Define Cpe_external / Cpi는 추후 wind pressure 기능으로 따로 제작해야 할 듯
def au_cpe_leeward(theta) :
    # Cpe_Lee Ward 계산
    bx = [float(x) for x in st.session_state["story_data"]["loaded_bx"]]
    by = [float(y) for y in st.session_state["story_data"]["loaded_by"]]
    db_x_dir = [bx/by for bx, by in zip(bx, by)]
    db_y_dir = [by/bx for bx, by in zip(bx, by)]

    # alpha = 0 으로 가정
    if theta == 0 or theta == 180 :
        cpe_lw = np.interp(db_x_dir, au_cpe_lw["d/b"], au_cpe_lw["Cpe"])
    elif theta == 90 or theta == 270 :
        cpe_lw = np.interp(db_y_dir, au_cpe_lw["d/b"], au_cpe_lw["Cpe"])
    else :
        raise ValueError(f"Invalid theta: {theta}. Only 0, 90, 180, 270 degrees are supported.")
    cpe_lw_list = cpe_lw.tolist()

    return {
        "Cpe_lw" : cpe_lw_list
    }
# # calculate sample
# d_b = au_cpe()
# print(f"d/b = {d_b}")

# copy value (층 개수만큼 복사)
def copy_value_by_story(number, input_value) :
    return [input_value] * number if not isinstance(input_value, (list, pd.Series)) else input_value

# Modal Dialog : Mz,cat
# 모달 창에서는 다수의 Terrain Cateogry 및 길이만을 입력받도록 정리
# 최종 정리된 값을 au_mzcat() 함수에 전달하여 Mz,cat을 계산
@st.dialog("Select Multiple Terrain Category")
def au_mzcat_modal(cardinal_direction) :
    def add_row() :
        selected_value = st.session_state[f"terrain_cat_drop_multi_{cardinal_direction}"]
        if selected_value is not None:
            st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Category"].append(selected_value)
            st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Length(m)"].append(1000)

    def renew_table_data() :
        edited_row = st.session_state["edited_df"]["edited_rows"]  # dictionary type (index + value)
        added_row = st.session_state["edited_df"]["added_rows"]  # list type (index는 무조건 0부터 시작)
        indices_deleted = st.session_state["edited_df"]["deleted_rows"]  # index list type

        for index in edited_row :
            if len(edited_row[index].keys()) == 0 :
                continue
            elif len(edited_row[index].keys()) == 1 :
                if "Terrain Category" in edited_row[index] :
                    st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Category"][index] = edited_row[index]["Terrain Category"]
                elif "Terrain Length(m)" in edited_row[index] :
                    st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Length(m)"][index] = edited_row[index]["Terrain Length(m)"]
            elif len(edited_row[index].keys()) == 2 :
                st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Category"][index] = edited_row[index]["Terrain Category"]
                st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Length(m)"][index] = edited_row[index]["Terrain Length(m)"]
            else :
                st.error("Please check the data")

        for rows_value in added_row :
            if len(rows_value.keys()) == 0 :
                continue
            elif len(rows_value.keys()) == 2 :
                st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Category"].append(rows_value["Terrain Category"])
                st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Length(m)"].append(rows_value["Terrain Length(m)"])
            else :
                st.error("Please check the data")

        for index in sorted(indices_deleted, reverse=True) :
            st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Category"].pop(index)
            st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Length(m)"].pop(index)

    st.info("Start entering values from the location nearest to the structure.")

    if f"terrain_table_data_{cardinal_direction}" not in st.session_state:
        st.session_state[f"terrain_table_data_{cardinal_direction}"] = {
            "Terrain Category": [],
            "Terrain Length(m)": []
        }

    mzcat_col_1, mzcat_col_2 = st.columns(2, vertical_alignment ="center")
    with mzcat_col_1 :
        st.write("Terrain Category")
    with mzcat_col_2 :
        st.selectbox("Terrain Category", au_terrain_category_modal["Category"], index=None, key=f"terrain_cat_drop_multi_{cardinal_direction}", label_visibility="collapsed", on_change=add_row)
    
    # 다수 Terrain Category 입력 테이블 출력
    st.data_editor(
        data=st.session_state[f"terrain_table_data_{cardinal_direction}"],
        hide_index=True,
        use_container_width=True,
        num_rows="dynamic",
        key="edited_df",
        on_change=renew_table_data
    )
    
    if len(st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Category"]) * len(st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Length(m)"]) == 0 :
        st.warning("Please select or input terrain category", icon="⚠️")
    
    if st.button("Calculate", use_container_width=True) :
        try :
            mzcat_result = au_mzcat(
                st.session_state["wind_region"],
                st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Category"],
                st.session_state[f"terrain_table_data_{cardinal_direction}"]["Terrain Length(m)"]
            )
            st.session_state[f"mzcat_result_{cardinal_direction}"] = mzcat_result

            del st.session_state["edited_df"]
                                
            if st.session_state["length_modified_flag"] == True :
                st.warning(f"Terrain Category Length has been modified to align with xa(max)")
                time.sleep(2)

            st.rerun()

        except Exception as e:
            st.warning(f"❗ Exception: {e}")

# Modal Dialog : Ms
@st.dialog("Calculate Mₛ")
def au_ms_modal(height, wind_direction) :
    st.session_state["calc_ms_button"] = False
    h_key = f"h_{wind_direction}"
    hs_key = f"hs_{wind_direction}"
    bs_key = f"bs_{wind_direction}"
    ns_key = f"ns_{wind_direction}"
    st.session_state[h_key] = height
    if hs_key not in st.session_state:
        st.session_state[hs_key] = 100.0
    if bs_key not in st.session_state:
        st.session_state[bs_key] = 10.0
    if ns_key not in st.session_state:
        st.session_state[ns_key] = 5

    ms_col1, ms_col2 = st.columns(2, vertical_alignment ="center")
    with ms_col1 :
        st.markdown("h<sub>s</sub>(m)", unsafe_allow_html=True)
    with ms_col2 :
        st.number_input("hs(m)", min_value=0.0, step=1.0, value=float(st.session_state[hs_key]), key="hs_temp", label_visibility="collapsed")
        st.session_state[hs_key] = st.session_state["hs_temp"]

    ms_col3, ms_col4 = st.columns(2, vertical_alignment ="center")
    with ms_col3 :
        st.markdown("b<sub>s</sub>(m)", unsafe_allow_html=True)
    with ms_col4 :
        st.number_input("bs(m)", min_value=0.0, step=1.0, value=float(st.session_state[bs_key]), key="bs_temp", label_visibility="collapsed")
        st.session_state[bs_key] = st.session_state["bs_temp"]

    ms_col5, ms_col6 = st.columns(2, vertical_alignment ="center")
    with ms_col5 :
        st.markdown("n<sub>s</sub>", unsafe_allow_html=True)
    with ms_col6 :
        st.number_input("ns", min_value=0, step=1, value=int(st.session_state[ns_key]), key="ns_temp", label_visibility="collapsed")
        st.session_state[ns_key] = st.session_state["ns_temp"]
    if st.button("Calculate", use_container_width=True) :
        try :
            ms_result = au_ms(height, st.session_state[hs_key], st.session_state[bs_key], st.session_state[ns_key])
            st.session_state[f"ms_result_{wind_direction}"] = ms_result
            st.session_state["calc_ms_button"] = True
            st.rerun()
        except Exception as e:
            st.warning(f"❗ Exception: {e}")

# Modal Dialog : Mh
@st.dialog("Calculate Mh")
def au_mt_modal(hill_type, wind_direction) :
    hill_height_temp_key = f"hill_height_temp_{wind_direction}"
    lu_temp_key = f"lu_temp_{wind_direction}"
    x_dist_temp_key = f"x_dist_temp_{wind_direction}"
    hill_height_key = f"hill_height_{wind_direction}"
    lu_key = f"lu_{wind_direction}"
    x_dist_key = f"x_dist_{wind_direction}"
    e_key = f"E_{wind_direction}"

    if hill_height_key not in st.session_state:
        st.session_state[hill_height_key] = 5.0
    if lu_key not in st.session_state:
        st.session_state[lu_key] = 10.0
    if x_dist_key not in st.session_state:
        st.session_state[x_dist_key] = 0.0
    if e_key not in st.session_state:
        st.session_state[e_key] = 0.0

    col1_1, col1_2 = st.columns(2, vertical_alignment ="center")
    with col1_1 :
        st.write("Height of Hill(m)")
    with col1_2 :
        st.number_input("hill_height(m)", min_value=0.1, step=1.0, value=float(st.session_state[hill_height_key]), key=hill_height_temp_key, label_visibility="collapsed")
        st.session_state[hill_height_key] = st.session_state[hill_height_temp_key]

    col1_3, col1_4 = st.columns(2, vertical_alignment ="center")
    with col1_3 :
        st.markdown("L<sub>u</sub>(m)", unsafe_allow_html=True)
    with col1_4 :
        st.number_input("Lu(m)", min_value=0.1, step=1.0, value=float(st.session_state[lu_key]), key=lu_temp_key, label_visibility="collapsed")
        st.session_state[lu_key] = st.session_state[lu_temp_key]

    col1_5, col1_6 = st.columns(2, vertical_alignment ="center")
    with col1_5 :
        st.write("x(m)")
    with col1_6 :
        st.number_input("x(m)", min_value=0.0, step=1.0, value=float(st.session_state[x_dist_key]), key=x_dist_temp_key, label_visibility="collapsed")
        st.session_state[x_dist_key] = st.session_state[x_dist_temp_key]

    col1_7, col1_8 = st.columns(2, vertical_alignment ="center")
    with col1_7 :
        st.write("E(m)")
    with col1_8 :
        st.number_input("E", min_value=0.0, step=0.1, value=float(st.session_state[e_key]), key=f"E_temp_{wind_direction}", label_visibility="collapsed", help="When the wind region is A4")
        st.session_state[e_key] = st.session_state[f"E_temp_{wind_direction}"]

    if st.button("Calculate", use_container_width=True) :
        try :
            if st.session_state["wind_region"] == "A4" :
                mt_result = au_mt(height, hill_type, st.session_state[hill_height_key], st.session_state[lu_key], st.session_state[x_dist_key], st.session_state[e_key])
                st.session_state[f"mt_result_{wind_direction}"] = mt_result
                st.rerun()
            else :
                mt_result = au_mt(height, hill_type, st.session_state[hill_height_key], st.session_state[lu_key], st.session_state[x_dist_key])
                st.session_state[f"mt_result_{wind_direction}"] = mt_result
                st.rerun()
        except Exception as e:
            st.warning(f"❗ Exception: {e}")

def velocity_line_chart(df: pd.DataFrame, selected_stories: list, sort: list, title: str, height: int):
    """
    Altair 라인차트로 층별 방향 데이터 시각화

    Parameters:
    - df (pd.DataFrame): 'Story'와 'N', 'NE', ..., 'NW' 컬럼을 포함한 데이터프레임
    - selected_stories (list): 시각화할 층 이름 리스트
    """

    sort_str = [str(angle) for angle in sort]

    # 데이터를 melt하여 long-format으로 변환
    melted_df = df.melt(
        id_vars=['Story'],
        value_vars=sort_str,
        var_name='Direction',
        value_name='Value'
    )

    # 선택한 층만 필터링
    filtered_df = melted_df[melted_df['Story'].isin(selected_stories)]
    min_val = filtered_df['Value'].min()
    max_val = filtered_df['Value'].max()
    padding = int(2 * zoom_ratio)
    y_min = min_val - padding
    y_max = max_val + padding

    # Altair 차트 생성
    chart = alt.Chart(filtered_df).mark_line(point=True, opacity=1).encode(
        x=alt.X('Direction:N', sort=sort_str, axis=alt.Axis(title=None)),
        y=alt.Y('Value:Q', scale=alt.Scale(domain=[y_min, y_max])),
        color=alt.Color('Story:N', sort=selected_stories, legend=alt.Legend(title=None))
    ).properties(
        # width=700 * zoom_ratio,
        height=height,
        title=title
    )

    st.altair_chart(chart, use_container_width=True)

def vdes_extract(df: pd.DataFrame, target_angle, orientation_angle) -> pd.DataFrame:
    """
    주어진 풍속 DataFrame에서 각 층별로 주어진 각도에 대해 선형 보간한 풍속 값을 반환합니다.

    Parameters:
    - df: 8방향(N, NE, E, SE, S, SW, W, NW) 풍속 데이터를 포함한 DataFrame
    - target_angles: 보간을 원하는 각도 리스트 (ex. [50, 140, 230, 320])

    Returns:
    - 보간된 풍속 값을 담은 DataFrame
    """

    target_angle = float(target_angle)
    orientation_angle = float(orientation_angle)

    # 원본 복사 후 360도 방향(N 중복) 추가
    df = df.copy()
    df["360"] = df["N"]

    # 각 방향별 각도 정의
    direction_angles = {
        "N": 0,
        "NE": 45,
        "E": 90,
        "SE": 135,
        "S": 180,
        "SW": 225,
        "W": 270,
        "NW": 315,
        "360": 360
    }

    # 보간을 위한 기준 각도 및 순서
    direction_cols = list(direction_angles.keys())
    angle_keys = list(direction_angles.values())

    # 결과 저장용
    result_list = []

    for _, row in df.iterrows():
        values = row[direction_cols].astype(float).values
        interpolated = np.interp(target_angle, angle_keys, values)
        interpolated = np.maximum(interpolated, 30)
        result_list.append(interpolated)

    # 결과 DataFrame 구성
    # result_df = pd.DataFrame(result_list, columns=[f"{int(target_angle-orientation_angle)}°"])
    return result_list

def line_chart_hor_multi(df: pd.DataFrame, unit: str):
    df = df.copy()
    
    col_story = df.columns[0]
    value_cols = df.columns[1:]  # 예: Axial, Shear, Moment, Displacement

    # long-form으로 변환
    df_melted = df.melt(id_vars=[col_story], value_vars=value_cols,
                        var_name='Series', value_name='Value')

    # Story 순서 정렬 (Roof → G.L.)
    story_order = list(df[col_story])
    df_melted[col_story] = pd.Categorical(df_melted[col_story], categories=story_order, ordered=True)
    df_melted["Value"] = df_melted["Value"].astype(float)

    # Altair 수평 라인차트
    chart = (
        alt.Chart(df_melted)
        .mark_line(point=True)
        .encode(
            y=alt.Y(f"{col_story}:N", sort=story_order, title=col_story),
            x=alt.X("Value:Q", title=f"Value ({unit})"),
            color=alt.Color("Series:N", title="Type", sort=au_building_angle_new),
            tooltip=[col_story, "Series", "Value"]
        )
        .properties(height=int(622 * zoom_ratio))
    )

    return chart

def bar_chart_hor(df: pd.DataFrame, unit : str):
    # 데이터 복사 및 컬럼 이름 자동 감지
    df = df.copy()
    col_y = df.columns[0]  # Story (범주형)
    col_x = df.columns[1]  # Force 등 (수치형)

    # Story 순서 역순으로 지정 (Roof → G.L.)
    story_order = list(df[col_y])
    df[col_y] = pd.Categorical(df[col_y], categories=story_order, ordered=True)

    # 수치형으로 안전하게 변환
    df[col_x] = df[col_x].astype(float)

    # Altair 그래프 생성
    bar = (
        alt.Chart(df)
        .mark_bar()
        .encode(
            y=alt.Y(f"{col_y}:N", sort=story_order, title=col_y),
            x=alt.X(f"{col_x}:Q", title=f"{col_x} {unit}"),
            tooltip=[col_y, col_x]
        )
    )
    text = (
        alt.Chart(df)
        .mark_text(
            align='right',
            baseline='middle',
            dx=-6,
            color='white',  # 흰색 텍스트
            fontSize=int(14 * zoom_ratio)  # 텍스트 크기 조정
        )
        .encode(
            y=alt.Y(f"{col_y}:N", sort=story_order),
            x=alt.X(f"{col_x}:Q"),
            text=alt.Text(f"{col_x}:Q", format=".2f")
        )
    )
    chart = (bar + text).properties(height=int(670 * zoom_ratio))
    return chart

def bar_chart_hor_multi(df: pd.DataFrame, unit: str):
    df = df.copy()
    
    col_story = df.columns[0]
    value_cols = df.columns[1:]

    # long-form으로 변환
    df_melted = df.melt(
        id_vars=[col_story],
        value_vars=value_cols,
        var_name='Series',
        value_name='Value'
    )

    # Story 순서 고정
    story_order = list(df[col_story])
    df_melted[col_story] = pd.Categorical(df_melted[col_story], categories=story_order, ordered=True)

    # 누적 수평 막대 차트 (stacked bar chart)
    chart = (
        alt.Chart(df_melted)
        .mark_bar(opacity=0.7)
        .encode(
            y=alt.Y(f"{col_story}:N", sort=story_order, title=col_story),
            x=alt.X("Value:Q", title=f"Value ({unit})").stack(None),
            color=alt.Color("Series:N", title="Series"),
            tooltip=[col_story, "Series", "Value"]
        )
        .properties(height=int(670 * zoom_ratio))
    )

    return chart

@st.dialog("Select Load Case")
def select_stld_modal(stld_name) :
    select_apply_direction = st.toggle("Apply X & Y Only", key="critical_direcction", help="Based on top floor [p] value")
    st.divider()
    if select_apply_direction :
        col1, col2 = st.columns(2)
        with col1 :
            st.write("Load Case for X Dir")
        with col2 :
            st.selectbox("Load Case for X-Dir", options=stld_name, key="stld_x_temp", label_visibility="collapsed")

        col3, col4 = st.columns(2)
        with col3 :
            st.write("Load Case for Y Dir")
        with col4 :
            st.selectbox("Load Case for Y-Dir", options=stld_name, key="stld_y_temp", label_visibility="collapsed")

        st.session_state["stld_x"] = st.session_state["stld_x_temp"]
        st.session_state["stld_y"] = st.session_state["stld_y_temp"]

        stld_x_name = st.session_state["stld_x"]
        stld_y_name = st.session_state["stld_y"]

        stld_x_idx = st.session_state["stld_data"]["NAME"].index(stld_x_name)
        stld_y_idx = st.session_state["stld_data"]["NAME"].index(stld_y_name)

        stld_x_type = st.session_state["stld_data"]["TYPE"][stld_x_idx]
        stld_y_type = st.session_state["stld_data"]["TYPE"][stld_y_idx]

        if st.button("Apply", use_container_width=True) :
            if stld_x_type != "W" or stld_y_type != "W" :
                st.warning("Wind load case must be selected.", icon="❗")
            elif stld_x_name == stld_y_name :
                st.warning("Load Case - X and Load Case - Y must be different.", icon="❗")
            else :
                error_flag = False
                # 방향 별 변수를 입력 받는지 여부에 따라 df 이름이 다르기 때문에 구분
                if st.session_state["input_condition"] == "Input Only Worst-Case Value" :
                    calculated_df_name = "wind_calc_all_df_"
                else :
                    calculated_df_name = "wind_calc_each_df_"
               
                # 최상층 p 값을 기준으로 X, Y 방향에 어떤 DF을 적용할 지 정리
                if st.session_state[f"{calculated_df_name}0"]["p [kPa]"].iloc[0] >= st.session_state[f"{calculated_df_name}180"]["p [kPa]"].iloc[0] :
                    selected_df_x = st.session_state[f"{calculated_df_name}0"]
                    scale_x = 1
                    building_angle_x = 0
                else :
                    selected_df_x = st.session_state[f"{calculated_df_name}180"]
                    scale_x = -1
                    building_angle_x = 180
 
                if st.session_state[f"{calculated_df_name}90"]["p [kPa]"].iloc[0] >= st.session_state[f"{calculated_df_name}270"]["p [kPa]"].iloc[0] :
                    selected_df_y = st.session_state[f"{calculated_df_name}90"]
                    scale_y = 1
                    building_angle_y = 90
                else :
                    selected_df_y = st.session_state[f"{calculated_df_name}270"]
                    scale_y = -1
                    building_angle_y = 270
 
                for direction in ["x", "y"] :
                    try :
                        header = {"MAPI-Key":_mapi_key}
                        body = {
                            "Assign": {
                                 st.session_state["stld_data"]["INDEX"][stld_x_idx if direction == "x" else stld_y_idx]: {
                                    "WIND_CODE": "USER TYPE",
                                    "DESC": "",
                                    "SCALE_FACTOR_X": scale_x if direction == "x" else 0,
                                    "SCALE_FACTOR_Y": scale_y if direction == "y" else 0,
                                    "SCALE_FACTOR_R": 0,
                                    "USER": apply_wind_data(selected_df_x if direction == "x" else selected_df_y, building_angle_x if direction == "x" else building_angle_y)
                                }
                            }
                        }
 
                        res = requests.put(url=f"{base_url}/db/wind", headers=header, json=body)
                        if res.status_code == 200:
                            st.success(f"{stld_x_name if direction == 'x' else stld_y_name} wind loads successfully applied.", icon="✅")
                        else:
                            error_detail = res.json().get("message") or res.text
                            st.error(f"❌ Failed to apply {stld_x_name if direction == 'x' else stld_y_name} loads.\n\n**Status Code:** {res.status_code}\n**Details:** {error_detail}")
                            error_flag = True
 
                    except Exception as e:
                        st.error(f"❗ Unexpected error occurred while sending API request.\n\n**Exception:** {e}")
                   
                if not error_flag :
                    time.sleep(1)
                    st.rerun()
    else :
        col1, col2 = st.columns(2)
        with col1 :
            st.write("Load Case for +X Dir")
        with col2 :
            st.selectbox("Load Case for +X-Dir", options=stld_name, key="stld_pos_x_temp", label_visibility="collapsed")

        col3, col4 = st.columns(2)
        with col3 :
            st.write("Load Case for +Y Dir")
        with col4 :
            st.selectbox("Load Case for +Y-Dir", options=stld_name, key="stld_pos_y_temp", label_visibility="collapsed")

        col5, col6 = st.columns(2)
        with col5 :
            st.write("Load Case for -X Dir")
        with col6 :
            st.selectbox("Load Case for -X-Dir", options=stld_name, key="stld_neg_x_temp", label_visibility="collapsed")

        col7, col8 = st.columns(2)
        with col7 :
            st.write("Load Case for -Y Dir")
        with col8 :
            st.selectbox("Load Case for -Y-Dir", options=stld_name, key="stld_neg_y_temp", label_visibility="collapsed")
            
        selected_stld_name = [st.session_state["stld_pos_x_temp"], st.session_state["stld_pos_y_temp"], st.session_state["stld_neg_x_temp"], st.session_state["stld_neg_y_temp"]]
        has_duplicates = len(selected_stld_name) != len(set(selected_stld_name))
        
        st.session_state["stld_pos_x"] = st.session_state["stld_pos_x_temp"]
        st.session_state["stld_pos_y"] = st.session_state["stld_pos_y_temp"]
        st.session_state["stld_neg_x"] = st.session_state["stld_neg_x_temp"]
        st.session_state["stld_neg_y"] = st.session_state["stld_neg_y_temp"]

        stld_pos_x_name = st.session_state["stld_pos_x"]
        stld_pos_y_name = st.session_state["stld_pos_y"]
        stld_neg_x_name = st.session_state["stld_neg_x"]
        stld_neg_y_name = st.session_state["stld_neg_y"]

        stld_pos_x_idx = st.session_state["stld_data"]["NAME"].index(stld_pos_x_name)
        stld_pos_y_idx = st.session_state["stld_data"]["NAME"].index(stld_pos_y_name)
        stld_neg_x_idx = st.session_state["stld_data"]["NAME"].index(stld_neg_x_name)
        stld_neg_y_idx = st.session_state["stld_data"]["NAME"].index(stld_neg_y_name)

        stld_pos_x_type = st.session_state["stld_data"]["TYPE"][stld_pos_x_idx]
        stld_pos_y_type = st.session_state["stld_data"]["TYPE"][stld_pos_y_idx]
        stld_neg_x_type = st.session_state["stld_data"]["TYPE"][stld_neg_x_idx]
        stld_neg_y_type = st.session_state["stld_data"]["TYPE"][stld_neg_y_idx]

        if st.button("Apply", use_container_width=True) :
            if stld_pos_x_type != "W" or stld_pos_y_type != "W" or stld_neg_x_type != "W" or stld_neg_y_type != "W" :
                st.warning("Wind load case must be selected.", icon="❗")
            elif has_duplicates == True :
                st.warning("Load Case - X and Load Case - Y must be different.", icon="❗")
            else :
                error_flag = False
                for building_angle in au_building_angle :
                    try :
                        header = {"MAPI-Key": mapi_key}
                        
                        # building angle에 따른 scale factor와 index값 지정
                        if building_angle == 0 :
                            scale_x = 1
                            scale_y = 0
                            stld_idx = stld_pos_x_idx
                            stld_type = st.session_state["stld_pos_x"]
                        elif building_angle == 90 :
                            scale_x = 0
                            scale_y = 1
                            stld_idx = stld_pos_y_idx
                            stld_type = st.session_state["stld_neg_x"]
                        elif building_angle == 180 :
                            scale_x = -1
                            scale_y = 0
                            stld_idx = stld_neg_x_idx
                            stld_type = st.session_state["stld_pos_y"]
                        elif building_angle == 270 :
                            scale_x = 0
                            scale_y = -1
                            stld_idx = stld_neg_y_idx
                            stld_type = st.session_state["stld_neg_y"]

                        # 방향 별 변수를 입력 받는지 여부에 따라 df 이름이 다르기 때문에 구분
                        if st.session_state["input_condition"] == "Input Only Worst-Case Value" :
                            calculated_df_name = "wind_calc_all_df_"
                        else :
                            calculated_df_name = "wind_calc_each_df_"

                        selected_df_name = calculated_df_name + str(building_angle)
                        selected_df = st.session_state[selected_df_name]

                        body = {
                            "Assign": {
                                 st.session_state["stld_data"]["INDEX"][stld_idx]: {
                                    "WIND_CODE": "USER TYPE",
                                    "DESC": "",
                                    "SCALE_FACTOR_X": scale_x,
                                    "SCALE_FACTOR_Y": scale_y,
                                    "SCALE_FACTOR_R": 0,
                                    "USER": apply_wind_data(selected_df, building_angle)
                                }
                            }
                        }

                        res = requests.put(url=f"{base_url}/db/wind", headers=header, json=body)
                        if res.status_code == 200:
                            st.success(f"{stld_type} wind loads successfully applied.", icon="✅")
                        else:
                            error_detail = res.json().get("message") or res.text
                            st.error(f"❌ Failed to apply {stld_type} loads.\n\n**Status Code:** {res.status_code}\n**Details:** {error_detail}")
                            error_flag = True

                    except Exception as e:
                        st.error(f"❗ Unexpected error occurred while sending API request.\n\n**Exception:** {e}")
                    
                if not error_flag :
                    time.sleep(1)
                    st.rerun()
                    

def apply_wind_data(df, wind_direction) :
    return [
        {
            "STORY_NAME": row["STORY"],
            "ELIV": float(row["ELEV [m]"]),
            "LOAD_H": float(row["Loaded_H [m]"]),
            "LOAD_BX": float(row["Loaded_Bx [m]"]),
            "LOAD_BY": float(row["Loaded_By [m]"]),
            "PRESS_X": row["p [kPa]"] if wind_direction in [0, 180] else 0,
            "PRESS_Y": row["p [kPa]"] if wind_direction in [90, 270] else 0
        }
        for _, row in df.iterrows()
    ]

def dataframe_to_excel(ws, df, building_angle, start_row=2, start_col=1):
    # 1. start_row - 1 행의 서식을 저장
    template_styles = []
    for col_idx in range(start_col, start_col + len(df.columns)):
        cell = ws.cell(row=start_row - 1, column=col_idx)
        style = {
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format
        }
        template_styles.append(style)

    # 2. DataFrame 내용을 시트에 쓰기 + 서식 적용
    for i, (_, row) in enumerate(df.iterrows()):
        current_row = start_row + i
        for j, value in enumerate(row):
            cell = ws.cell(row=current_row, column=start_col + j)
            cell.value = value
            
            # 서식 복사
            style = template_styles[j]
            cell.font = style['font']
            cell.fill = style['fill']
            cell.border = style['border']
            cell.alignment = style['alignment']
            cell.number_format = style['number_format']
        
        # 행 높이를 20으로 고정
        ws.row_dimensions[current_row].height = int(20 * zoom_ratio)

    # 3. 서식 복사 후 start_row - 1 행 삭제
    ws.delete_rows(start_row - 1)

# =======================================================================================================
# ============================================== Streamlit ==============================================
# =======================================================================================================

# ============================================== Side Bar ==============================================
# st.set_page_config(layout="wide")

st.session_state["api_connected"] = None
print(st.session_state["api_connected"])
print(base_url)
print(_mapi_key)

if not st.session_state["api_connected"] and base_url and _mapi_key :
        try :
            header = {"MAPI-Key":_mapi_key}
            res = requests.get(f"{base_url}/db/unit", headers=header)

            if res.status_code == 200:
                st.session_state["api_connected"] = True

            else :
                st.session_state["api_connected"] = False
                st.toast(f"❌ Failed to connect. Status code: {res.status_code}")

        except Exception as e:
            st.session_state["api_connected"] = False
            st.toast(f"❌ Connection failed: {e}")

# ============================================== Body ==============================================
if st.session_state["api_connected"] == True :
    story_data_body = {
        "Argument": {
            "FORCE_UNIT": "KN",
            "LENGTH_UNIT": "M"
        }
    }
    get_story_data = get_stor_data(body = story_data_body)
    stld_data = get_stld_data()

    if st.session_state["stor_valid_flag"] == True :
        st.session_state["story_data"] = get_story_data
    if st.session_state["stld_valid_flag"] == True :
        st.session_state["stld_data"] = stld_data

    # Calculate wind load 버튼 눌렸는지 확인하기 위한 flag 추가
    if st.session_state.get("wind_load_calculated") is None :
        st.session_state["wind_load_calculated"] = False

    # 방향 별 Parameters 결과 테이블 작성
    if st.session_state.get("parameters_merge_dict_all_NW") is None :
        for selected_cardinal_direction in au_cardinal_directions :
            parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
            if parameters_table_name not in st.session_state:
                # Parameters 계산 결과
                parameters_table_temp = {
                    "Story" : st.session_state["story_data"]["story_name"],
                    "VR" : [],
                    "Mc" : [],
                    "Md" : [],
                    "Mz,cat" : [],
                    "Ms" : [],
                    "Mt(ww)" : [],
                    "Mt(lw)" : []
                }
                st.session_state[parameters_table_name] = parameters_table_temp

    story_count = len(st.session_state["story_data"]["story_name"])

    c1, c2 = st.columns([1, 1.5])
    # 왼쪽 화면 설정
    with c1 : 
        with st.container(border=False, height=int(848 * zoom_ratio)) :
            # Parameters for wind pressure
            st.info("Parameters for wind actions")
            with st.container(border=False) :
                # Input VR
                param_col1, param_col2 = st.columns(2, vertical_alignment ="center")
                with param_col1 :
                    st.markdown(
												"""
												<span>V<sub>R</sub>(m/s)</span>
												<a href="https://tools.midasuser.com/en/fundamental-basic-wind-velocity-map" target="_blank" style="text-decoration: none; margin-left: 12px;">[🗺️Map]</a>
                        """,
                        unsafe_allow_html=True
                    )
                with param_col2 :
                    st.number_input("VR", min_value=0.0, max_value=100.0, step=0.1, value=39.00, key="vr", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": False}))
                for selected_cardinal_direction in au_cardinal_directions :
                    parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
                    st.session_state[parameters_table_name]["VR"] = [st.session_state["vr"]] * story_count

                # Input Wind Region
                param_col3, param_col4 = st.columns(2, vertical_alignment ="center")
                with param_col3 :
                    st.markdown("Wind Region", unsafe_allow_html=True)
                with param_col4 :
                    wind_region = st.selectbox("Wind Region", au_wind_region["Region"], index=5, key="wind_region", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": False}))

                # Input θ
                param_col5, param_col6 = st.columns(2, vertical_alignment ="center")
                with param_col5 :
                    st.markdown("Orientation Angle", unsafe_allow_html=True)
                with param_col6 :
                    orientation_angle = st.number_input("orientation angle", min_value=-360.0, max_value=360.0 ,step=0.1, value=0.0, key="orientation_angle", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": False}))

            st.info("Parameters for site exposure")
            with st.container(border=False) :
                # Input Type (worst case / each direction)
                popver_direction_select = st.popover(label="Wind Load Input Condition", use_container_width=True)
                input_condition = popver_direction_select.radio(label="",options=["Input Only Worst-Case Value", "Input for Each Direction"], key="input_condition", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": False}))

                if st.session_state.get("mfactor_display") is None:
                    st.session_state["mfactor_display"] = True

                if st.session_state["input_condition"] == "Input Only Worst-Case Value" :
                    st.session_state["mfactor_display"] = False
                else :
                    st.session_state["mfactor_display"] = True
                mfactor_display = st.session_state["mfactor_display"]

                # Calculate mc
                mc_index = au_mc["Region"].index(wind_region)
                st.session_state["mc"] = au_mc["Mc"][mc_index]
                if input_condition == "Input Only Worst-Case Value" :
                    for selected_cardinal_direction in au_cardinal_directions :
                        parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
                        st.session_state[parameters_table_name]["Mc"] = [st.session_state["mc"]] * story_count
                else :
                    for selected_cardinal_direction in au_cardinal_directions :
                        parameters_table_name = f"parameters_merge_dict_each_{selected_cardinal_direction}"
                        st.session_state[parameters_table_name]["Mc"] = [st.session_state["mc"]] * story_count

                # Calculate Md
                st.session_state["md"] = au_md[wind_region]
                if input_condition == "Input Only Worst-Case Value" :
                    for index, selected_cardinal_direction in enumerate(au_cardinal_directions) :
                        parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
                        st.session_state[parameters_table_name]["Md"] = [st.session_state["md"][index]] * story_count
                else :
                    for index, selected_cardinal_direction in enumerate(au_cardinal_directions) :
                        parameters_table_name = f"parameters_merge_dict_each_{selected_cardinal_direction}"
                        st.session_state[parameters_table_name]["Md"] = [st.session_state["md"][index]] * story_count

                # Calculate Mz,cat
                if "show_modal_mzcat" not in st.session_state:
                    st.session_state["show_modal_mzcat"] = False

                param_col7, param_col8 = st.columns(2, vertical_alignment ="center")
                def mzcat_on_change():
                    st.session_state["wind_load_calculated"] = False
                    st.session_state["show_modal_mzcat"] = True
                with param_col7 :
                    st.markdown("M<sub>z,cat</sub>", unsafe_allow_html=True)
                with param_col8 :
                    mzcat_input_type = st.selectbox("Terrain Category", au_terrain_category["Category"], index=3, key="terrain_cat_drop_single" ,label_visibility="collapsed", disabled=mfactor_display, on_change=mzcat_on_change)

                # Worst Case로 8 방향 풍하중 계산하는 경우
                if input_condition == "Input Only Worst-Case Value" :
                    if mzcat_input_type == au_terrain_category["Category"][5] :
                        # 새롭게 Select Multiple 선택한 경우
                        if st.session_state["show_modal_mzcat"] == True :
                            # 데이터 입력 전 오류 방지
                            if st.session_state.get("mzcat_result_all") is None:
                                mzcat_result = au_mzcat(st.session_state["wind_region"], [au_terrain_category["Category"][3]])
                            # 처음으로 Select Multiple로 입력한 경우
                            else :
                                mzcat_result = st.session_state["mzcat_result_all"]
                            au_mzcat_modal("all")
                            st.session_state["show_modal_mzcat"] = False
                            
                        # Select Multiple 이 미리 선택되어 유지돼 있는 경우
                        elif mzcat_input_type == au_terrain_category["Category"][5] and st.session_state["show_modal_mzcat"] == False :
                            if wind_region == "A0" :
                                au_mzcat(wind_region)
                            if st.session_state.get("mzcat_result_all") is None:
                                st.error("Please input terrain category data and click [Calculate] button.")
                                st.stop()
                            else :
                                mzcat_result = au_mzcat(
                                    st.session_state["wind_region"],
                                    st.session_state["terrain_table_data_all"]["Terrain Category"],
                                    st.session_state["terrain_table_data_all"]["Terrain Length(m)"]
                                )
                    else :
                        mzcat_result = au_mzcat(st.session_state["wind_region"], [st.session_state["terrain_cat_drop_single"]])
                    
                    for selected_cardinal_direction in au_cardinal_directions :
                        parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
                        st.session_state[parameters_table_name]["Mz,cat"] = mzcat_result["Mz,cat"]

                # 8방향 각각 풍하중 계산하는 경우
                else :
                    # col2의 값이 렌더링 되기 전 예외처리
                    if st.session_state.get("modified_mzcat_data") is None or st.session_state["modified_mzcat_data"] == {} :
                        mzcat_result = au_mzcat(st.session_state["wind_region"], [au_terrain_category["Category"][3]])
                        for selected_cardinal_direction in au_cardinal_directions :
                            parameters_table_name = f"parameters_merge_dict_each_{selected_cardinal_direction}"
                            st.session_state[parameters_table_name]["Mz,cat"] = mzcat_result["Mz,cat"]
                    else :
                        modified_terrain_category_index = list(st.session_state["modified_mzcat_data"].keys())[0]
                        modified_terrain_category_value = list(st.session_state["modified_mzcat_data"].values())[0]["Mzcat"]
                        modified_terrain_category_cardinal_direction = au_cardinal_directions[modified_terrain_category_index]
          
                        if modified_terrain_category_value == au_terrain_category["Category"][5] :
                            # 새롭게 Select Multiple 선택한 경우
                            if st.session_state["show_modal_mzcat"] == True :
                                # 데이터 입력 전 오류 방지
                                if st.session_state.get(f"mzcat_result_{modified_terrain_category_cardinal_direction}") is None:
                                    mzcat_result = au_mzcat(st.session_state["wind_region"], [au_terrain_category["Category"][3]])
                                # 처음으로 Select Multiple로 입력한 경우
                                else :
                                    mzcat_result = st.session_state[f"mzcat_result_{modified_terrain_category_cardinal_direction}"]
                                au_mzcat_modal(modified_terrain_category_cardinal_direction)
                                st.session_state["show_modal_mzcat"] = False

                            # Select Multiple 이 미리 선택되어 유지돼 있는 경우
                            elif st.session_state["show_modal_mzcat"] == False :
                                if wind_region == "A0" :
                                    au_mzcat(wind_region)
                                if st.session_state.get(f"mzcat_result_{modified_terrain_category_cardinal_direction}") is None:
                                    st.error("Please input terrain category data and click [Calculate] button.")
                                    st.stop()
                                else :
                                    mzcat_result = au_mzcat(
                                        st.session_state["wind_region"],
                                        st.session_state[f"terrain_table_data_{modified_terrain_category_cardinal_direction}"]["Terrain Category"],
                                        st.session_state[f"terrain_table_data_{modified_terrain_category_cardinal_direction}"]["Terrain Length(m)"]
                                    )
                        else :
                            mzcat_result = au_mzcat(st.session_state["wind_region"], [modified_terrain_category_value])

                        parameters_table_name = f"parameters_merge_dict_each_{modified_terrain_category_cardinal_direction}"
                        st.session_state[parameters_table_name]["Mz,cat"] = mzcat_result["Mz,cat"]
                    
                # Parameters for Ms
                if "show_modal_ms" not in st.session_state:
                    st.session_state["show_modal_ms"] = False
                if "calc_ms_button" not in st.session_state :
                    st.session_state["calc_ms_button"] = False
                    
                param_col9, param_col10 = st.columns(2, vertical_alignment ="center")
                def ms_on_change():
                    st.session_state["wind_load_calculated"] = False
                    st.session_state["show_modal_ms"] = True
                with param_col9 :
                    st.markdown("M<sub>s</sub>", unsafe_allow_html=True)
                with param_col10 :
                    shileding_category = st.selectbox("Shielding Category", au_shielding_category["shielding_catetory"], key="shielding_catetory", label_visibility="collapsed", disabled=mfactor_display, on_change=ms_on_change)

                height = st.session_state["story_data"]["story_elev"][0]

                # Worst Case로 8 방향 풍하중 계산하는 경우
                if input_condition == "Input Only Worst-Case Value" :
                    if shileding_category == "Shielding" :
                        # 새롭게 shielding parameters 입력하는 경우
                        if st.session_state["show_modal_ms"] == True :
                            if st.session_state.get("ms_result_all") is None:
                                ms_result = au_ms(height, 100, 10, 5)
                            else :
                                ms_result = st.session_state["ms_result_all"]
                            au_ms_modal(height, "all")
                            st.session_state["show_modal_ms"] = False
                        # 이미 입력하고 유지하고 있는 경우
                        elif st.session_state["show_modal_ms"] == False :
                            if st.session_state.get("ms_result_all") is None:
                                st.error("Please input shielding data and click [Calculate] button.")
                                st.stop()
                            else :
                                ms_result = st.session_state["ms_result_all"]
                    else :
                        ms_result = {"Ms" : 1.0}

                    for selected_cardinal_direction in au_cardinal_directions :
                        parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
                        st.session_state[parameters_table_name]["Ms"] = [ms_result["Ms"]] * story_count

                # 8방향 각각 풍하중 계산하는 경우
                else :
                    # col2의 값이 렌더링 되기 전 예외처리
                    if st.session_state.get("modified_ms_data") is None or st.session_state["modified_ms_data"] == {} :
                        ms_result = {"Ms" : 1.0}
                        for selected_cardinal_direction in au_cardinal_directions :
                            parameters_table_name = f"parameters_merge_dict_each_{selected_cardinal_direction}"
                            st.session_state[parameters_table_name]["Ms"] = [ms_result["Ms"]] * story_count
                    else :
                        modified_ms_index = list(st.session_state["modified_ms_data"].keys())[0]
                        modified_ms_value = list(st.session_state["modified_ms_data"].values())[0]["Ms"]
                        modified_ms_cardinal_direction = au_cardinal_directions[modified_ms_index]

                        if modified_ms_value == "Shielding" :
                            # 새롭게 shielding parameters 입력하는 경우
                            if st.session_state["show_modal_ms"] == True :
                                if st.session_state.get(f"ms_result_{modified_ms_cardinal_direction}") is None:
                                    ms_result = au_ms(height, 100, 10, 5)
                                else :
                                    ms_result = st.session_state[f"ms_result_{modified_ms_cardinal_direction}"]
                                au_ms_modal(height, modified_ms_cardinal_direction)
                                st.session_state["show_modal_ms"] = False
                                
                            # 이미 입력하고 유지하고 있는 경우
                            elif st.session_state["show_modal_ms"] == False :
                                if st.session_state.get(f"ms_result_{modified_ms_cardinal_direction}") is None:
                                    st.error("Please input shielding data and click [Calculate] button.")
                                    st.stop()
                                else :
                                    ms_result = st.session_state[f"ms_result_{modified_ms_cardinal_direction}"]
                        else :
                            ms_result = {"Ms" : 1.0}
                        
                        parameters_table_name = f"parameters_merge_dict_each_{modified_ms_cardinal_direction}"
                        st.session_state[parameters_table_name]["Ms"] = [ms_result["Ms"]] * story_count

                # Parameters for Mt
                if "show_modal_mt" not in st.session_state:
                    st.session_state["show_modal_mt"] = False

                param_col11, param_col12 = st.columns(2, vertical_alignment ="center")
                def mt_on_change():
                    st.session_state["wind_load_calculated"] = False
                    st.session_state["show_modal_mt"] = True
                with param_col11 :
                    st.markdown("M<sub>t</sub>", unsafe_allow_html=True)
                with param_col12 :
                    hill_type = st.selectbox("Hill Type", au_hill_type["hill_type"], key="hill_type", label_visibility="collapsed", disabled=mfactor_display, on_change=mt_on_change)

                # Worst Case로 8 방향 풍하중 계산하는 경우
                if input_condition == "Input Only Worst-Case Value" :
                    if hill_type != au_hill_type["hill_type"][0] :
                        # 새롭게 Mt 값 선택한 경우
                        if st.session_state["show_modal_mt"] == True :
                            mt_result = au_mt(height, au_hill_type["hill_type"][0])
                            au_mt_modal(hill_type, "all")
                            st.session_state["show_modal_mt"] = False
                        # Mt 옵션이 미리 선택된 경우
                        else :
                            if st.session_state.get("mt_result_all") is None:
                                st.error("Please input hill data and click [Calculate] button.")
                                st.stop()
                            else :
                                mt_result = st.session_state["mt_result_all"]                               
                    else :
                        mt_result = au_mt(height, hill_type)
                    for selected_cardinal_direction in au_cardinal_directions :
                        parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
                        st.session_state[parameters_table_name]["Mt(ww)"] = mt_result["Mt(ww)"]
                        st.session_state[parameters_table_name]["Mt(lw)"] = mt_result["Mt(lw)"]

                # 8방향 각각 풍하중 계산하는 경우
                else :
                    # col2의 값이 렌더링 되기 전 예외처리
                    if st.session_state.get("modified_mt_data") is None or st.session_state["modified_mt_data"] == {} :
                        mt_result = au_mt(height, au_hill_type["hill_type"][0])
                        for selected_cardinal_direction in au_cardinal_directions :
                            parameters_table_name = f"parameters_merge_dict_each_{selected_cardinal_direction}"
                            st.session_state[parameters_table_name]["Mt(ww)"] = mt_result["Mt(ww)"]
                            st.session_state[parameters_table_name]["Mt(lw)"] = mt_result["Mt(lw)"]
                    else :
                        modified_mt_index = list(st.session_state["modified_mt_data"].keys())[0]
                        modified_mt_value = list(st.session_state["modified_mt_data"].values())[0]["Mt"]
                        modified_mt_cardinal_direction = au_cardinal_directions[modified_mt_index]

                        if modified_mt_value != au_hill_type["hill_type"][0] :
                            # 새롭게 Mt parameters 입력하는 경우
                            if st.session_state["show_modal_mt"] == True :
                                if st.session_state.get(f"mt_result_{modified_mt_cardinal_direction}") is None:
                                    mt_result = au_mt(height, au_hill_type["hill_type"][0])
                                else :
                                    mt_result = st.session_state[f"mt_result_{modified_mt_cardinal_direction}"]
                                au_mt_modal(modified_mt_value, modified_mt_cardinal_direction)
                                st.session_state["show_modal_mt"] = False

                            # 이미 입력하고 유지하고 있는 경우
                            elif st.session_state["show_modal_mt"] == False :
                                if st.session_state.get(f"mt_result_{modified_mt_cardinal_direction}") is None:
                                    st.error("Please input hill data and click [Calculate] button.")
                                    st.stop()

                                else :
                                    mt_result = st.session_state[f"mt_result_{modified_mt_cardinal_direction}"]
                        else :
                            mt_result = au_mt(height, modified_mt_value)

                        parameters_table_name = f"parameters_merge_dict_each_{modified_mt_cardinal_direction}"
                        st.session_state[parameters_table_name]["Mt(ww)"] = mt_result["Mt(ww)"]
                        st.session_state[parameters_table_name]["Mt(lw)"] = mt_result["Mt(lw)"]

                # Worst Case Parameters로 작성된 st.session_state[f"parameters_merge_dict_all_{selected_cardinal_direction}"]을
                # 8방향 각각 Parameters로 작성된 st.session_state[f"parameters_merge_dict_each_{selected_cardinal_direction}"]으로 변경
                if st.session_state.get("parameters_merge_dict_each_NW") is None :
                    for selected_cardinal_direction in au_cardinal_directions :
                        parameters_table_name = f"parameters_merge_dict_all_{selected_cardinal_direction}"
                        parameters_table_name_new = f"parameters_merge_dict_each_{selected_cardinal_direction}"
                        if parameters_table_name_new not in st.session_state :
                            st.session_state[parameters_table_name_new] = st.session_state[parameters_table_name].copy()
                
                # ==============================================================================================================================================================
                # ==============================================================================================================================================================
                # Calculate Vsit,β (8방향) /  
                if "vsit_data(ww)_all" not in st.session_state:
                    st.session_state["vsit_data(ww)_all"] = {}
                    st.session_state["vsit_data(lw)_all"] = {}
                    st.session_state["vsit_data(ww)_all"]["Story"] = st.session_state["story_data"]["story_name"]
                    st.session_state["vsit_data(lw)_all"]["Story"] = st.session_state["story_data"]["story_name"]
                if "vsit_data(ww)_each" not in st.session_state:
                    st.session_state["vsit_data(ww)_each"] = {}
                    st.session_state["vsit_data(lw)_each"] = {}
                    st.session_state["vsit_data(ww)_each"]["Story"] = st.session_state["story_data"]["story_name"]
                    st.session_state["vsit_data(lw)_each"]["Story"] = st.session_state["story_data"]["story_name"]
                
                for selected_cardinal_direction in au_cardinal_directions :
                    df_temp_all = pd.DataFrame(st.session_state[f"parameters_merge_dict_all_{selected_cardinal_direction}"])
                    df_temp_each = pd.DataFrame(st.session_state[f"parameters_merge_dict_each_{selected_cardinal_direction}"])
                    columns_to_multiply_ww = ["VR", "Mc", "Md", "Mz,cat", "Ms", "Mt(ww)"]
                    columns_to_multiply_lw = ["VR", "Mc", "Md", "Mz,cat", "Ms", "Mt(lw)"]
                    st.session_state[f"parameters_merge_dict_all_{selected_cardinal_direction}"]["Vsit,β(ww)"] = list(df_temp_all[columns_to_multiply_ww].product(axis=1))
                    st.session_state[f"parameters_merge_dict_all_{selected_cardinal_direction}"]["Vsit,β(lw)"] = list(df_temp_all[columns_to_multiply_lw].product(axis=1))
                    st.session_state["vsit_data(ww)_all"][f"{selected_cardinal_direction}"] = st.session_state[f"parameters_merge_dict_all_{selected_cardinal_direction}"]["Vsit,β(ww)"]
                    st.session_state["vsit_data(lw)_all"][f"{selected_cardinal_direction}"] = st.session_state[f"parameters_merge_dict_all_{selected_cardinal_direction}"]["Vsit,β(lw)"]

                    st.session_state[f"parameters_merge_dict_each_{selected_cardinal_direction}"]["Vsit,β(ww)"] = list(df_temp_each[columns_to_multiply_ww].product(axis=1))
                    st.session_state[f"parameters_merge_dict_each_{selected_cardinal_direction}"]["Vsit,β(lw)"] = list(df_temp_each[columns_to_multiply_lw].product(axis=1))
                    st.session_state["vsit_data(ww)_each"][f"{selected_cardinal_direction}"] = st.session_state[f"parameters_merge_dict_each_{selected_cardinal_direction}"]["Vsit,β(ww)"]
                    st.session_state["vsit_data(lw)_each"][f"{selected_cardinal_direction}"] = st.session_state[f"parameters_merge_dict_each_{selected_cardinal_direction}"]["Vsit,β(lw)"]

                df_vsit_ww_all = pd.DataFrame(st.session_state["vsit_data(ww)_all"])
                df_vsit_lw_all = pd.DataFrame(st.session_state["vsit_data(lw)_all"])
                df_vsit_ww_each = pd.DataFrame(st.session_state["vsit_data(ww)_each"])
                df_vsit_lw_each = pd.DataFrame(st.session_state["vsit_data(lw)_each"])

                # ==============================================================================================================================================================
                # ==============================================================================================================================================================
                # Calculate Vdes,θ (4방향 : 건축물에 적용될 4방향에 대한 데이터만 추출)
                st.session_state["building_direction"] = [beta + orientation_angle for beta in au_building_angle]

                if "vdes_data(ww)_all" not in st.session_state:
                    st.session_state["vdes_data(ww)_all"] = {}
                    st.session_state["vdes_data(lw)_all"] = {}
                    st.session_state["vdes_data(ww)_all"]["Story"] = st.session_state["story_data"]["story_name"]
                    st.session_state["vdes_data(lw)_all"]["Story"] = st.session_state["story_data"]["story_name"]
                if "vdes_data(ww)_each" not in st.session_state:
                    st.session_state["vdes_data(ww)_each"] = {}
                    st.session_state["vdes_data(lw)_each"] = {}
                    st.session_state["vdes_data(ww)_each"]["Story"] = st.session_state["story_data"]["story_name"]
                    st.session_state["vdes_data(lw)_each"]["Story"] = st.session_state["story_data"]["story_name"]

                for idx, building_angle in enumerate(au_building_angle) :
                    angle = st.session_state["building_direction"][idx]
                    vdes_ww_temp_all_value = vdes_extract(df_vsit_ww_all, angle, orientation_angle)
                    vdes_lw_temp_all_value = vdes_extract(df_vsit_lw_all, angle, orientation_angle)
                    vdes_ww_temp_each_value = vdes_extract(df_vsit_ww_each, angle, orientation_angle)
                    vdes_lw_temp_each_value = vdes_extract(df_vsit_lw_each, angle, orientation_angle)
                    
                    st.session_state[f"vdes_data(ww)_all"][f"{building_angle}°"] = vdes_ww_temp_all_value
                    st.session_state[f"vdes_data(lw)_all"][f"{building_angle}°"] = vdes_lw_temp_all_value
                    st.session_state[f"vdes_data(ww)_each"][f"{building_angle}°"] = vdes_ww_temp_each_value
                    st.session_state[f"vdes_data(lw)_each"][f"{building_angle}°"] = vdes_lw_temp_each_value

                df_vdes_ww_all = pd.DataFrame(st.session_state["vdes_data(ww)_all"])
                df_vdes_lw_all = pd.DataFrame(st.session_state["vdes_data(lw)_all"])
                df_vdes_ww_each = pd.DataFrame(st.session_state["vdes_data(ww)_each"])
                df_vdes_lw_each = pd.DataFrame(st.session_state["vdes_data(lw)_each"])

            st.info("Parameters for wind pressure(p)")
            with st.container(border=False) :
                param_col13, param_col14 = st.columns(2, vertical_alignment ="center")
                with param_col13 :
                    st.markdown("C<sub>pe</sub>(windward)", unsafe_allow_html=True)
                with param_col14 :
                    cpe_windward_options = [f"{val:.2f}" for val in au_cpe_ww["Cpe_windward"]]
                    input_cpe = float(st.selectbox("Cpe_windward", cpe_windward_options, key="cpe_windward", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": False})))
                
                # Calculate Cpe_leeward
                param_col15, param_col16 = st.columns(2, vertical_alignment ="center")
                with param_col15 :
                    st.markdown("C<sub>pe_x</sub>(leeward)", unsafe_allow_html=True)
                with param_col16 :
                    input_cpe_lw_x = st.number_input("Cpe_leeward", value=au_cpe_leeward(0)["Cpe_lw"][0], key="cpe_leeward_x", label_visibility="collapsed", disabled=True, on_change=lambda: st.session_state.update({"wind_load_calculated": False}))

                param_col17, param_col18 = st.columns(2, vertical_alignment ="center")
                with param_col17 :
                    st.markdown("C<sub>pe_y</sub>(leeward)", unsafe_allow_html=True)
                with param_col18 :
                    input_cpe_lw_y = st.number_input("Cpe_leeward", value=au_cpe_leeward(90)["Cpe_lw"][0], key="cpe_leeward_y", label_visibility="collapsed", disabled=True, on_change=lambda: st.session_state.update({"wind_load_calculated": False}))

                # Input Cdyn
                param_col19, param_col20 = st.columns([1, 1], vertical_alignment ="center")
                with param_col19 :
                    st.markdown("C<sub>dyn</sub>", unsafe_allow_html=True)
                with param_col20 :
                    cdyn = st.number_input("Cdyn", 1.0, step=0.1, key="cdyn", label_visibility="collapsed", disabled=mfactor_display, on_change=lambda: st.session_state.update({"wind_load_calculated": False}))
                
    if input_condition == "Input Only Worst-Case Value" :
        with c2 :
            with st.container(border=False, height=int(848 * zoom_ratio)) :
                if st.session_state.get("wind_load_calculated", True) :
                    st.info("Result Graphs and Tables")
                    if st.session_state["stor_valid_flag"] == True :
                        for angle in au_building_angle :
                            wind_calc_all_name = f"wind_calc_all_df_{angle}"

                            wind_calc_all_df = pd.DataFrame({})
                            wind_calc_all_df["STORY"] = st.session_state["story_data"]["story_name"]
                            wind_calc_all_df["ELEV [m]"] = st.session_state["story_data"]["story_elev"]
                            wind_calc_all_df["p [kPa]"] = st.session_state["wind_pressure_data_all"][f"{angle}°"]
                            wind_calc_all_df["Loaded_H [m]"] = st.session_state["story_data"]["loaded_h"]
                            wind_calc_all_df["Loaded_Bx [m]"] = st.session_state["story_data"]["loaded_bx"]
                            wind_calc_all_df["Loaded_By [m]"] = st.session_state["story_data"]["loaded_by"]

                            # Story Force 계산
                            if angle in [0, 180] :
                                wind_calc_all_df["Story Force [kN]"] = wind_calc_all_df["p [kPa]"] * wind_calc_all_df["Loaded_H [m]"].astype(float) * wind_calc_all_df["Loaded_Bx [m]"].astype(float)
                            else :
                                wind_calc_all_df["Story Force [kN]"] = wind_calc_all_df["p [kPa]"] * wind_calc_all_df["Loaded_H [m]"].astype(float) * wind_calc_all_df["Loaded_By [m]"].astype(float)
                            wind_calc_all_df["Story Force [kN]"].iloc[-1] = 0

                            # 제품과 동일한 로직으로 shear force 계산 (해당 층보다 위에 해당하는 층의 story force 합)
                            story_shear_all_list = []
                            for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                                if story_idx == 0 :
                                    story_shear_all_list.append(0)
                                else :
                                    story_shear_all_list.append(sum(wind_calc_all_df["Story Force [kN]"][0:story_idx]))
                            wind_calc_all_df["Story Shear [kN]"] = story_shear_all_list

                            # Overturning Moment 계산
                            story_overturning_moment_all_list = []
                            for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                                if story_idx == 0 :
                                    story_overturning_moment_all_list.append(0)
                                else :
                                    overturning_moment_all_temp_list = []
                                    for i in range(story_idx) :
                                        height = wind_calc_all_df["ELEV [m]"].astype(float).iloc[i] - wind_calc_all_df["ELEV [m]"].astype(float).iloc[story_idx]
                                        overturning_moment_all_temp_list.append(wind_calc_all_df["Story Force [kN]"].iloc[i] * height)
                                    story_overturning_moment_all_list.append(sum(overturning_moment_all_temp_list))
                            wind_calc_all_df["Overturning M [kN·m]"] = story_overturning_moment_all_list

                            st.session_state[wind_calc_all_name] = wind_calc_all_df
                        
                        # st.dataframe(st.session_state[wind_calc_all_name])
                        
                        tab_graph, tab_table = st.tabs(["Graph", "Table"])
                        with tab_table :
                            table_col1, table_col2 = st.columns(2, vertical_alignment="center")
                            with table_col1 :
                                st.write("Select Cardinal Direction")
                            with table_col2 :
                                au_building_angle_new = [f"{building_angle}°" for building_angle in au_building_angle]
                                table_cardinal_direction = st.selectbox("Graph Type", au_building_angle_new, key="table_cardinal_direction_all", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": True}))
                                table_cardinal_direction = table_cardinal_direction.replace("°", "")

                            st.dataframe(data=st.session_state[f"wind_calc_all_df_{table_cardinal_direction}"], height=int(630 * zoom_ratio), row_height=int(44 * zoom_ratio) ,hide_index=True, use_container_width=True ,column_config={
                                "ELEV [m]": st.column_config.NumberColumn("ELEV [m]", format="%.2f"),
                                "p [kPa]": st.column_config.NumberColumn("p [kPa]", format="%.2f"),
                                "Loaded_H [m]": st.column_config.NumberColumn("Loaded_H [m]", format="%.2f"),
                                "Loaded_Bx [m]": st.column_config.NumberColumn("Loaded_Bx [m]", format="%.2f"),
                                "Loaded_By [m]": st.column_config.NumberColumn("Loaded_By [m]", format="%.2f"),
                                "Story Force [kN]": st.column_config.NumberColumn("Story Force [kN]", format="%.2f"),
                                "Story Shear [kN]": st.column_config.NumberColumn("Story Shear [kN]", format="%.2f"),
                                "Overturning M [kN·m]": st.column_config.NumberColumn("Overturning M [kN·m]", format="%.2f"),
                            })
                            
                        with tab_graph :
                            graph_col1, graph_col2 = st.columns(2, vertical_alignment="center")
                            with graph_col1 :
                                st.write("Select Graph Type")
                            with graph_col2 :
                                graph_type = st.selectbox("Graph Type", ["Story Force", "Story Shear", "Overturning Moment"], key="graph_type_all", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": True}))
                            
                            story_force_all_df = pd.DataFrame({})
                            story_shear_all_df = pd.DataFrame({})
                            over_turning_moment_all_df = pd.DataFrame({})
                            story_force_all_df["STORY"] = st.session_state["story_data"]["story_name"]
                            story_shear_all_df["STORY"] = st.session_state["story_data"]["story_name"]
                            over_turning_moment_all_df["STORY"] = st.session_state["story_data"]["story_name"]

                            for angle in au_building_angle :
                                wind_calc_column_name = f"{angle}°"
                                story_force_all_df[f"Story Force[{angle}°]"] = st.session_state[f"wind_calc_all_df_{angle}"].copy()["Story Force [kN]"]
                                story_shear_all_df[f"Story Shear[{angle}°]"] = st.session_state[f"wind_calc_all_df_{angle}"].copy()["Story Shear [kN]"]
                                over_turning_moment_all_df[f"Story_Force[{angle}°]"] = st.session_state[f"wind_calc_all_df_{angle}"].copy()["Overturning M [kN·m]"]

                            # 최종 결과 그래프
                            story_order = list(story_force_all_df["STORY"])

                            story_force_all_df = story_force_all_df.iloc[:-1]
                            story_force_all_df.columns = ["Story", "0°", "90°", "180°", "270°"]
                            story_force_all_df["Story"] = pd.Categorical(story_force_all_df["Story"], categories=story_order, ordered=True)

                            story_shear_all_df.columns = ["Story", "0°", "90°", "180°", "270°"]
                            story_shear_all_df["Story"] = pd.Categorical(story_shear_all_df["Story"], categories=story_order, ordered=True)

                            over_turning_moment_all_df.columns = ["Story", "0°", "90°", "180°", "270°"]
                            over_turning_moment_all_df["Story"] = pd.Categorical(over_turning_moment_all_df["Story"], categories=story_order, ordered=True)
                            
                            if graph_type == "Story Force" :
                                # 그래프 그리기
                                chart = line_chart_hor_multi(story_force_all_df, "[kN]")
                                st.altair_chart(chart, use_container_width=True)
                            elif graph_type == "Story Shear" :
                                # 그래프 그리기
                                chart = line_chart_hor_multi(story_shear_all_df, "[kN]")
                                st.altair_chart(chart, use_container_width=True)
                            else :
                                # 그래프 그리기
                                chart = line_chart_hor_multi(over_turning_moment_all_df, "[kN·m]")
                                st.altair_chart(chart, use_container_width=True)
                    elif st.session_state["stor_valid_flag"] == None :
                        st.warning("Please check the story data.", icon="❗")
                else :
                    st.info("Graphs for Wind Velocity")
                    with st.container(height=int(800 * zoom_ratio), border=False) :
                        velocity_line_chart(df_vsit_ww_all, st.session_state["story_data"]["story_name"], au_cardinal_directions, 'Vsit,β by Story', height=int(360 * zoom_ratio))
                        velocity_line_chart(df_vdes_ww_all, st.session_state["story_data"]["story_name"], list(df_vdes_ww_all.columns[1:]), 'Vdes,θ by Story', height=int(360 * zoom_ratio))	

                        ## Mt의 카테고리가 Escarpment일 경우 Windward, Leeward에 따라서 Velocity 값이 바뀔 수 있음. 해당 데이터는 다운로드 제공하는 엑셀에 데이터로 제공. UI상에서는 windward로만 제공
                        # graph_col1, graph_col2 = st.columns(2, vertical_alignment="center")
                        # with graph_col1 :
                        #     st.write("Select Direction")
                        # with graph_col2 :
                        #     all_graph_direction = st.selectbox("Graph Direction", ["Windward", "Leeward"], key="all_graph_direction", label_visibility="collapsed")
                        
                        # if all_graph_direction == "Windward" :
                        #     velocity_line_chart(df_vsit_ww_all, st.session_state["story_data"]["story_name"], au_cardinal_directions, 'Vsit,β by Story', height=360 * zoom_ratio)
                        #     velocity_line_chart(df_vdes_ww_all, st.session_state["story_data"]["story_name"], list(df_vdes_ww_all.columns[1:]), 'Vdes,θ by Story', height=360 * zoom_ratio)
                        # else :
                        #     velocity_line_chart(df_vsit_lw_all, st.session_state["story_data"]["story_name"], au_cardinal_directions, 'Vsit,β by Story', height=360 * zoom_ratio)
                        #     velocity_line_chart(df_vdes_lw_all, st.session_state["story_data"]["story_name"], list(df_vdes_lw_all.columns[1:]), 'Vdes,θ by Story', height=360 * zoom_ratio)

    elif input_condition == "Input for Each Direction" :
        with c2 :
            with st.container(border=False, height=int(848 * zoom_ratio)) :
                if st.session_state.get("wind_load_calculated", True) :
                    st.info("Result Graphs and Tables")
                    if st.session_state["stor_valid_flag"] == True :
                        for building_angle in au_building_angle :
                            wind_calc_each_name = f"wind_calc_each_df_{building_angle}"
                            wind_calc_each_df = pd.DataFrame({})
                            wind_calc_each_df["STORY"] = st.session_state["story_data"]["story_name"]
                            wind_calc_each_df["ELEV [m]"] = st.session_state["story_data"]["story_elev"]
                            wind_calc_each_df["p [kPa]"] = st.session_state["wind_pressure_data_each"][f"{building_angle}°"]
                            wind_calc_each_df["Loaded_H [m]"] = st.session_state["story_data"]["loaded_h"]
                            wind_calc_each_df["Loaded_Bx [m]"] = st.session_state["story_data"]["loaded_bx"]
                            wind_calc_each_df["Loaded_By [m]"] = st.session_state["story_data"]["loaded_by"]
                            
                            # Story Force 계산
                            if building_angle in [0, 180] :
                                wind_calc_each_df["Story Force [kN]"] = wind_calc_each_df["p [kPa]"] * wind_calc_each_df["Loaded_H [m]"].astype(float) * wind_calc_each_df["Loaded_Bx [m]"].astype(float)
                            else :
                                wind_calc_each_df["Story Force [kN]"] = wind_calc_each_df["p [kPa]"] * wind_calc_each_df["Loaded_H [m]"].astype(float) * wind_calc_each_df["Loaded_By [m]"].astype(float)
                            wind_calc_each_df["Story Force [kN]"].iloc[-1] = 0

                            # 제품과 동일한 로직으로 shear force 계산 (해당 층보다 위에 해당하는 층의 story force 합)
                            story_shear_each_list = []
                            for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                                if story_idx == 0 :
                                    story_shear_each_list.append(0)
                                else :
                                    story_shear_each_list.append(sum(wind_calc_each_df["Story Force [kN]"][0:story_idx]))
                            wind_calc_each_df["Story Shear [kN]"] = story_shear_each_list

                            # Overturning Moment 계산
                            story_overturning_moment_each_list = []
                            for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                                if story_idx == 0 :
                                    story_overturning_moment_each_list.append(0)
                                else :
                                    overturning_moment_each_temp_list = []
                                    for i in range(story_idx) :
                                        height = wind_calc_each_df["ELEV [m]"].astype(float).iloc[i] - wind_calc_each_df["ELEV [m]"].astype(float).iloc[story_idx]
                                        overturning_moment_each_temp_list.append(wind_calc_each_df["Story Force [kN]"].iloc[i] * height)
                                    story_overturning_moment_each_list.append(sum(overturning_moment_each_temp_list))
                            wind_calc_each_df["Overturning M [kN·m]"] = story_overturning_moment_each_list

                            st.session_state[wind_calc_each_name] = wind_calc_each_df
                        
                        # st.dataframe(st.session_state[wind_calc_each_name])
                        
                        tab_graph, tab_table = st.tabs(["Graph", "Table"])
                        with tab_table :
                            table_col1, table_col2 = st.columns(2, vertical_alignment="center")
                            with table_col1 :
                                st.write("Select Cardinal Direction")
                            with table_col2 :
                                au_building_angle_new = [f"{building_angle}°" for building_angle in au_building_angle]
                                table_cardinal_direction = st.selectbox("Graph Type", au_building_angle_new, key="table_cardinal_direction_each", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": True}))
                                table_cardinal_direction = table_cardinal_direction.replace("°", "")

                            st.dataframe(data=st.session_state[f"wind_calc_each_df_{table_cardinal_direction}"], height=int(630 * zoom_ratio), row_height=int(44 * zoom_ratio) ,hide_index=True, use_container_width=True ,column_config={
                                "ELEV [m]": st.column_config.NumberColumn("ELEV [m]", format="%.2f"),
                                "p [kPa]": st.column_config.NumberColumn("p [kPa]", format="%.2f"),
                                "Loaded_H [m]": st.column_config.NumberColumn("Loaded_H [m]", format="%.2f"),
                                "Loaded_Bx [m]": st.column_config.NumberColumn("Loaded_Bx [m]", format="%.2f"),
                                "Loaded_By [m]": st.column_config.NumberColumn("Loaded_By [m]", format="%.2f"),
                                "Story Force [kN]": st.column_config.NumberColumn("Story Force [kN]", format="%.2f"),
                                "Story Shear [kN]": st.column_config.NumberColumn("Story Shear [kN]", format="%.2f"),
                                "Overturning M [kN·m]": st.column_config.NumberColumn("Overturning M [kN·m]", format="%.2f"),
                            })
                            
                        with tab_graph :
                            graph_col1, graph_col2 = st.columns(2, vertical_alignment="center")
                            with graph_col1 :
                                st.write("Select Graph Type")
                            with graph_col2 :
                                graph_type = st.selectbox("Graph Type", ["Story Force", "Story Shear", "Overturning Moment"], key="graph_type_each", label_visibility="collapsed", on_change=lambda: st.session_state.update({"wind_load_calculated": True}))
                            
                            story_force_each_df = pd.DataFrame({})
                            story_shear_each_df = pd.DataFrame({})
                            over_turning_moment_each_df = pd.DataFrame({})
                            story_force_each_df["STORY"] = st.session_state["story_data"]["story_name"]
                            story_shear_each_df["STORY"] = st.session_state["story_data"]["story_name"]
                            over_turning_moment_each_df["STORY"] = st.session_state["story_data"]["story_name"]

                            for building_angle in au_building_angle :
                                wind_calc_column_name = f"{building_angle}°"
                                story_force_each_df[f"Story Force[{building_angle}°]"] = st.session_state[f"wind_calc_each_df_{building_angle}"].copy()["Story Force [kN]"]
                                story_shear_each_df[f"Story Shear[{building_angle}°]"] = st.session_state[f"wind_calc_each_df_{building_angle}"].copy()["Story Shear [kN]"]
                                over_turning_moment_each_df[f"Story_Force[{building_angle}°]"] = st.session_state[f"wind_calc_each_df_{building_angle}"].copy()["Overturning M [kN·m]"]

                            # 최종 결과 그래프
                            story_order = list(story_force_each_df["STORY"])

                            story_force_each_df = story_force_each_df.iloc[:-1]
                            story_force_each_df.columns = ["Story", "0°", "90°", "180°", "270°"]
                            story_force_each_df["Story"] = pd.Categorical(story_force_each_df["Story"], categories=story_order, ordered=True)

                            story_shear_each_df.columns = ["Story", "0°", "90°", "180°", "270°"]
                            story_shear_each_df["Story"] = pd.Categorical(story_shear_each_df["Story"], categories=story_order, ordered=True)

                            over_turning_moment_each_df.columns = ["Story", "0°", "90°", "180°", "270°"]
                            over_turning_moment_each_df["Story"] = pd.Categorical(over_turning_moment_each_df["Story"], categories=story_order, ordered=True)
                            
                            if graph_type == "Story Force" :
                                # 그래프 그리기
                                chart = line_chart_hor_multi(story_force_each_df, "[kN]")
                                st.altair_chart(chart, use_container_width=True)
                            elif graph_type == "Story Shear" :
                                # 그래프 그리기
                                chart = line_chart_hor_multi(story_shear_each_df, "[kN]")
                                st.altair_chart(chart, use_container_width=True)
                            else :
                                # 그래프 그리기
                                chart = line_chart_hor_multi(over_turning_moment_each_df, "[kN·m]")
                                st.altair_chart(chart, use_container_width=True)
                    elif st.session_state["stor_valid_flag"] == None :
                        st.warning("Please check the story data.", icon="❗")
                else :
                    st.info("Graphs for Wind Velocity")
                    with st.container(height=int(800 * zoom_ratio), border=False) :
                        with st.expander("Each Direction Parameters for Site Exposure", expanded=True):
                            with st.container(border=False) :
                                if st.session_state.get("modified_mzcat_data") is None :
                                    st.session_state["modified_mzcat_data"] = {}
                                if st.session_state.get("modified_ms_data") is None :
                                    st.session_state["modified_ms_data"] = {}
                                if st.session_state.get("modified_mt_data") is None :
                                    st.session_state["modified_mt_data"] = {}

                                def handle_mfactor_edit():
                                    st.session_state["modified_mzcat_data"] = {}
                                    st.session_state["modified_ms_data"] = {}
                                    st.session_state["modified_mt_data"] = {}
                                    st.session_state["show_modal_mzcat"] = True
                                    st.session_state["show_modal_ms"] = True
                                    st.session_state["show_modal_mt"] = True
                                    
                                    field_type = list(st.session_state["mfactor_data_each_direction"].keys())[1:]
                                    for field in field_type :
                                        if f"previous_{field}" not in st.session_state:
                                            st.session_state[f"previous_{field}"] = {}

                                    current = st.session_state["edited_mfactor_data"]["edited_rows"]
                                    if "prev_edited_rows" not in st.session_state :
                                        st.session_state["prev_edited_rows"] = {}
                                    previous = st.session_state["prev_edited_rows"]
                                    
                                    # 현재 mfactor_data_each_direction 데이터 정리
                                    for current_idx, current_value_dict in current.items() :
                                        for current_field, current_value in current_value_dict.items() :
                                            if f"current_{current_field}" not in st.session_state:
                                                st.session_state[f"current_{current_field}"] = {}
                                            st.session_state[f"current_{current_field}"][current_idx] = current_value

                                    # 이전 mfactor_data_each_direction 데이터 정리
                                    for previous_idx, previous_value_dict in previous.items() :
                                        for previous_field, previous_value in previous_value_dict.items() :
                                            if f"previous_{previous_field}" not in st.session_state:
                                                st.session_state[f"previous_{previous_field}"] = {}
                                            st.session_state[f"previous_{previous_field}"][previous_idx] = previous_value

                                    # field 별 변경된 사항 추출
                                    for field in field_type :
                                        if st.session_state.get(f"current_{field}") is None :
                                            pass
                                        else :
                                            for current_idx, current_value in st.session_state[f"current_{field}"].items() :
                                                if field == "Mzcat" :
                                                    if current_idx not in st.session_state[f"previous_{field}"] :
                                                        st.session_state["mfactor_data_each_direction"][field][current_idx] = current_value
                                                        st.session_state["modified_mzcat_data"][current_idx] = {field: current_value}
                                                    else :
                                                        if str(current_value) == str(st.session_state[f"previous_{field}"][current_idx]) :
                                                            continue
                                                        else :
                                                            st.session_state["mfactor_data_each_direction"][field][current_idx] = current_value
                                                            st.session_state["modified_mzcat_data"][current_idx] = {field: current_value}
                                                elif field == "Ms" :
                                                    if current_idx not in st.session_state[f"previous_{field}"] :
                                                        st.session_state["mfactor_data_each_direction"][field][current_idx] = current_value
                                                        st.session_state["modified_ms_data"][current_idx] = {field: current_value}
                                                    else :
                                                        if str(current_value) == str(st.session_state[f"previous_{field}"][current_idx]) :
                                                            continue
                                                        else :
                                                            st.session_state["mfactor_data_each_direction"][field][current_idx] = current_value
                                                            st.session_state["modified_ms_data"][current_idx] = {field: current_value}
                                                elif field == "Mt" :
                                                    if current_idx not in st.session_state[f"previous_{field}"] :
                                                        st.session_state["mfactor_data_each_direction"][field][current_idx] = current_value
                                                        st.session_state["modified_mt_data"][current_idx] = {field: current_value}
                                                    else :
                                                        if str(current_value) == str(st.session_state[f"previous_{field}"][current_idx]) :
                                                            continue
                                                        else :
                                                            st.session_state["mfactor_data_each_direction"][field][current_idx] = current_value
                                                            st.session_state["modified_mt_data"][current_idx] = {field: current_value}

                                    st.session_state["prev_edited_rows"] = current.copy()
                                    
                                # st.markdown("M<sub>factor</sub>", unsafe_allow_html=True)
                                if "mfactor_multiple_data" not in st.session_state :
                                    st.session_state["mfactor_multiple_data"] = {
                                        "Cardinal Direction" : au_cardinal_directions,
                                        "Mzcat" : [au_terrain_category["Category"][3]] * len(au_cardinal_directions),
                                        "Ms" : [au_shielding_category["shielding_catetory"][0]] * len(au_cardinal_directions),
                                        "Mt" : [au_hill_type["hill_type"][0]] * len(au_cardinal_directions)
                                    }
                                else :
                                    st.session_state["mfactor_multiple_data"] = {
                                        "Cardinal Direction" : au_cardinal_directions,
                                        "Mzcat" : st.session_state["mfactor_data_each_direction"]["Mzcat"],
                                        "Ms" : st.session_state["mfactor_data_each_direction"]["Ms"],
                                        "Mt" : st.session_state["mfactor_data_each_direction"]["Mt"]
                                    }

                                # column_config 설정: 각 열마다 드롭다운
                                column_config = {
                                    "Cardinal Direction": st.column_config.TextColumn(
                                        label="Cardinal Direction",
                                        disabled=True
                                    ),
                                    "Mzcat": st.column_config.SelectboxColumn(
                                        label="Mzcat",
                                        options=au_terrain_category["Category"],
                                        required=True
                                    ),
                                    "Ms": st.column_config.SelectboxColumn(
                                        label="Ms",
                                        options=au_shielding_category["shielding_catetory"],
                                        required=True
                                    ),
                                    "Mt": st.column_config.SelectboxColumn(
                                        label="Mt",
                                        options=au_hill_type["hill_type"],
                                        required=True
                                    )
                                }
                              
                                mfactor_data_each_direction = st.data_editor(
                                    data=st.session_state["mfactor_multiple_data"],
                                    column_config=column_config,
                                    use_container_width=True,
                                    hide_index=True,
                                    key="edited_mfactor_data",
                                    on_change=handle_mfactor_edit
                                )
                                st.session_state["mfactor_data_each_direction"] = mfactor_data_each_direction

                        with st.container(border=False) :
                            velocity_line_chart(df_vsit_ww_each, st.session_state["story_data"]["story_name"], au_cardinal_directions, 'Vsit,β by Story', height=int(320 * zoom_ratio))
                            velocity_line_chart(df_vdes_ww_each, st.session_state["story_data"]["story_name"], list(df_vdes_ww_each.columns[1:]), 'Vdes,θ by Story', height=int(320 * zoom_ratio))

                            # st.write(st.session_state["parameters_merge_dict_each_E"]["Mz,cat"])
                            # st.write(st.session_state["parameters_merge_dict_each_E"]["Ms"])
                            # st.write(st.session_state["parameters_merge_dict_each_E"]["Mt"])
                            # st.write(st.session_state["edited_mfactor_data"])
                            # st.write(st.session_state["modified_mzcat_data"])
                            # st.write(st.session_state["modified_ms_data"])
                            # st.write(st.session_state["modified_mt_data"])

    with c1 :
        col_button1, col_button2 = st.columns(spec=[1,1], vertical_alignment ="center")

        # Wind Loac 계산 및 결과 그래프/테이블 출력
        with col_button1 :
            calculate_button_label = "Back to Page" if st.session_state["wind_load_calculated"] == True else "Calculate wind load"
            if st.button(calculate_button_label, key="calc_wind_load", use_container_width=True) :
                st.session_state["wind_load_calculated"] = not st.session_state["wind_load_calculated"]
                try :
                    with st.spinner("Calculating Data...") :
                        if "wind_pressure_data_all" not in st.session_state:
                            st.session_state["wind_pressure_data_all"] = {}
                            st.session_state["wind_pressure_data_all"]["Story"] = st.session_state["story_data"]["story_name"]
                        if "vdes_data_each" not in st.session_state:
                            st.session_state["wind_pressure_data_each"] = {}
                            st.session_state["wind_pressure_data_each"]["Story"] = st.session_state["story_data"]["story_name"]
                        
                        for building_angle in au_building_angle :
                            if building_angle in [0, 180] :
                                applied_cpe_lw = input_cpe_lw_x
                                vdes_all = df_vdes_ww_all[f"{building_angle}°"]
                                vdes_each = df_vdes_ww_each[f"{building_angle}°"]

                                wind_pressure_ww_temp_all_value = 0.5 * au_rho_air * vdes_all**2 * input_cpe * cdyn
                                wind_pressure_lw_temp_all_value = 0.5 * au_rho_air * vdes_all**2 * applied_cpe_lw * cdyn
                                wind_pressure_temp_all_value = wind_pressure_ww_temp_all_value - wind_pressure_lw_temp_all_value

                                wind_pressure_ww_temp_each_value = 0.5 * au_rho_air * vdes_each**2 * input_cpe * cdyn
                                wind_pressure_lw_temp_each_value = 0.5 * au_rho_air * vdes_each**2 * applied_cpe_lw * cdyn
                                wind_pressure_temp_each_value = wind_pressure_ww_temp_each_value - wind_pressure_lw_temp_each_value

                            else :
                                applied_cpe_lw = input_cpe_lw_y
                                vdes_all = df_vdes_ww_all[f"{building_angle}°"]
                                vdes_each = df_vdes_ww_each[f"{building_angle}°"]

                                wind_pressure_ww_temp_all_value = 0.5 * au_rho_air * vdes_all**2 * input_cpe * cdyn
                                wind_pressure_lw_temp_all_value = 0.5 * au_rho_air * vdes_all**2 * applied_cpe_lw * cdyn
                                wind_pressure_temp_all_value = wind_pressure_ww_temp_all_value - wind_pressure_lw_temp_all_value

                                wind_pressure_ww_temp_each_value = 0.5 * au_rho_air * vdes_each**2 * input_cpe * cdyn
                                wind_pressure_lw_temp_each_value = 0.5 * au_rho_air * vdes_each**2 * applied_cpe_lw * cdyn
                                wind_pressure_temp_each_value = wind_pressure_ww_temp_each_value - wind_pressure_lw_temp_each_value

                            st.session_state[f"wind_pressure_data_all"][f"{building_angle}°"] = wind_pressure_temp_all_value / 1000
                            st.session_state[f"wind_pressure_data_each"][f"{building_angle}°"] = wind_pressure_temp_each_value / 1000

                        df_wind_pressure_all = st.session_state[f"wind_pressure_data_all"]
                        df_wind_pressure_each = st.session_state[f"wind_pressure_data_each"]

                except Exception as e:
                    st.toast(f"🔁 Refresh failed: {e}")
                st.rerun()

        # Export Excel Data
        with col_button2:
            if st.session_state["stor_valid_flag"] == True and st.session_state["wind_load_calculated"] == True:
                # 엑셀 원본을 바이너리로 열되, BytesIO로 복사
                with open("./resources/as1170_2_2021/AS1170_2_2021.xlsx", "rb") as base_excel:
                    in_memory_excel = io.BytesIO(base_excel.read())

                # openpyxl로 메모리상에서 workbook 로드
                wb = load_workbook(in_memory_excel)
                
                # 데이터 위치 지정
                start_row = 6
                start_col = 1

                if st.session_state["input_condition"] == "Input Only Worst-Case Value" :
                    for building_angle in au_building_angle :
                        original_ws = wb["Wind Loads"]
                        ws = wb.copy_worksheet(original_ws)
                        ws.title = f"Wind Loads_{building_angle}"

                        selected_vdes_list = st.session_state[f"vdes_data(ww)_all"][f"{building_angle}°"]
                        selected_vdes_df = pd.DataFrame(selected_vdes_list)
                        selected_vdes_df.columns = ["Vdes(ww)"]
                        selected_vdes_df["Vdes(lw)"] = st.session_state[f"vdes_data(lw)_all"][f"{building_angle}°"]
                        selected_vdes_df["Cpe(ww)"] = st.session_state["cpe_windward"]

                        if building_angle in [0, 180] :
                            selected_vdes_df["Cpe(lw)"] = st.session_state["cpe_leeward_x"]
                        else :
                            selected_vdes_df["Cpe(lw)"] = st.session_state["cpe_leeward_y"]
                        selected_vdes_df["Cdyn"] = st.session_state["cdyn"]

                        selected_result_df_name = f"wind_calc_all_df_{building_angle}"
                        selected_result_df = st.session_state[selected_result_df_name]
                        merged_resultf_df = pd.concat([selected_result_df.iloc[:, :2], selected_vdes_df, selected_result_df.iloc[:, 2:]], axis=1)
                        dataframe_to_excel(ws, merged_resultf_df, building_angle, start_row, start_col)

                        direction_name = {
                            0: "+ X",
                            90: "+ Y",
                            180: "- X",
                            270: "- Y"
                        }
                        angle_text = direction_name.get(building_angle, "UNKNOWN")
                        ws["A1"].value = f"W I N D   L O A D   G E N E R A T I O N   D A T A   {angle_text}   D I R E C T I O N"

                    for cardinal_direction in au_cardinal_directions :
                        original_ws = wb["Wind Params for Vsit"]
                        ws = wb.copy_worksheet(original_ws)
                        ws.title = f"Wind Params for Vsit_{cardinal_direction}"

                        selected_param_df_name = f"parameters_merge_dict_all_{cardinal_direction}"
                        selected_param_df = pd.DataFrame(st.session_state[selected_param_df_name])
                        dataframe_to_excel(ws, selected_param_df, cardinal_direction, start_row, start_col)

                        ws["A1"].value = f"P A R A M E T E R S   F O R    V S I T   ( {cardinal_direction} )"

                elif st.session_state["input_condition"] == "Input for Each Direction" :
                    for building_angle in au_building_angle :
                        original_ws = wb["Wind Loads"]
                        ws = wb.copy_worksheet(original_ws)
                        ws.title = f"Wind Loads_{building_angle}"

                        selected_vdes_list = st.session_state[f"vdes_data(ww)_each"][f"{building_angle}°"]
                        selected_vdes_df = pd.DataFrame(selected_vdes_list)
                        selected_vdes_df.columns = ["Vdes(ww)"]
                        selected_vdes_df["Vdes(lw)"] = st.session_state[f"vdes_data(lw)_each"][f"{building_angle}°"]
                        selected_vdes_df["Cpe(ww)"] = st.session_state["cpe_windward"]
                        if building_angle in [0, 180] :
                            selected_vdes_df["Cpe(lw)"] = st.session_state["cpe_leeward_x"]
                        else :
                            selected_vdes_df["Cpe(lw)"] = st.session_state["cpe_leeward_y"]
                        selected_vdes_df["Cdyn"] = st.session_state["cdyn"]

                        selected_result_df_name = f"wind_calc_each_df_{building_angle}"
                        selected_result_df = st.session_state[selected_result_df_name]

                        merged_resultf_df = pd.concat([selected_result_df.iloc[:, :2], selected_vdes_df, selected_result_df.iloc[:, 2:]], axis=1)
                        dataframe_to_excel(ws, merged_resultf_df, building_angle, start_row, start_col)

                        direction_name = {
                            0: "+ X",
                            90: "+ Y",
                            180: "- X",
                            270: "- Y"
                        }
                        angle_text = direction_name.get(building_angle, "UNKNOWN")
                        ws["A1"].value = f"W I N D   L O A D   G E N E R A T I O N   D A T A   {angle_text} - D I R E C T I O N"

                    for cardinal_direction in au_cardinal_directions :
                        original_ws = wb["Wind Params for Vsit"]
                        ws = wb.copy_worksheet(original_ws)
                        ws.title = f"Wind Params for Vsit_{cardinal_direction}"

                        selected_param_df_name = f"parameters_merge_dict_each_{cardinal_direction}"
                        selected_param_df = pd.DataFrame(st.session_state[selected_param_df_name])
                        dataframe_to_excel(ws, selected_param_df, cardinal_direction, start_row, start_col)

                        ws["A1"].value = f"P A R A M E T E R S   F O R    V S I T   ( {cardinal_direction} )"

                # 시트 삭제
                if "Wind Loads" in wb.sheetnames:
                    wb.remove(wb["Wind Loads"])
                if "Wind Params for Vsit" in wb.sheetnames:
                    wb.remove(wb["Wind Params for Vsit"])

                # 결과를 메모리로 저장
                output_excel = io.BytesIO()
                wb.save(output_excel)
                output_excel.seek(0)  # 파일 포인터를 처음으로 돌려야 다운로드 가능

                b64_excel = base64.b64encode(output_excel.getvalue()).decode()
                st.markdown(
                    f"""
                    <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}"
                  download="AS1170_2_2021_WindLoads.xlsx"
                  target="_parent"
                  style="text-decoration: none;">
                    <button style="
                        background-color: #3498db;
                        color: white;
                        padding: 0.3rem 0.3rem;
                        border: none;
                        border-radius: 0.25rem;
                        cursor: pointer;
                        width: 100%;
                        font-size: 1rem;">
                        Download Excel Data
                    </button>
                  </a>
                      """,
                      unsafe_allow_html=True
                  )
            else :
                st.markdown(
                    '''
                    <button style="
                        background-color: #bdc3c7;
                        color: white;
                        padding: 0.3rem 0.3rem;
                        border: none;
                        border-radius: 0.25rem;
                        width: 100%;
                        font-size: 1rem;
                        cursor: not-allowed;"
                        disabled>
                        Download Excel Data
                    </button>
                    ''',
                    unsafe_allow_html=True
                )

    
    with c2 :
        if st.session_state["stor_valid_flag"] == True and st.session_state["stld_valid_flag"] == True and st.session_state["wind_load_calculated"] == True:
            apply_button = st.button("Apply", type="primary", use_container_width=True)
            if apply_button :
                select_stld_modal(st.session_state["stld_data"]["NAME"])
        else :
            st.button("Apply", type="primary", use_container_width=True, disabled=True, help="Calculate wind load first")
import requests
import numpy as np
import pandas as pd
import streamlit as st
import altair as alt
import time
import io
from openpyxl import load_workbook
from copy import copy
from dotenv import load_dotenv
import base64
from utils.ParamHandler import ParamHandler

# Load environment variables
load_dotenv()

# Set page config must be the first Streamlit command
st.set_page_config(page_title="MS 1553:2002, AMD.1:2013", page_icon="༄", layout="wide")  # wide

st.markdown("""
    <style>
        header[data-testid="stHeader"] {
            display: none !important;
        }
        .main > div {
            padding-top: 0rem;
        }
        .block-container {
            padding-top: 0rem;
            padding-bottom: 0rem;
        }
        [data-testid="stToolbar"] {
            display: none !important;
        }
        [data-testid="stDecoration"] {
            display: none !important;
        }
        #MainMenu {
            display: none !important;
        }
        [data-testid="stStatusWidget"] {
            display: none !important;
        }
        footer {
            display: none !important;
        }
    </style>
""", unsafe_allow_html=True)

# --- URL 파라미터 읽기 ---
_base_url = ParamHandler.base_url
_mapi_key = ParamHandler.mapi_key
base_url = f"{_base_url}/gen"

# =======================================================================================================
# ============================================ DEFAULT VALUE ============================================
# =======================================================================================================
# STOR / STLD 데이터 유효성 체크 FLAG
st.session_state["stor_valid_flag"] = False
st.session_state["stld_valid_flag"] = False

# Importance Factor
my_importance_factor = {
    "Category" : [1, 2, 3, 4],
    "I" : [0.87, 1.00, 1.15, 1.15]
}
# terrain category
my_terrain_category = {
    "Category" : ["Category1", "Category2", "Category3", "Category4", "Select Multiple"],
    "Explain" : [
        "Exposed open terrain with few or no obstructions.",
        "Water surfaces, open terrain, grassland with few well scattered obstructions having height generally from 1.5 m to 10.0 m.",
        "Terrain with numerous closely spaced obstructions 3.0 m to 5.0 m high such as areas of suburban housing.",
        "Terrain with numerous large, high (10.0 m to 30.0 m high) and closely spaced obstructions such as large city centers and well-developed industrial complexes."
    ]
}
my_terrain_category_modal = {
    "Category" : ["Category1", "Category2", "Category3", "Category4"],
    "Explain" : [
        "Exposed open terrain with few or no obstructions.",
        "Water surfaces, open terrain, grassland with few well scattered obstructions having height generally from 1.5 m to 10.0 m.",
        "Terrain with numerous closely spaced obstructions 3.0 m to 5.0 m high such as areas of suburban housing.",
        "Terrain with numerous large, high (10.0 m to 30.0 m high) and closely spaced obstructions such as large city centers and well-developed industrial complexes."
    ]
}
# Md (Directional Factor)
my_md = 1.00
# Define terrain category data (for mzcat)
my_mzcat_db = {
    "Heights" : [3, 5, 10, 15, 20, 30, 40, 50, 75, 100, 150, 200, 250, 300, 400, 500],
    "Category1" : [0.99, 1.05, 1.12, 1.16, 1.19, 1.22, 1.24, 1.25, 1.27, 1.29, 1.31, 1.32, 1.34, 1.35, 1.37, 1.38],
    "Category2" : [0.85, 0.91, 1.00, 1.05, 1.08, 1.12, 1.16, 1.18, 1.22, 1.24, 1.27, 1.29, 1.31, 1.32, 1.35, 1.37],
    "Category3" : [0.75, 0.75, 0.83, 0.89, 0.94, 1.00, 1.04, 1.07, 1.12, 1.16, 1.21, 1.24, 1.27, 1.29, 1.32, 1.35],
    "Category4" : [0.75, 0.75, 0.75, 0.75, 0.75, 0.80, 0.85, 0.90, 0.98, 1.03, 1.11, 1.16, 1.20, 1.23, 1.28, 1.31]
}
# Define roughness length by terrain category
my_roughness_length = {
    "Category" : ["Category1", "Category2", "Category3", "Category4"],
    "Roughness_Length" : [0.002, 0.02, 0.2, 2.0]
}
# calculate shielding parameter, s
my_shielding_parameter = {
    "s" : [1.5, 3, 6, 12],
    "Ms" : [0.7, 0.8, 0.9, 1.0],
    "Ms_normal" : [0.85, 0.85, 0.85, 0.85],
    "Ms_noShield" : [1.0, 1.0, 1.0, 1.0]
}
# Mh
my_mh_table = {
    "up_slope" : [0.05, 0.05, 0.1, 0.2, 0.3, 0.44],
    "mh" : [1.0, 1.08, 1.16, 1.33, 1.49, 1.71]
}
# rho_air
# Shielding Category
my_shileding_category = {
    "shielding_catetory" : ["No shielding", "Shielding", "Normal suburban housing"]
}
# Hill Tpe (default : "DO not consider")
my_hill_type = {
    "hill_type" : ["Do not consider", "Ridge", "Escarpment"]
}
my_rho_air = 1.200 # kg/m3
# Cpe_windward
my_cpe_ww = {
    "Cpe_windward" : [0.8, 0.7]
}
# Cpe_leeward
my_cpe_lw = {
    "d/b" : [1, 2, 4],
    "Cpe" : [-0.5, -0.3, -0.2]
}
# Shielding Type (default : "No shielding")
my_shielding_category = {
    "shielding_catetory" : ["No shielding", "Shielding", "Normal suburban housing"]
}
# Wind Direction 
my_wind_direction = {
    "wind_direction" : ["x_dir", "y_dir"]
}
# Kc (combination factor)
my_kc = {
    "kc" : [1.0, 0.95, 0.8]
}
my_structure_position = ["Upwind", "Downwind"]

# =======================================================================================================
# ============================================= Definitions =============================================
# =======================================================================================================
# Extract Story Data
def get_stor_data(body) :
    header = {"MAPI-Key": _mapi_key}
    try:
        response = requests.post(url=base_url + "/ope/storyprop", headers=header, json=body)
        response.raise_for_status()  # HTTP 오류 확인
        get_story_data_origin = response.json()
        st.session_state["stor_valid_flag"] = True
    except requests.exceptions.RequestException as e:
        get_story_data_origin = {
            "STORYPROP": {
                "FORCE": "KN",
                "LENGTH": "M",
                "HEAD": [
                    "Story",
                    "Weight",
                    "Elev.",
                    "Loaded H",
                    "Loaded Bx",
                    "Loaded By"
                ],
                "DATA": [
                    {
                        "STORY": "ROOF",
                        "WEIGHT": "46482.2030",
                        "ELEV": "7.5000",
                        "LOADED_H": "2.2500",
                        "LOADED_BX": "29.1000",
                        "LOADED_BY": "36.0000"
                    },
                    {
                        "STORY": "1F",
                        "WEIGHT": "6694.3802",
                        "ELEV": "3.0000",
                        "LOADED_H": "2.2500",
                        "LOADED_BX": "29.1000",
                        "LOADED_BY": "36.0000"
                    },
                    {
                        "STORY": "G.L.",
                        "WEIGHT": "0.0000",
                        "ELEV": "0.0000",
                        "LOADED_H": "0.0000",
                        "LOADED_BX": "29.1000",
                        "LOADED_BY": "36.0000"
                    }
                ]
            }
        }
        st.session_state["stor_valid_flag"] = None
    except ValueError:
        st.error("Please check the structure of the Story data")
        st.stop()

    story_name = []
    story_elev = []
    story_weight = []
    loaded_h = []
    bx = []
    by = []

    for stor_cnt in range(len(get_story_data_origin["STORYPROP"]["DATA"])) :
        # Get Story Data
        story_name.append(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["STORY"])
        story_elev.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["ELEV"]))
        story_weight.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["WEIGHT"]))
        loaded_h.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["LOADED_H"]))
        bx.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["LOADED_BX"]))
        by.append(float(get_story_data_origin["STORYPROP"]["DATA"][stor_cnt]["LOADED_BY"]))

    return {
        "story_name" : story_name,
        "story_elev" : story_elev,
        "story_weight" : story_weight,
        "loaded_h" : loaded_h,
        "loaded_bx" : bx,
        "loaded_by" : by
    }
# # calculate sample
# story_data = get_stor_data()
# print(story_data)

def get_stld_data() :
    header = {"MAPI-Key": _mapi_key}
    try:
        response = requests.get(url=base_url + "/db/stld", headers=header)
        response.raise_for_status()  # HTTP 오류 확인
        stld_data = response.json()
        if stld_data["STLD"] == {} :
            st.session_state["stld_valid_flag"] = False
        else :
            st.session_state["stld_valid_flag"] = True
    except requests.exceptions.RequestException as e:
        st.error(f"서버 요청에 실패했습니다: {e}")
        st.stop()
    except ValueError:
        st.error("응답을 JSON으로 파싱하는 데 실패했습니다.")
        st.stop()

    stld_idx = []
    stld_name = []
    stld_type = []

    stld_idx = list(stld_data["STLD"].keys())
    for index in stld_idx :
        stld_name.append(stld_data["STLD"][str(index)]["NAME"])
        stld_type.append(stld_data["STLD"][str(index)]["TYPE"])

    return {
        "INDEX" : stld_idx,
        "NAME" : stld_name,
        "TYPE" : stld_type
    }

# Define Mz,cat for MY
def my_mzcat(input_terrain_category, input_roughness_legnth = None) :
    # Define Terrain Category
    mzcat = {}
    if len(input_terrain_category) == 0 :
        st.error('Check the input(Mz,cat) data', icon="🚨")
    elif len(input_terrain_category) == 1 :
        mzcat_calc = []
        for cnt in range(len(st.session_state["story_data"]["story_name"])) :
            # Input mzcat
            mzcat_calc.append(np.interp(st.session_state["story_data"]["story_elev"][cnt], my_mzcat_db["Heights"], my_mzcat_db[input_terrain_category[0]]))
        mzcat["Mz,cat"] = mzcat_calc

    elif len(input_terrain_category) >= 2 :
        zor = []
        mzcat_calc = []

        # category lenght 계산
        if sum(input_roughness_legnth) > 3000 :
            input_roughness_legnth[-1] = input_roughness_legnth[-1] - (sum(input_roughness_legnth)-3000)
        else :
            input_roughness_legnth[-1] = input_roughness_legnth[-1] + (3000-sum(input_roughness_legnth))

        # zor 계산
        for index in range(len(input_roughness_legnth)) :
            if index+1 == len(input_roughness_legnth) :
                inner_index = my_roughness_length["Category"].index(input_terrain_category[-1])
                zor.append(my_roughness_length["Roughness_Length"][inner_index])
            else :
                cat_1 = input_terrain_category[index]
                cat_2 = input_terrain_category[index+1]
                inner_index_1 = my_roughness_length["Category"].index(cat_1)
                inner_index_2 = my_roughness_length["Category"].index(cat_2)
                zor_1 = my_roughness_length["Roughness_Length"][inner_index_1]
                zor_2 = my_roughness_length["Roughness_Length"][inner_index_2]
                zor.append(max(zor_1, zor_2))

        # x_i, x_t 계산
        for height in st.session_state["story_data"]["story_elev"] :
            xi_list = []
            xt_list = []
            mzcat_temp = []
            for index in range(len(zor)) :
                # index == 0일 경우, 원래 길이에서 뺄 x_i는 있으나 더해줄 x_i-1이 없음
                if index == 0 :
                    xi = zor[index] * ((float(height)/(0.3*zor[index]))**1.25)
                    xi_list.append(xi)
                    xt_list.append(input_roughness_legnth[index] - xi)
                    mzcat_temp.append(np.interp(height, my_mzcat_db["Heights"], my_mzcat_db[input_terrain_category[index]]))
                # 마지막 index의 경우, 맨 마지막 길이이기 때문에 뺄 x_i가 없음
                elif index == len(zor)-1 :
                    xi_list.append(0)
                    xt_list.append(input_roughness_legnth[index] + xi_list[index-1])
                    mzcat_temp.append(np.interp(height, my_mzcat_db["Heights"], my_mzcat_db[input_terrain_category[index]]))
                else :
                    xi = zor[index] * ((float(height)/(0.3*zor[index]))**1.25)
                    xi_list.append(xi)
                    xt_list.append(input_roughness_legnth[index] - xi + xi_list[index-1])
                    mzcat_temp.append(np.interp(height, my_mzcat_db["Heights"], my_mzcat_db[input_terrain_category[index]]))

            # 마지막으로 최종 Mzcat 계산
            mzcat_calc.append(sum([x * y for x, y in zip(xt_list, mzcat_temp)]) / 3000)
        # mzcat_dict["xt"] = xt_list
        mzcat["Mz,cat"] = mzcat_calc
    else :
        raise ValueError("Check the terrain category input data")
    return mzcat
# calculate sample
# mzcat = mzcat_my(input_terrain_category)
# print(mzcat)

# Define Ms for MY
def my_ms(input_shielding_category, hs, bs, ns) :
    # calculate shielding parameter, s
    db_shielding_parameter = {
        "s" : [1.5, 3, 6, 12],
        "Ms" : [0.7, 0.8, 0.9, 1.0],
        "Ms_normal" : [0.85, 0.85, 0.85, 0.85],
        "Ms_noShield" : [1.0, 1.0, 1.0, 1.0]
    }
    # Calculate Ms
    if input_shielding_category == "Normal suburban housing" :
        ms = 0.85
    elif input_shielding_category == "No shielding" :
        ms = 1
    else :
        # Calculate l_s
        ls = hs * (10/ns + 5)
        # Calculate s
        s = ls / ((hs * bs)**0.5)
        # Calculate Ms
        ms = np.interp(s, db_shielding_parameter["s"], db_shielding_parameter["Ms"])
    return {
        "Ms" : ms
    }
# # calculate sample
# ms_dict = my_ms("Shielding", 10, 20, 5)
# print(ms_dict)

# Define Mh for MY
def my_mh(input_hill_type, hill_height=None, Lu=None, x=None) :
    mh_ww_list = []
    mh_lw_list = []

    hill_shape = input_hill_type
    h = hill_height
    lu = Lu
    dist = x

    db_mh_table = {
        "up_slope" : [0.05, 0.05, 0.1, 0.2, 0.3, 0.44],
        "mh" : [1.0, 1.08, 1.16, 1.33, 1.49, 1.71]
    }
    
    if input_hill_type == "Do not consider" :
        mh_ww = mh_lw = 1.0
        for z in st.session_state["story_data"]["story_elev"] :
            mh_ww_list.append(mh_ww)
            mh_lw_list.append(mh_lw)
    elif input_hill_type != "Do not consider" :
        if hill_height == None or Lu == None or x == None :
            raise ValueError("Check the hill type input data")
        else :
            slope = h/(2*lu)

            # Calculate L1
            l1 = max(0.4*h, 0.36*lu)
            # Calculate L2
            if hill_shape == "Escarpment" :
                l2_up = 4 * l1
                l2_dn = 10 * l1
            else :
                l2_up = l2_dn = 4 * l1

            for z in st.session_state["story_data"]["story_elev"] :
                # Calculate Mh (up / down)
                if z == 0 and x == 0 : # Table 4.4 에 따라 Mh 산정
                    if slope < 0.05 :
                        mh_ww = mh_lw = db_mh_table["mh"][0]
                    else :
                        mh_ww = np.interp(slope, db_mh_table["up_slope"][1:], db_mh_table["mh"][1:])
                        mh_lw = mh_ww
                else : # 4.4 a) ~ c)에 따라 계산
                    if slope < 0.05 :
                        mh_ww = mh_lw = 1
                    elif slope <= 0.44 :
                        mh_ww = 1 + (h / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_up)
                        mh_lw = 1 + (h / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_dn)
                    elif slope > 0.44 :
                        if dist < h/4 and z <= h/10 :
                            mh_ww = 1 + 0.71 * (1 - abs(dist)/l2_up)
                            mh_lw = 1 + 0.71 * (1 - abs(dist)/l2_dn)
                        else :
                            mh_ww = 1 + (h / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_up)
                            mh_lw = 1 + (h / (3.5*(float(z) + l1))) * (1 - abs(dist)/l2_dn)
                mh_ww_list.append(mh_ww)
                mh_lw_list.append(mh_lw)

    return {
        "Mh(ww)" : mh_ww_list,
        "Mh(lw)" : mh_lw_list
    }

# Define Cpe_external / Cpi는 추후 wind pressure 기능으로 따로 제작해야 할 듯
def my_cpe_leeward(wind_direction) :
    # Cpe_Lee Ward 계산
    bx = [float(x) for x in st.session_state["story_data"]["loaded_bx"]]
    by = [float(y) for y in st.session_state["story_data"]["loaded_by"]]
    db_x_dir = [bx/by for bx, by in zip(bx, by)]
    db_y_dir = [by/bx for bx, by in zip(bx, by)]

    # alpha = 0 으로 가정
    if wind_direction == "x_dir" :
        cpe_lw = np.interp(db_x_dir, my_cpe_lw["d/b"], my_cpe_lw["Cpe"])
    else :
        cpe_lw = np.interp(db_y_dir, my_cpe_lw["d/b"], my_cpe_lw["Cpe"])
    cpe_lw_list = cpe_lw.tolist()

    return {
        "Cpe_lw" : cpe_lw_list
    }
# # calculate sample
# d_b = my_cpe()
# print(f"d/b = {d_b}")

# copy value (층 개수만큼 복사)
def copy_value_by_story(number, input_value) :
    return [input_value] * number if not isinstance(input_value, (list, pd.Series)) else input_value

# Modal Dialog : Mz,cat
@st.dialog("Select Multiple Terrain Category")
def my_mzcat_modal(wind_direction) :
    st.session_state["my_mzcat_button_flag"] = False

    def add_row() :
        selected_value = st.session_state[f"terrain_cat_drop_multi_{wind_direction}"]
        if selected_value is not None:
            st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"].append(selected_value)
            st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"].append(1000)

    def renew_table_data() :
        edited_row = st.session_state[f"edited_df_{wind_direction}"]["edited_rows"]  # dictionary type (index + value)
        added_row = st.session_state[f"edited_df_{wind_direction}"]["added_rows"]  # list type (index는 무조건 0부터 시작)
        indices_deleted = st.session_state[f"edited_df_{wind_direction}"]["deleted_rows"]  # index list type

        for index in edited_row :
            if len(edited_row[index].keys()) == 0 :
                continue
            elif len(edited_row[index].keys()) == 1 :
                if "Terrain Category" in edited_row[index] :
                    st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"][index] = edited_row[index]["Terrain Category"]
                elif "Roughness Length(m)" in edited_row[index] :
                    st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][index] = edited_row[index]["Roughness Length(m)"]
            elif len(edited_row[index].keys()) == 2 :
                st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"][index] = edited_row[index]["Terrain Category"]
                st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][index] = edited_row[index]["Roughness Length(m)"]
            else :
                st.error("Please check the data")

        for rows_value in added_row :
            if len(rows_value.keys()) == 0 :
                continue
            elif len(rows_value.keys()) == 2 :
                st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"].append(rows_value["Terrain Category"])
                st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"].append(rows_value["Roughness Length(m)"])
            else :
                st.error("Please check the data")

        for index in sorted(indices_deleted, reverse=True) :
            st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"].pop(index)
            st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"].pop(index)

    st.info("Start entering values from the location nearest to the structure.")

    if f"terrain_table_data_{wind_direction}" not in st.session_state:
        st.session_state[f"terrain_table_data_{wind_direction}"] = {
            "Terrain Category": [],
            "Roughness Length(m)": []
        }

    mzcat_col_1, mzcat_col_2 = st.columns(2, vertical_alignment ="center")
    with mzcat_col_1 :
        st.write("Terrain Category")
    with mzcat_col_2 :
        st.selectbox("Terrain Category", my_terrain_category_modal["Category"], index=None, key=f"terrain_cat_drop_multi_{wind_direction}", label_visibility="collapsed", on_change=add_row)
    
    # 다수 Terrain Category 입력 테이블 출력
    st.data_editor(
        data=st.session_state[f"terrain_table_data_{wind_direction}"],
        hide_index=True,
        use_container_width=True,
        num_rows="dynamic",
        key=f"edited_df_{wind_direction}",
        on_change=renew_table_data
    )

    if len(st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"]) * len(st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"]) == 0 :
        st.warning("Please select or input terrain category", icon="⚠️")
    
    if st.button("Calculate", use_container_width=True) :
        st.session_state["my_mzcat_button_flag"] = True
        try :
            if st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][0] >= 3000 or len(st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"]) == 1 :
                mzcat_result = my_mzcat(st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"][0:1])
                st.session_state[f"mzcat_result_{wind_direction}"] = mzcat_result
                st.warning(f"Single Terrain Category has been applied")
                time.sleep(2)
            else :
                roughness_length = st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"]
                if sum(roughness_length) == 3000 :
                    pass
                elif sum(roughness_length) < 3000 :
                    st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][-1] = st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][-1] + (3000 - sum(st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"]))
                    st.warning(f"Last category adjusted to 3000m total")
                    time.sleep(2)
                else :
                    length_sum = 0
                    for roughness_idx, length in enumerate(roughness_length) :
                        length_sum += length
                        if length_sum >= 3000 :
                            selected_index = roughness_idx
                            st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"] = st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"][:selected_index+1]
                            st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"] = st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][:selected_index+1]
                            st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][-1] = st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"][-1] - (sum(st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"]) - 3000)
                            break
                    st.warning(f"Last category adjusted to 3000m total")
                    time.sleep(2)

                mzcat_result = my_mzcat(
                    st.session_state[f"terrain_table_data_{wind_direction}"]["Terrain Category"],
                    st.session_state[f"terrain_table_data_{wind_direction}"]["Roughness Length(m)"]
                )
                st.session_state[f"mzcat_result_{wind_direction}"] = mzcat_result

            del st.session_state[f"edited_df_{wind_direction}"]
            st.session_state["mzcat_button"] = False
            st.rerun()

        except Exception as e:
            st.warning(f"❗ Exception: {e}")

# Modal Dialog : Ms
@st.dialog("Calculate Mₛ")
def my_ms_modal(wind_direction) :
    hs_key = f"hs_{wind_direction}"
    bs_key = f"bs_{wind_direction}"
    ns_key = f"ns_{wind_direction}"
    if hs_key not in st.session_state:
        st.session_state[hs_key] = 10.0
    if bs_key not in st.session_state:
        st.session_state[bs_key] = 10.0
    if ns_key not in st.session_state:
        st.session_state[ns_key] = 5

    ms_col1, ms_col2 = st.columns(2, vertical_alignment ="center")
    with ms_col1 :
        st.markdown("h<sub>s</sub>(m)", unsafe_allow_html=True)
    with ms_col2 :
        st.number_input("hs(m)", min_value=0.0, step=1.0, value=float(st.session_state[hs_key]), key="hs_temp", label_visibility="collapsed")
        st.session_state[hs_key] = st.session_state["hs_temp"]

    ms_col3, ms_col4 = st.columns(2, vertical_alignment ="center")
    with ms_col3 :
        st.markdown("b<sub>s</sub>(m)", unsafe_allow_html=True)
    with ms_col4 :
        st.number_input("bs(m)", min_value=0.0, step=1.0, value=float(st.session_state[bs_key]), key="bs_temp", label_visibility="collapsed")
        st.session_state[bs_key] = st.session_state["bs_temp"]

    ms_col5, ms_col6 = st.columns(2, vertical_alignment ="center")
    with ms_col5 :
        st.markdown("n<sub>s</sub>", unsafe_allow_html=True)
    with ms_col6 :
        st.number_input("ns", min_value=0, step=1, value=int(st.session_state[ns_key]), key="ns_temp", label_visibility="collapsed")
        st.session_state[ns_key] = st.session_state["ns_temp"]

    if st.button("Calculate", key="calc_ms_button", use_container_width=True) :
        try :
            ms_result = my_ms("Shielding", st.session_state[hs_key], st.session_state[bs_key], st.session_state[ns_key])
            st.session_state[f"ms_result_{wind_direction}"] = ms_result
            st.rerun()
        except Exception as e:
            st.warning(f"❗ Exception: {e}")

# Modal Dialog : Mh
@st.dialog("Calculate Mh")
def my_mh_modal(hill_type, wind_direction) :
    hill_height_temp_key = f"hill_height_temp_{wind_direction}"
    lu_temp_key = f"lu_temp_{wind_direction}"
    x_dist_temp_key = f"x_dist_temp_{wind_direction}"
    hill_height_key = f"hill_height_{wind_direction}"
    lu_key = f"lu_{wind_direction}"
    x_dist_key = f"x_dist_{wind_direction}"
    if hill_height_key not in st.session_state:
        st.session_state[hill_height_key] = 5.0
    if lu_key not in st.session_state:
        st.session_state[lu_key] = 10.0
    if x_dist_key not in st.session_state:
        st.session_state[x_dist_key] = 0.0

    col1_1, col1_2 = st.columns(2, vertical_alignment ="center")
    with col1_1 :
        st.write("Height of Hill(m)")
    with col1_2 :
        st.number_input("hill_height(m)", min_value=0.1, step=1.0, value=float(st.session_state[hill_height_key]), key=hill_height_temp_key, label_visibility="collapsed")
        st.session_state[hill_height_key] = st.session_state[hill_height_temp_key]

    col1_3, col1_4 = st.columns(2, vertical_alignment ="center")
    with col1_3 :
        st.markdown("L<sub>u</sub>(m)", unsafe_allow_html=True)
    with col1_4 :
        st.number_input("Lu(m)", min_value=0.1, step=1.0, value=float(st.session_state[lu_key]), key=lu_temp_key, label_visibility="collapsed")
        st.session_state[lu_key] = st.session_state[lu_temp_key]

    col1_5, col1_6 = st.columns(2, vertical_alignment ="center")
    with col1_5 :
        st.write("x(m)")
    with col1_6 :
        st.number_input("x(m)", min_value=0.0, step=1.0, value=float(st.session_state[x_dist_key]), key=x_dist_temp_key, label_visibility="collapsed")
        st.session_state[x_dist_key] = st.session_state[x_dist_temp_key]

    if st.button("Calculate", use_container_width=True) :
        try :
            mh_result = my_mh(hill_type, st.session_state[hill_height_key], st.session_state[lu_key], st.session_state[x_dist_key])
            st.session_state[f"mh_result_{wind_direction}"] = mh_result
            st.rerun()
        except Exception as e:
            st.warning(f"❗ Exception: {e}")

def bar_chart_hor(df: pd.DataFrame, unit : str):
    # 데이터 복사 및 컬럼 이름 자동 감지
    df = df.copy()
    col_y = df.columns[0]  # Story (범주형)
    col_x = df.columns[1]  # Force 등 (수치형)

    # Story 순서 역순으로 지정 (Roof → G.L.)
    story_order = list(df[col_y])
    df[col_y] = pd.Categorical(df[col_y], categories=story_order, ordered=True)

    # 수치형으로 안전하게 변환
    df[col_x] = df[col_x].astype(float)

    # Altair 그래프 생성
    bar = (
        alt.Chart(df)
        .mark_bar()
        .encode(
            y=alt.Y(f"{col_y}:N", sort=story_order, title=col_y),
            x=alt.X(f"{col_x}:Q", title=f"{col_x} {unit}"),
            tooltip=[col_y, col_x]
        )
    )
    text = (
        alt.Chart(df)
        .mark_text(
            align='right',
            baseline='middle',
            dx=-6,
            color='white',  # 흰색 텍스트
            fontSize=14  # 텍스트 크기 조정
        )
        .encode(
            y=alt.Y(f"{col_y}:N", sort=story_order),
            x=alt.X(f"{col_x}:Q"),
            text=alt.Text(f"{col_x}:Q", format=".2f")
        )
    )
    chart = (bar + text).properties(height=423)
    return chart

@st.dialog("Select Load Case")
def select_stld_modal(stld_name) :
    col1, col2 = st.columns(2)
    with col1 :
        st.write("Load Case for X-Dir")
    with col2 :
        st.selectbox("Load Case for X-Dir", options=stld_name, key="stld_x_temp", label_visibility="collapsed")

    col3, col4 = st.columns(2)
    with col3 :
        st.write("Load Case for Y-Dir")
    with col4 :
        st.selectbox("Load Case for Y-Dir", options=stld_name, key="stld_y_temp", label_visibility="collapsed")

    if st.button("Apply", use_container_width=True) :
        st.session_state["stld_x"] = st.session_state["stld_x_temp"]
        st.session_state["stld_y"] = st.session_state["stld_y_temp"]
        stld_x_idx = st.session_state["stld_data"]["NAME"].index(st.session_state["stld_x"])
        stld_y_idx = st.session_state["stld_data"]["NAME"].index(st.session_state["stld_y"])
        stld_x_type = st.session_state["stld_data"]["TYPE"][stld_x_idx]
        stld_y_type = st.session_state["stld_data"]["TYPE"][stld_y_idx]

        if stld_x_type != "W" or stld_y_type != "W" :
            st.warning("Wind load case must be selected.", icon="❗")
        elif st.session_state["stld_x"] == st.session_state["stld_y"] :
            st.warning("Load Case - X and Load Case - Y must be different.", icon="❗")
        else :
            error_flag = False
            try :
                header = {"MAPI-Key":_mapi_key}

                # X방향 하중 입력
                body_x = {
                    "Assign": {
                         st.session_state["stld_data"]["INDEX"][stld_x_idx]: {
                            "WIND_CODE": "USER TYPE",
                            "DESC": "",
                            "SCALE_FACTOR_X": 1,
                            "SCALE_FACTOR_Y": 0,
                            "SCALE_FACTOR_R": 0,
                            "USER": apply_wind_data(wind_calc_df_x, "x_dir")
                        }
                    }
                }

                res_x = requests.put(url=f"{base_url}/db/wind", headers=header, json=body_x)
                if res_x.status_code == 200:
                    st.success("X-direction wind loads successfully applied.", icon="✅")
                else:
                    error_detail = res_x.json().get("message") or res_x.text
                    st.error(f"❌ Failed to apply X-direction loads.\n\n**Status Code:** {res_x.status_code}\n**Details:** {error_detail}")
                    error_flag = True

                # Y방향 하중 입력
                body_y = {
                    "Assign": {
                         st.session_state["stld_data"]["INDEX"][stld_y_idx]: {
                            "WIND_CODE": "USER TYPE",
                            "DESC": "",
                            "SCALE_FACTOR_X": 0,
                            "SCALE_FACTOR_Y": 1,
                            "SCALE_FACTOR_R": 0,
                            "USER": apply_wind_data(wind_calc_df_y, "y_dir")
                        }
                    }
                }
                res_y = requests.put(url=f"{base_url}/db/wind", headers=header, json=body_y)
                if res_y.status_code == 200:
                    st.success("Y-direction wind loads successfully applied.", icon="✅")
                else:
                    error_detail = res_y.json().get("message") or res_y.text
                    st.error(f"❌ Failed to apply Y-direction loads.\n\n**Status Code:** {res_y.status_code}\n**Details:** {error_detail}")
                    error_flag = True

            except Exception as e:
                st.error(f"❗ Unexpected error occurred while sending API request.\n\n**Exception:** {e}")
                
            if not error_flag :
                time.sleep(2)
                st.rerun()
                

def apply_wind_data(df, wind_direction) :
    return [
        {
            "STORY_NAME": row["STORY"],
            "ELIV": float(row["ELEV [m]"]),
            "LOAD_H": float(row["Loaded_H [m]"]),
            "LOAD_BX": float(row["Loaded_Bx [m]"]),
            "LOAD_BY": float(row["Loaded_By [m]"]),
            "PRESS_X": row["p [kPa]"] if wind_direction == "x_dir" else 0,
            "PRESS_Y": row["p [kPa]"] if wind_direction == "y_dir" else 0
        }
        for _, row in df.iterrows()
    ]

def dataframe_to_excel(ws, df, start_row=2, start_col=1):
    # 1. start_row - 1 행의 서식을 저장
    template_styles = []
    for col_idx in range(start_col, start_col + len(df.columns)):
        cell = ws.cell(row=start_row - 1, column=col_idx)
        style = {
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format
        }
        template_styles.append(style)

    # 2. DataFrame 내용을 시트에 쓰기 + 서식 적용
    for i, (_, row) in enumerate(df.iterrows()):
        current_row = start_row + i
        for j, value in enumerate(row):
            cell = ws.cell(row=current_row, column=start_col + j)
            cell.value = value
            
            # 서식 복사
            style = template_styles[j]
            cell.font = style['font']
            cell.fill = style['fill']
            cell.border = style['border']
            cell.alignment = style['alignment']
            cell.number_format = style['number_format']
        
        # 행 높이를 20으로 고정
        ws.row_dimensions[current_row].height = 20

    # 3. 서식 복사 후 start_row - 1 행 삭제
    ws.delete_rows(start_row - 1)

# =======================================================================================================
# ============================================== Streamlit ==============================================
# =======================================================================================================
# ============================================== Side Bar ==============================================


st.session_state["api_connected"] = None
print(st.session_state["api_connected"])
print(base_url)
print(_mapi_key)



if not st.session_state["api_connected"] and base_url and _mapi_key :
        try :
            header = {"MAPI-Key":_mapi_key}
            res = requests.get(f"{base_url}/db/unit", headers=header)

            if res.status_code == 200:
                st.session_state["api_connected"] = True

            else :
                st.session_state["api_connected"] = False
                st.toast(f"❌ Failed to connect. Status code: {res.status_code}")

        except Exception as e:
            st.session_state["api_connected"] = False
            st.toast("❌ Connection failed: {e}")

# ============================================== Body ==============================================
if st.session_state["api_connected"] == True :
    story_data_body = {
        "Argument": {
            "FORCE_UNIT": "KN",
            "LENGTH_UNIT": "M"
        }
    }
    get_story_data = get_stor_data(body = story_data_body)
    stld_data = get_stld_data()

    if st.session_state["stor_valid_flag"] == True or st.session_state["stor_valid_flag"] == None :
        st.session_state["story_data"] = get_story_data
    elif st.session_state["stld_valid_flag"] == True :
        st.session_state["stld_data"] = stld_data


    tab_x, tab_y = st.tabs(["X-Dir", "Y-Dir"])
    with tab_x :
        selected_wind_direction = my_wind_direction["wind_direction"][0]       # "x_dir"
        c1, c2 = st.columns([1, 1.5])
        # 왼쪽 화면 설정
        with c1 : 
            # Parameters for wind pressure
            st.info("Parameters for Vdes")
            # st.markdown(
            # """
            # <div style='background-color:#d0ebff; padding: 10px; border-radius: 5px;'>
            #     <b>Parameters for V<sub>des</sub></b>
            # </div>
            # """,
            # unsafe_allow_html=True
            # ) 
            with st.container(border=False) :

                # Input Vs
                col1_1, col1_2 = st.columns(2, vertical_alignment ="center")
                with col1_1 :
                    st.markdown("V<sub>s</sub>(m/s)", unsafe_allow_html=True)
                with col1_2 :
                    st.number_input("Vs", min_value=0.0, max_value=100.0, step=0.1, value=31.4, key="vs", label_visibility="collapsed")

                # Calculate Md
                col1_3, col1_4 = st.columns(2, vertical_alignment ="center")
                with col1_3 :
                    st.markdown("M<sub>d</sub>", unsafe_allow_html=True)
                with col1_4 :
                    md = st.number_input(label="Md", value=my_md, format="%0.2f", key="md", label_visibility="collapsed", disabled=True)

                md_dict = {}
                md_list = []
                for i in st.session_state["story_data"]["story_elev"] :
                    md_list.append(my_md)
                md_dict["Md"] = md_list

                # Calculate Mz,cat
                # Flag 추가
                mzcat_modal_falg_x = "show_modal_mzcat_"+selected_wind_direction
                if mzcat_modal_falg_x not in st.session_state:
                    st.session_state[mzcat_modal_falg_x] = False

                mzcat_selectbox_name = "terrain_cat_drop_single_" + selected_wind_direction
                col1_5, col1_6 = st.columns(2, vertical_alignment ="center")
                with col1_5 :
                    st.markdown("M<sub>z,cat</sub>", unsafe_allow_html=True)
                with col1_6 :
                    terrain_cat = st.selectbox("Terrain Category", my_terrain_category["Category"], key=mzcat_selectbox_name ,label_visibility="collapsed", on_change=lambda: st.session_state.update({mzcat_modal_falg_x: True}))

                if terrain_cat == my_terrain_category["Category"][4] :
                    if st.session_state[mzcat_modal_falg_x] == True : # Select Multiple
                        if st.session_state.get(f"mzcat_result_{selected_wind_direction}") is None:
                            mzcat_result_x = my_mzcat([my_terrain_category["Category"][0]])
                        else :
                            mzcat_result_x = st.session_state[f"mzcat_result_{selected_wind_direction}"]
                        my_mzcat_modal(selected_wind_direction)
                        st.session_state[mzcat_modal_falg_x] = False
                    elif terrain_cat == my_terrain_category["Category"][4] and st.session_state[mzcat_modal_falg_x] == False :
                        if st.session_state.get(f"mzcat_result_{selected_wind_direction}") is None:
                            st.error("Please input terrain category data and click [Calculate] button.")
                            st.stop()
                        else :
                            mzcat_result_x = st.session_state[f"mzcat_result_{selected_wind_direction}"]
                else :
                    mzcat_result_x = my_mzcat([st.session_state[mzcat_selectbox_name]])
                
                # Parameters for Ms
                ms_modal_flag_x = "show_modal_ms_"+selected_wind_direction
                if ms_modal_flag_x not in st.session_state:
                    st.session_state[ms_modal_flag_x] = False
                    
                col1_7, col1_8 = st.columns(2, vertical_alignment ="center")
                with col1_7 :
                    st.markdown("M<sub>s</sub>", unsafe_allow_html=True)
                with col1_8 :
                    shileding_category = st.selectbox("Shielding Category", my_shileding_category["shielding_catetory"], key="shielding_catetory_x", label_visibility="collapsed", on_change=lambda: st.session_state.update({ms_modal_flag_x: True}))

                if shileding_category == "Shielding" and st.session_state[ms_modal_flag_x] == True :
                    if st.session_state.get(f"ms_result_{selected_wind_direction}") is None:
                        ms_result_x = my_ms(shileding_category, 10, 10, 5)
                    else :
                        ms_result_x = st.session_state[f"ms_result_{selected_wind_direction}"]
                    my_ms_modal(selected_wind_direction)
                    st.session_state[ms_modal_flag_x] = False
                elif shileding_category == "Shielding" and st.session_state[ms_modal_flag_x] == False :
                    if st.session_state.get(f"ms_result_{selected_wind_direction}") is None:
                        st.error("Please input shielding data and click [Calculate] button.")
                        st.stop()
                    else :
                        ms_result_x = st.session_state[f"ms_result_{selected_wind_direction}"]
                else :
                    ms_result_x = my_ms(shileding_category, 10, 10, 5)


                # Parameters for Mh
                mh_modal_flag_x = "show_modal_mh_"+selected_wind_direction
                if mh_modal_flag_x not in st.session_state:
                    st.session_state[mh_modal_flag_x] = False
                col1_9, col1_10 = st.columns(2, vertical_alignment ="center")
                with col1_9 :
                    st.markdown("M<sub>h</sub>", unsafe_allow_html=True)
                with col1_10 :
                    hill_type_x = st.selectbox("Hill Type", my_hill_type["hill_type"], key="hill_type_x", label_visibility="collapsed", on_change=lambda: st.session_state.update({mh_modal_flag_x: True}))

                if hill_type_x != my_hill_type["hill_type"][0] and st.session_state[mh_modal_flag_x] == True :
                    mh_result_x = my_mh(my_hill_type["hill_type"][0])
                    my_mh_modal(hill_type_x, selected_wind_direction)
                    st.session_state[mh_modal_flag_x] = False
                elif hill_type_x != my_hill_type["hill_type"][0] and st.session_state[mh_modal_flag_x] == False :
                    if st.session_state.get(f"mh_result_{selected_wind_direction}") is None:
                        st.error("Please input hill data and click [Calculate] button.")
                        st.stop()
                    else :
                        mh_result_x = st.session_state[f"mh_result_{selected_wind_direction}"]
                else :
                    mh_result_x = my_mh(hill_type_x)
                    
                # Parameters for Vdes (Importance factor)
                col1_11, col1_12 = st.columns(2, vertical_alignment ="center")
                with col1_11 :
                    st.write("Importance Factor")
                with col1_12 :
                    importance_options = [f"{val:.2f}" for val in my_importance_factor["I"]]
                    importance_factor_visible = importance_options[:-1]
                    importance_factor = float(st.selectbox("Importance Factor", importance_factor_visible, key="importance_factor", index=1, label_visibility="collapsed"))

            st.info("Parameters for Cfig")
            with st.container(border=False) :
                # col1_21, col1_22 = st.columns(2, vertical_alignment ="center")
                # with col1_21 :
                #     st.markdown("ρ<sub>air</sub>", unsafe_allow_html=True)
                # with col1_22 :
                #     rho_air = st.number_input("ρ_air", value=my_rho_air, key="rho_air", label_visibility="collapsed", disabled=True)

                # Calculate Cfig for enclosed RC Building
                # Calculate Cpe_windward
                col1_23, col1_24 = st.columns(2, vertical_alignment ="center")
                with col1_23 :
                    st.markdown("C<sub>pe</sub>(windward)", unsafe_allow_html=True)
                with col1_24 :
                    cpe_windward_options = [f"{val:.2f}" for val in my_cpe_ww["Cpe_windward"]]
                    input_cpe_ww = float(st.selectbox("Cpe_windward", cpe_windward_options, key="cpe_windward_x", label_visibility="collapsed"))
                
                # Calculate Cpe_leeward
                col1_25, col1_26 = st.columns(2, vertical_alignment ="center")
                with col1_25 :
                    st.markdown("C<sub>pe</sub>(leeward)", unsafe_allow_html=True)
                with col1_26 :
                    st.number_input("Cpe_leeward", my_cpe_leeward(wind_direction=selected_wind_direction)["Cpe_lw"][0], key="cpe_leeward_x", label_visibility="collapsed", disabled=True)
                    input_cpe_lw = my_cpe_leeward(wind_direction=selected_wind_direction)["Cpe_lw"]

                # # Calculate Ka
                # col1_27, col1_28 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_27 :
                #     st.markdown("K<sub>a</sub>", unsafe_allow_html=True)
                # with col1_28 :
                #     input_ka = st.number_input("Ka", 1.0, key="ka_x", disabled=True, label_visibility="collapsed")
                #     ka = float(input_ka) # Area Reduction Factor

                # # Calculate Kc (Combination Factor)
                # col1_29, col1_30 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_29 :
                #     st.markdown("K<sub>c</sub>", unsafe_allow_html=True)
                # with col1_30 :
                #     kc_options = [f"{val:.2f}" for val in my_kc["kc"]]
                #     input_kc = st.selectbox("Kc", kc_options, key="kc_x", label_visibility="collapsed")
                #     kc = float(input_kc)

                # # Calculate Kl (Local Pressure factor)
                # col1_31, col1_32 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_31 :
                #     st.markdown("K<sub>l</sub>", unsafe_allow_html=True)
                # with col1_32 :
                #     input_kl = st.number_input("Kl", 1.0, step=0.1, key="kl_x", label_visibility="collapsed")
                #     kl = float(input_kl)

                # # Calculate Kp
                # col1_33, col1_34 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_33 :
                #     st.markdown("K<sub>p</sub>", unsafe_allow_html=True)
                # with col1_34 :
                #     input_kp = st.number_input("Kp", 1.0, step=0.1, key="kp_x", label_visibility="collapsed")
                #     kp = float(input_kp)
                
                # Input Kfactor = Ka * Kc * Kl * Kp
                kfactor_help = """
                Kfactor = Ka * Kc * Kl * Kp
                """

                col1_35, col1_36= st.columns([1, 1], vertical_alignment ="top")
                with col1_35:
                    # st.markdown("K<sub>factor</sub>", unsafe_allow_html=True)
                    # st.markdown("K<sub>factor</sub>", help=kfactor_help, unsafe_allow_html=True)
                    st.markdown(
                        f"""
                        <div style="display: flex; align-items: center; font-size: 12px;">
                            <span>K<sub>factor</sub></span>
                            <span style="margin-left: 5px; cursor: pointer;" title="{kfactor_help}">❔</span>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                with col1_36 :
                    k_factor_x = st.number_input("Kfactor", 1.0, step=0.1, key="k_factor_x", label_visibility="collapsed")
                
                # Calculate Cfig
                # k_factor = ka * kc * kl * kp
                # cfig_list_ww = [input_cpe_ww * k_factor for cpe_lw in input_cpe_lw]
                # cfig_list_lw = [cpe_lw * k_factor for cpe_lw in input_cpe_lw]
                cfig_list_ww = [input_cpe_ww * k_factor_x for cpe_lw in input_cpe_lw]
                cfig_list_lw = [cpe_lw * k_factor_x for cpe_lw in input_cpe_lw]

            st.info("Parameters for wind pressure(p)")
            with st.container(border=False) :
                # Input Cdyn
                col1_41, col1_42 = st.columns([1, 1], vertical_alignment ="center")
                with col1_41 :
                    st.markdown("C<sub>dyn</sub>", unsafe_allow_html=True)
                with col1_42 :
                    cdyn = st.number_input("Cdyn", 1.0, step=0.1, key="cdyn_x", label_visibility="collapsed")

                # 중간 결과 테이블 테이블 계산
                data_merge_dict_x = {
                    "Story" : st.session_state["story_data"]["story_name"],
                    "Vs" : st.session_state["vs"],
                    **md_dict,
                    **mzcat_result_x,
                    **ms_result_x,
                    "Mh(ww)" : mh_result_x["Mh(ww)"],
                    "Mh(lw)" : mh_result_x["Mh(lw)"]
                }
                param_merge_df_x = pd.DataFrame(data_merge_dict_x)

                # ********** Vsit 계산 **********
                param_merge_df_x["Vsit(ww)"] = param_merge_df_x.drop(columns=["Story", "Mh(lw)"]).product(axis=1)

                param_merge_df_x["Vsit(lw)"] = param_merge_df_x.drop(columns=["Story", "Mh(ww)", "Vsit(ww)"]).product(axis=1)

                # Importance Factor 추가
                param_merge_df_x["I"] = importance_factor

                # ********** Vdes 계산 **********
                param_merge_df_x["Vdes(ww)"] = param_merge_df_x[["Vsit(ww)", "I"]].product(axis=1)

                param_merge_df_x["Vdes(lw)"] = param_merge_df_x[["Vsit(lw)", "I"]].product(axis=1)

                # ********** pressure 계산 **********
                # Cfig 계산
                param_merge_df_x["Cfig(ww)"] = cfig_list_ww
                param_merge_df_x["Cfig(lw)"] = cfig_list_lw
             

                # Cdyn 계산
                repeat_cnt = len(st.session_state["story_data"]["story_name"])
                param_merge_df_x["Cdyn"] = copy_value_by_story(repeat_cnt, cdyn)

                # pressure 계산
                param_merge_df_x["p(ww)(Pa)"] = 0.5 * my_rho_air * param_merge_df_x["Vdes(ww)"]**2 * cfig_list_ww * param_merge_df_x["Cdyn"]
                param_merge_df_x["p(lw)(Pa)"] = 0.5 * my_rho_air * param_merge_df_x["Vdes(lw)"]**2 * cfig_list_lw * param_merge_df_x["Cdyn"]
                param_merge_df_x["p(kPa)"] = np.maximum(0.65, (param_merge_df_x["p(ww)(Pa)"] - param_merge_df_x["p(lw)(Pa)"]) / 1000)

                # # 중간 결과 테이블 테이블 출력
                # param_df_final = st.dataframe(param_merge_df_x, hide_index=True)

            # Story 결과 테이블 출력
            wind_calc_df_x = pd.DataFrame({})
            wind_calc_df_x["STORY"] = st.session_state["story_data"]["story_name"]
            wind_calc_df_x["ELEV [m]"] = st.session_state["story_data"]["story_elev"]
            wind_calc_df_x["p [kPa]"] = param_merge_df_x["p(kPa)"]
            wind_calc_df_x["Loaded_H [m]"] = st.session_state["story_data"]["loaded_h"]
            wind_calc_df_x["Loaded_Bx [m]"] = st.session_state["story_data"]["loaded_bx"]
            wind_calc_df_x["Loaded_By [m]"] = st.session_state["story_data"]["loaded_by"]
            
        with c2 :
            st.info("Graphs and Tables")
            if st.session_state["stor_valid_flag"] == True :
                tab_graph, tab_table = st.tabs(["Graph", "Table"])
                with tab_table :
                    wind_calc_df_x["Story Force [kN]"] = param_merge_df_x["p(kPa)"] * wind_calc_df_x["Loaded_H [m]"].astype(float) * wind_calc_df_x["Loaded_Bx [m]"].astype(float)
                    wind_calc_df_x["Story Force [kN]"].iloc[-1] = 0
                    
                    # 제품과 동일한 로직으로 shear force 계산 (해당 층보다 위에 해당하는 층의 story force 합)
                    story_shear_list = []
                    for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                        if story_idx == 0 :
                            story_shear_list.append(0)
                        else :
                            story_shear_list.append(sum(wind_calc_df_x["Story Force [kN]"][0:story_idx]))
                    wind_calc_df_x["Story Shear [kN]"] = story_shear_list

                    # # 해당 층 위에 해당하는 모든 면적에 대한 force 합으로 shear force 계산 로직
                    # story_height_list = []
                    # story_force_for_shear = []
                    # for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                    #     if story_idx == 0 :
                    #         story_height_list.append(0)
                    #         story_force_temp = float(param_merge_df_x["p(Pa)"][story_idx]/1000 * pd.Series(st.session_state["story_data"]["loaded_bx"][story_idx]).astype(float)/1000 * story_height_list[story_idx]/1000)
                    #         story_force_for_shear.append(story_force_temp)
                    #     else :
                    #         story_height_list.append(float(st.session_state["story_data"]["story_elev"][story_idx-1]) - float(st.session_state["story_data"]["story_elev"][story_idx]))
                    #         story_force_temp = float(param_merge_df_x["p(Pa)"][story_idx]/1000 * pd.Series(st.session_state["story_data"]["loaded_bx"][story_idx]).astype(float)/1000 * story_height_list[story_idx]/1000)
                    #         story_force_for_shear.append(story_force_temp)
                    # wind_calc_df_x["story hegiht"] = story_height_list
                    # wind_calc_df_x["Story Shear [kN]"] = story_force_for_shear
                    
                    story_height_list = []
                    story_overturning_moment = []
                    for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                        if story_idx == 0 :
                            story_height_list.append(0)
                            story_overturning_moment.append(0)
                        else :
                            overturning_moment_temp_list = []
                            for i in range(story_idx) :
                                height = wind_calc_df_x["ELEV [m]"].astype(float).iloc[i] - wind_calc_df_x["ELEV [m]"].astype(float).iloc[story_idx]
                                overturning_moment_temp_list.append(wind_calc_df_x["Story Force [kN]"].iloc[i] * height)
                            story_overturning_moment.append(sum(overturning_moment_temp_list))
                    wind_calc_df_x["Overturning M [kN·m]"] = story_overturning_moment

                    st.dataframe(data=wind_calc_df_x, height=465, row_height=25 ,hide_index=True, use_container_width=True ,column_config={
                        "ELEV [m]": st.column_config.NumberColumn("ELEV [m]", format="%.2f"),
                        "p [kPa]": st.column_config.NumberColumn("p [kPa]", format="%.2f"),
                        "Loaded_H [m]": st.column_config.NumberColumn("Loaded_H [m]", format="%.2f"),
                        "Loaded_Bx [m]": st.column_config.NumberColumn("Loaded_Bx [m]", format="%.2f"),
                        "Loaded_By [m]": st.column_config.NumberColumn("Loaded_By [m]", format="%.2f"),
                        "Story Force [kN]": st.column_config.NumberColumn("Story Force [kN]", format="%.2f"),
                        "Story Shear [kN]": st.column_config.NumberColumn("Story Shear [kN]", format="%.2f"),
                        "Overturning M [kN·m]": st.column_config.NumberColumn("Overturning M [kN·m]", format="%.2f"),
                    })

                with tab_graph :
                    col3, col_4 = st.columns(2, vertical_alignment="center")
                    with col3 :
                        st.write("Select Graph Type")
                    with col_4 :
                        graph_type_x = st.selectbox("Graph Type", ["Story Force", "Story Shear", "Overturning Moment"], key="graph_type_x", label_visibility="collapsed")
                    # 최종 결과 그래프
                    story_force_df_x = wind_calc_df_x[["STORY", "Story Force [kN]"]].copy()
                    story_order = list(story_force_df_x["STORY"])
                    story_force_df_x.columns = ["Story", "Story_Force"]
                    story_force_df_x["Story"] = pd.Categorical(story_force_df_x["Story"], categories=story_order, ordered=True)

                    story_shear_df_x = wind_calc_df_x[["STORY", "Story Shear [kN]"]].copy()
                    story_shear_df_x.columns = ["Story", "Story_Shear"]
                    story_shear_df_x["Story"] = pd.Categorical(story_shear_df_x["Story"], categories=story_order, ordered=True)

                    over_turning_moment_df_x = wind_calc_df_x[["STORY", "Overturning M [kN·m]"]].copy()
                    over_turning_moment_df_x.columns = ["Story", "Over_Turning_M"]
                    over_turning_moment_df_x["Story"] = pd.Categorical(over_turning_moment_df_x["Story"], categories=story_order, ordered=True)
                    
                    if graph_type_x == "Story Force" :
                        # 그래프 그리기
                        chart = bar_chart_hor(story_force_df_x, "[kN]")
                        st.altair_chart(chart, use_container_width=True)
                    elif graph_type_x == "Story Shear" :
                        # 그래프 그리기
                        chart = bar_chart_hor(story_shear_df_x, "[kN]")
                        st.altair_chart(chart, use_container_width=True)
                    else :
                        # 그래프 그리기
                        chart = bar_chart_hor(over_turning_moment_df_x, "[kN·m]")
                        st.altair_chart(chart, use_container_width=True)
            elif st.session_state["stor_valid_flag"] == None :
                st.warning("Please check the story data.", icon="❗")

    with tab_y :
        selected_wind_direction = my_wind_direction["wind_direction"][1]       # "y_dir"
        c1, c2 = st.columns([1, 1.5])
        # 왼쪽 화면 설정
        with c1 : 
            # Parameters for Vdes
            st.info("Parameters for Vdes")
            with st.container(border=False) :

                # Input Vs
                col1_1, col1_2 = st.columns(2, vertical_alignment ="center")
                with col1_1 :
                    st.markdown("V<sub>s</sub>(m/s)", unsafe_allow_html=True)
                with col1_2 :
                    st.number_input("Vs", value=st.session_state["vs"], label_visibility="collapsed", disabled=True)

                # Calculate Md
                col1_3, col1_4 = st.columns(2, vertical_alignment ="center")
                with col1_3 :
                    st.markdown("M<sub>d</sub>", unsafe_allow_html=True)
                with col1_4 :
                    md = st.number_input("Md", st.session_state["md"], format="%0.2f", label_visibility="collapsed", disabled=True)

                md_dict = {}
                md_list = []
                for i in st.session_state["story_data"]["story_elev"] :
                    md_list.append(my_md)
                md_dict["Md"] = md_list

                # Calculate Mz,cat
                # Flag 추가
                mzcat_modal_falg_y = "show_modal_mzcat_"+selected_wind_direction
                if mzcat_modal_falg_y not in st.session_state:
                  st.session_state[mzcat_modal_falg_y] = False
                  
                col1_5, col1_6 = st.columns(2, vertical_alignment ="center")
                with col1_5 :
                    st.markdown("M<sub>z,cat</sub>", unsafe_allow_html=True)
                with col1_6 :
                    terrain_cat = st.selectbox("Terrain Category", my_terrain_category["Category"], key=f"terrain_cat_drop_single_{selected_wind_direction}" ,label_visibility="collapsed", on_change=lambda: st.session_state.update({mzcat_modal_falg_y: True}))

                if terrain_cat == my_terrain_category["Category"][4] and st.session_state[mzcat_modal_falg_y] == True : # Select Multiple
                    if st.session_state.get(f"mzcat_result_{selected_wind_direction}") is None:
                        mzcat_result_y = my_mzcat([my_terrain_category["Category"][0]])
                    else :
                        mzcat_result_y = st.session_state[f"mzcat_result_{selected_wind_direction}"]
                    my_mzcat_modal(selected_wind_direction)
                    st.session_state[mzcat_modal_falg_y] = False
                elif terrain_cat == my_terrain_category["Category"][4] and st.session_state[mzcat_modal_falg_y] == False :
                    if st.session_state.get(f"mzcat_result_{selected_wind_direction}") is None:
                        st.error("Please input terrain category data and click [Calculate] button.")
                        st.stop()
                    else :
                        mzcat_result_y = st.session_state[f"mzcat_result_{selected_wind_direction}"]
                else :
                    mzcat_result_y = my_mzcat([st.session_state[f"terrain_cat_drop_single_{selected_wind_direction}"]])
                
                # Parameters for Ms
                ms_modal_flag_y = "show_modal_ms_"+selected_wind_direction
                if ms_modal_flag_y not in st.session_state:
                    st.session_state[ms_modal_flag_y] = False
                col1_7, col1_8 = st.columns(2, vertical_alignment ="center")
                with col1_7 :
                    st.markdown("M<sub>s</sub>", unsafe_allow_html=True)
                with col1_8 :
                    shileding_category = st.selectbox("Shielding Category", my_shileding_category["shielding_catetory"], key="shielding_catetory_y", label_visibility="collapsed", on_change=lambda: st.session_state.update({ms_modal_flag_y: True}))

                if shileding_category == "Shielding" and st.session_state[ms_modal_flag_y] == True :
                    if st.session_state.get(f"ms_result_{selected_wind_direction}") is None:
                        ms_result_y = my_ms(shileding_category, 10, 10, 5)
                    else :
                        ms_result_y = st.session_state[f"ms_result_{selected_wind_direction}"]
                    my_ms_modal(selected_wind_direction)
                    st.session_state[ms_modal_flag_y] = False
                elif shileding_category == "Shielding" and st.session_state[ms_modal_flag_y] == False :
                    if st.session_state.get(f"ms_result_{selected_wind_direction}") is None:
                        st.error("Please input shielding data and click [Calculate] button.")
                        st.stop()
                    else :
                        ms_result_y = st.session_state[f"ms_result_{selected_wind_direction}"]
                else :
                    ms_result_y = my_ms(shileding_category, 10, 10, 5)


                # Parameters for Mh
                mh_modal_flag_y = "show_modal_mh_"+selected_wind_direction
                if mh_modal_flag_y not in st.session_state:
                    st.session_state[mh_modal_flag_y] = False
                col1_9, col1_10 = st.columns(2, vertical_alignment ="center")
                with col1_9 :
                    st.markdown("M<sub>h</sub>", unsafe_allow_html=True)
                with col1_10 :
                    hill_type_y = st.selectbox("Hill Type", my_hill_type["hill_type"], key="hill_type_y", label_visibility="collapsed", on_change=lambda: st.session_state.update({mh_modal_flag_y: True}))

                if hill_type_y != my_hill_type["hill_type"][0] and st.session_state[mh_modal_flag_y] == True :
                    if st.session_state.get(f"mh_result_{selected_wind_direction}") is None:
                        mh_result_y = my_mh(my_hill_type["hill_type"][0])
                    else :
                        mh_result_y = st.session_state[f"mh_result_{selected_wind_direction}"]
                    my_mh_modal(hill_type_y, selected_wind_direction)
                    st.session_state[mh_modal_flag_y] = False
                elif hill_type_y != my_hill_type["hill_type"][0] and st.session_state[mh_modal_flag_y] == False :
                    if st.session_state.get(f"mh_result_{selected_wind_direction}") is None:
                        st.error("Please input hill data and click [Calculate] button.")
                        st.stop()
                    else :
                        mh_result_y = st.session_state[f"mh_result_{selected_wind_direction}"]
                else :
                    mh_result_y = my_mh(hill_type_y)
                    
                # Parameters for Vdes (Importance factor)
                col1_11, col1_12 = st.columns(2, vertical_alignment ="center")
                with col1_11 :
                    st.write("Importance Factor")
                with col1_12 :
                    importance_factor = float(st.selectbox("Importance Factor", st.session_state["importance_factor"], disabled=True, label_visibility="collapsed"))

            st.info("Parameters for Cfig")
            with st.container(border=False) :
                # col1_21, col1_22 = st.columns(2, vertical_alignment ="center")
                # with col1_21 :
                #     st.markdown("ρ<sub>air</sub>", unsafe_allow_html=True)
                # with col1_22 :
                #     rho_air = st.number_input("ρ_air", value=st.session_state["rho_air"], label_visibility="collapsed", disabled=True)

                # Calculate Cfig for enclosed RC Building
                # Calculate Cpe_windward
                col1_23, col1_24 = st.columns(2, vertical_alignment ="center")
                with col1_23 :
                    st.markdown("C<sub>pe</sub>(windward)", unsafe_allow_html=True)
                with col1_24 :
                    cpe_ww_options = [f"{val:.2f}" for val in my_cpe_ww["Cpe_windward"]]
                    input_cpe_ww = float(st.selectbox("Cpe_windward", cpe_ww_options, key="cpe_windward_y", label_visibility="collapsed"))
                
                # Calculate Cpe_leeward
                col1_25, col1_26 = st.columns(2, vertical_alignment ="center")
                with col1_25 :
                    st.markdown("C<sub>pe</sub>(leeward)", unsafe_allow_html=True)
                with col1_26 :
                    st.number_input("Cpe_leeward", my_cpe_leeward(wind_direction=selected_wind_direction)["Cpe_lw"][0], key="cpe_leeward_y", label_visibility="collapsed", disabled=True)
                    input_cpe_lw = my_cpe_leeward(wind_direction=selected_wind_direction)["Cpe_lw"]

                # # Calculate Ka
                # col1_27, col1_28 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_27 :
                #     st.markdown("K<sub>a</sub>", unsafe_allow_html=True)
                # with col1_28 :
                #     input_ka = st.number_input("Ka", 1.0, key="ka_y", disabled=True, label_visibility="collapsed")
                #     ka = float(input_ka) # Area Reduction Factor

                # # Calculate Kc (Combination Factor)
                # col1_29, col1_30 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_29 :
                #     st.markdown("K<sub>c</sub>", unsafe_allow_html=True)
                # with col1_30 :
                #     kc_options = [f"{val:.2f}" for val in my_kc["kc"]]
                #     input_kc = st.selectbox("Kc", kc_options, key="kc_y", label_visibility="collapsed")
                #     kc = float(input_kc)

                # # Calculate Kl (Local Pressure factor)
                # col1_31, col1_32 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_31 :
                #     st.markdown("K<sub>l</sub>", unsafe_allow_html=True)
                # with col1_32 :
                #     input_kl = st.number_input("Kl", 1.0, step=0.1, key="kl_y", label_visibility="collapsed")
                #     kl = float(input_kl)

                # # Calculate Kp
                # col1_33, col1_34 = st.columns([1, 1], vertical_alignment ="center")
                # with col1_33 :
                #     st.markdown("K<sub>p</sub>", unsafe_allow_html=True)
                # with col1_34 :
                #     input_kp = st.number_input("Kp", 1.0, step=0.1, key="kp_y", label_visibility="collapsed")
                #     kp = float(input_kp)

                # Input Kfactor = Ka * Kc * Kl * Kp
                kfactor_markdown = """
                Kfactor = Ka * Kc * Kl * Kp
                """
                col1_35, col1_36 = st.columns([1, 1], vertical_alignment ="top")
                with col1_35 :
                    # st.markdown("K<sub>factor</sub>", help=kfactor_markdown, unsafe_allow_html=True)
                    st.markdown(
                        f"""
                        <div style="display: flex; align-items: center; font-size: 12px;">
                            <span>K<sub>factor</sub></span>
                            <span style="margin-left: 5px; cursor: pointer;" title="{kfactor_help}">❔</span>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                with col1_36 :
                    k_factor_y = st.number_input("Kfactor", 1.0, step=0.1, key="k_factor_y", label_visibility="collapsed")
                
                # Calculate Cfig
                # k_factor = ka * kc * kl * kp
                # cfig_list_ww = [input_cpe_ww * k_factor for cpe_lw in input_cpe_lw]
                # cfig_list_lw = [cpe_lw * k_factor for cpe_lw in input_cpe_lw]
                cfig_list_ww = [input_cpe_ww * k_factor_y for cpe_lw in input_cpe_lw]
                cfig_list_lw = [cpe_lw * k_factor_y for cpe_lw in input_cpe_lw]

            st.info("Parameters for wind pressure(p)")
            with st.container(border=False) :
                # Input Cdyn
                col1_41, col1_42 = st.columns([1, 1], vertical_alignment ="center")
                with col1_41 :
                    st.markdown("C<sub>dyn</sub>", unsafe_allow_html=True)
                with col1_42 :
                    cdyn = st.number_input("Cdyn", 1.0, step=0.1, key="cdyn_y", label_visibility="collapsed")

                # 중간 결과 테이블 테이블 계산
                # if mzcat_result is not None :
                data_merge_dict_y = {
                    "Story" : st.session_state["story_data"]["story_name"],
                    "Vs" : st.session_state["vs"],
                    **md_dict,
                    **mzcat_result_y,
                    **ms_result_y,
                    "Mh(ww)" : mh_result_y["Mh(ww)"],
                    "Mh(lw)" : mh_result_y["Mh(lw)"]
                }
                param_merge_df_y = pd.DataFrame(data_merge_dict_y)

                # ********** Vsit 계산 **********
                param_merge_df_y["Vsit(ww)"] = param_merge_df_y.drop(columns=["Story", "Mh(lw)"]).product(axis=1)

                param_merge_df_y["Vsit(lw)"] = param_merge_df_y.drop(columns=["Story", "Mh(ww)", "Vsit(ww)"]).product(axis=1)

                # Importance Factor 추가
                param_merge_df_y["I"] = importance_factor

                # ********** Vdes 계산 **********
                param_merge_df_y["Vdes(ww)"] = param_merge_df_y[["Vsit(ww)", "I"]].product(axis=1)

                param_merge_df_y["Vdes(lw)"] = param_merge_df_y[["Vsit(lw)", "I"]].product(axis=1)

                # ********** pressure 계산 **********
                # Cfig 계산
                param_merge_df_y["Cfig(ww)"] = cfig_list_ww
                param_merge_df_y["Cfig(lw)"] = cfig_list_lw
             

                # Cdyn 계산
                repeat_cnt = len(st.session_state["story_data"]["story_name"])
                param_merge_df_y["Cdyn"] = copy_value_by_story(repeat_cnt, cdyn)

                # pressure 계산
                param_merge_df_y["p(ww)(Pa)"] = 0.5 * my_rho_air * param_merge_df_y["Vdes(ww)"]**2 * cfig_list_ww * param_merge_df_y["Cdyn"]
                param_merge_df_y["p(lw)(Pa)"] = 0.5 * my_rho_air * param_merge_df_y["Vdes(lw)"]**2 * cfig_list_lw * param_merge_df_y["Cdyn"]
                param_merge_df_y["p(kPa)"] = np.maximum(0.65, (param_merge_df_y["p(ww)(Pa)"] - param_merge_df_y["p(lw)(Pa)"]) / 1000)

                # # 중간 결과 테이블 테이블 출력
                # param_df_final = st.dataframe(param_merge_df_y, hide_index=True)

            # Story 결과 테이블 출력
            wind_calc_df_y = pd.DataFrame({})
            wind_calc_df_y["STORY"] = st.session_state["story_data"]["story_name"]
            wind_calc_df_y["ELEV [m]"] = st.session_state["story_data"]["story_elev"]
            wind_calc_df_y["p [kPa]"] = param_merge_df_y["p(kPa)"]
            wind_calc_df_y["Loaded_H [m]"] = st.session_state["story_data"]["loaded_h"]
            wind_calc_df_y["Loaded_Bx [m]"] = st.session_state["story_data"]["loaded_bx"]
            wind_calc_df_y["Loaded_By [m]"] = st.session_state["story_data"]["loaded_by"]

        with c2 :
            st.info("Graphs and Tables")
            if st.session_state["stor_valid_flag"] == True :
                tab_graph, tab_table = st.tabs(["Graph", "Table"])
                with tab_table :
                    wind_calc_df_y["Story Force [kN]"] = param_merge_df_y["p(kPa)"] * wind_calc_df_y["Loaded_H [m]"].astype(float) * wind_calc_df_y["Loaded_By [m]"].astype(float)
                    wind_calc_df_y["Story Force [kN]"].iloc[-1] = 0
                    
                    # 제품과 동일한 로직으로 shear force 계산 (해당 층보다 위에 해당하는 층의 story force 합)
                    story_shear_list = []
                    for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                        if story_idx == 0 :
                            story_shear_list.append(0)
                        else :
                            story_shear_list.append(sum(wind_calc_df_y["Story Force [kN]"][0:story_idx]))
                    wind_calc_df_y["Story Shear [kN]"] = story_shear_list
                    
                    story_height_list = []
                    story_overturning_moment = []
                    for story_idx in range(len(st.session_state["story_data"]["story_name"])) :
                        if story_idx == 0 :
                            story_height_list.append(0)
                            story_overturning_moment.append(0)
                        else :
                            overturning_moment_temp_list = []
                            for i in range(story_idx) :
                                height = wind_calc_df_y["ELEV [m]"].astype(float).iloc[i] - wind_calc_df_y["ELEV [m]"].astype(float).iloc[story_idx]
                                overturning_moment_temp_list.append(wind_calc_df_y["Story Force [kN]"].iloc[i] * height)
                            story_overturning_moment.append(sum(overturning_moment_temp_list))
                    wind_calc_df_y["Overturning M [kN·m]"] = story_overturning_moment

                    st.dataframe(data=wind_calc_df_y, height=465, row_height=25 ,hide_index=True, use_container_width=True ,column_config={
                        "ELEV [m]": st.column_config.NumberColumn("ELEV [m]", format="%.2f"),
                        "p [kPa]": st.column_config.NumberColumn("p [kPa]", format="%.2f"),
                        "Loaded_H [m]": st.column_config.NumberColumn("Loaded_H [m]", format="%.2f"),
                        "Loaded_Bx [m]": st.column_config.NumberColumn("Loaded_Bx [m]", format="%.2f"),
                        "Loaded_By [m]": st.column_config.NumberColumn("Loaded_By [m]", format="%.2f"),
                        "Story Force [kN]": st.column_config.NumberColumn("Story Force [kN]", format="%.2f"),
                        "Story Shear [kN]": st.column_config.NumberColumn("Story Shear [kN]", format="%.2f"),
                        "Overturning M [kN·m]": st.column_config.NumberColumn("Overturning M [kN·m]", format="%.2f"),
                    })

                with tab_graph :
                    col3, col_4 = st.columns(2, vertical_alignment="center")
                    with col3 :
                        st.write("Select Graph Type")
                    with col_4 :
                        graph_type_y = st.selectbox("Graph Type", ["Story Force", "Story Shear", "Overturning Moment"], key="graph_type_y", label_visibility="collapsed")
                    # 최종 결과 그래프
                    story_force_df_y = wind_calc_df_y[["STORY", "Story Force [kN]"]].copy()
                    story_order = list(story_force_df_y["STORY"])
                    story_force_df_y.columns = ["Story", "Story_Force"]
                    story_force_df_y["Story"] = pd.Categorical(story_force_df_y["Story"], categories=story_order, ordered=True)

                    story_shear_df_y = wind_calc_df_y[["STORY", "Story Shear [kN]"]].copy()
                    story_shear_df_y.columns = ["Story", "Story_Shear"]
                    story_shear_df_y["Story"] = pd.Categorical(story_shear_df_y["Story"], categories=story_order, ordered=True)

                    over_turning_moment_df_y = wind_calc_df_y[["STORY", "Overturning M [kN·m]"]].copy()
                    over_turning_moment_df_y.columns = ["Story", "Over_Turning_M"]
                    over_turning_moment_df_y["Story"] = pd.Categorical(over_turning_moment_df_y["Story"], categories=story_order, ordered=True)
                    
                    if graph_type_y == "Story Force" :
                        # 그래프 그리기
                        chart = bar_chart_hor(story_force_df_y, "[kN]")
                        st.altair_chart(chart, use_container_width=True)
                    elif graph_type_y == "Story Shear" :
                        # 그래프 그리기
                        chart = bar_chart_hor(story_shear_df_y, "[kN]")
                        st.altair_chart(chart, use_container_width=True)
                    else :
                        # 그래프 그리기
                        chart = bar_chart_hor(over_turning_moment_df_y, "[kN·m]")
                        st.altair_chart(chart, use_container_width=True)
            elif st.session_state["stor_valid_flag"] == None :
                st.warning("Please check the story data.", icon="❗")

    # 엑셀 다운로드 버튼
    col_button1, col_button2, col_button3 = st.columns(spec=[1,1,3], vertical_alignment ="center")
    with col_button1 :
        if st.button("Refresh", use_container_width=True, type="secondary") :
            try :
                with st.spinner("Updating Data..."):
                    header = {"MAPI-Key":_mapi_key}
                    story_data_body = {
                        "Argument": {
                            "FORCE_UNIT": "KN",
                            "LENGTH_UNIT": "M"
                        }
                    }
                    get_story_data = get_stor_data(body = story_data_body)
                    st.session_state["story_data"] = get_story_data
                    
                    stld_data = get_stld_data()
                    st.session_state["stld_data"] = stld_data

            except Exception as e:
                st.toast(f"🔁 Refresh failed: {e}")
            st.rerun()

    with col_button2:
        if st.session_state["stor_valid_flag"] == True :
            # 엑셀 원본을 바이너리로 열되, BytesIO로 복사
            with open("./resources/ms_1553_2002/MS1553_2002.xlsx", "rb") as base_excel:
                in_memory_excel = io.BytesIO(base_excel.read())

            # openpyxl로 메모리상에서 workbook 로드
            wb = load_workbook(in_memory_excel)
            
            # 데이터 위치 지정
            start_row = 6
            start_col = 1

            # Parameter 입력 (X-Dir)
            if hill_type_x == "Escarpment" :
                param_merge_df_x_modified = param_merge_df_x.copy()
                param_merge_df_x_modified = param_merge_df_x_modified.drop(columns=["Vsit(ww)","Vsit(lw)", "Vdes(ww)", "Vdes(lw)", "p(ww)(Pa)", "p(lw)(Pa)", "p(kPa)"])

                # 필요없는 시트 삭제
                sheet_to_delete = wb["WindParameters X"]
                wb.remove(sheet_to_delete)

                # DataFrame 값 삽입
                ws = wb["WindParameters_Escarpment X"]
                dataframe_to_excel(ws, param_merge_df_x_modified, start_row, start_col)
                # for i, row in param_merge_df_x_modified.iterrows() :
                #     for j, value in enumerate(row) :
                #         cell = ws.cell(row=start_row+i, column=start_col+j)
                #         cell.value = value
                ws.title = "WindParameters X"

            else :
                param_merge_df_x_modified = param_merge_df_x.copy()
                param_merge_df_x_modified = param_merge_df_x_modified.drop(columns=["Mh(lw)", "Vsit(ww)","Vsit(lw)", "Vdes(ww)", "Vdes(lw)", "Cfig(lw)", "p(ww)(Pa)", "p(lw)(Pa)", "p(kPa)"])
                param_merge_df_x_modified["Cfig(ww)"] = param_merge_df_x["Cfig(ww)"] - param_merge_df_x["Cfig(lw)"]

                # 필요없는 시트 삭제
                sheet_to_delete = wb["WindParameters_Escarpment X"]
                wb.remove(sheet_to_delete)

                ws = wb["WindParameters X"]
                dataframe_to_excel(ws, param_merge_df_x_modified, start_row, start_col)
                # for i, row in param_merge_df_x_modified.iterrows() :
                #     for j, value in enumerate(row) :
                #         cell = ws.cell(row=start_row+i, column=start_col+j)
                #         cell.value = value

            # Parameter 입력 (Y-Dir)
            if hill_type_y == "Escarpment" :
                param_merge_df_y_modified = param_merge_df_y.copy()
                param_merge_df_y_modified = param_merge_df_y_modified.drop(columns=["Vsit(ww)","Vsit(lw)", "Vdes(ww)", "Vdes(lw)", "p(ww)(Pa)", "p(lw)(Pa)", "p(kPa)"])

                # 필요없는 시트 삭제
                sheet_to_delete = wb["WindParameters Y"]
                wb.remove(sheet_to_delete)

                # DataFrame 값 삽입 (Y-Dir)
                ws = wb["WindParameters_Escarpment Y"]
                dataframe_to_excel(ws, param_merge_df_y_modified, start_row, start_col)
                # for i, row in param_merge_df_y_modified.iterrows() :
                #     for j, value in enumerate(row) :
                #         cell = ws.cell(row=start_row+i, column=start_col+j)
                #         cell.value = value
                ws.title = "WindParameters Y"

            else :
                param_merge_df_y_modified = param_merge_df_y.copy()
                param_merge_df_y_modified = param_merge_df_y_modified.drop(columns=["Mh(lw)", "Vsit(ww)","Vsit(lw)", "Vdes(ww)", "Vdes(lw)", "Cfig(lw)", "p(ww)(Pa)", "p(lw)(Pa)", "p(kPa)"])
                param_merge_df_y_modified["Cfig(ww)"] = param_merge_df_y["Cfig(ww)"] - param_merge_df_y["Cfig(lw)"]

                # 필요없는 시트 삭제
                sheet_to_delete = wb["WindParameters_Escarpment Y"]
                wb.remove(sheet_to_delete)

                # DataFrame 값 삽입 (X-Dir)
                ws = wb["WindParameters Y"]
                dataframe_to_excel(ws, param_merge_df_y_modified, start_row, start_col)
                # for i, row in param_merge_df_y_modified.iterrows() :
                #     for j, value in enumerate(row) :
                #         cell = ws.cell(row=start_row+i, column=start_col+j)
                #         cell.value = value

            # Wind Load 계산 결과 삽입
            # DataFrame 값 삽입 (X-Dir)
            ws = wb["Wind Loads X"]
            dataframe_to_excel(ws, wind_calc_df_x, start_row, start_col)
            # for i, row in wind_calc_df_x.iterrows() :
            #     for j, value in enumerate(row) :
            #         cell = ws.cell(row=start_row+i, column=start_col+j)
            #         cell.value = value

            # DataFrame 값 삽입 (Y-Dir)
            ws = wb["Wind Loads Y"]
            dataframe_to_excel(ws, wind_calc_df_y, start_row, start_col)
            # for i, row in wind_calc_df_y.iterrows() :
            #     for j, value in enumerate(row) :
            #         cell = ws.cell(row=start_row+i, column=start_col+j)
            #         cell.value = value

            # 결과를 메모리로 저장
            output_excel = io.BytesIO()
            wb.save(output_excel)
            output_excel.seek(0)  # 파일 포인터를 처음으로 돌려야 다운로드 가능

            # 다운로드 버튼
            b64_excel = base64.b64encode(output_excel.getvalue()).decode()
            st.markdown(
                f"""
                <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}"
							download="MS1553_2002_WindLoads.xlsx"
							target="_parent"
							style="text-decoration: none;">
								<button style="
										background-color: #3498db;
										color: white;
										padding: 0.3rem 0.3rem;
										border: none;
										border-radius: 0.25rem;
										cursor: pointer;
										width: 100%;
										font-size: 1rem;">
										Download Excel Data
								</button>
						</a>
                """,
                unsafe_allow_html=True
            )
        else :
            # st.button("Download Excel Data", type="secondary", use_container_width=True, disabled=True)
            st.markdown(
                '''
                <button style="
                    background-color: #bdc3c7;
                    color: white;
                    padding: 0.3rem 0.3rem;
                    border: none;
                    border-radius: 0.25rem;
                    width: 100%;
                    font-size: 1rem;
                    cursor: not-allowed;"
                    disabled>
                    Download Excel Data
                </button>
                ''',
                unsafe_allow_html=True
            )
        
    with col_button3 :
        if st.session_state["stor_valid_flag"] == True and st.session_state["stld_valid_flag"] == True :
            apply_button = st.button("Apply", type="primary", use_container_width=True)
            if apply_button :
                select_stld_modal(st.session_state["stld_data"]["NAME"])
        else :
            st.button("Apply", type="primary", use_container_width=True, disabled=True)

    # st.write(param_merge_df_x)
    # st.write(wind_calc_df_x)
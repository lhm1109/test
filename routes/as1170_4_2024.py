import streamlit as st
import pandas as pd
import requests
from openpyxl import load_workbook
from io import BytesIO
import plotly.express as px
from st_aggrid import AgGrid,GridOptionsBuilder
import os
from urllib.parse import parse_qs
from dotenv import load_dotenv
import base64
from utils.ParamHandler import ParamHandler

# Load environment variables
load_dotenv()

# Set page configuration
st.set_page_config(page_title="AS 1170.4:2024", page_icon="", layout="wide")

# Hide header and remove top padding
st.markdown("""
    <style>
        header[data-testid="stHeader"] {
            display: none !important;
        }
        .main > div {
            padding-top: 0rem;
        }
        .block-container {
            padding-top: 0rem;
            padding-bottom: 0rem;
        }
        [data-testid="stToolbar"] {
            display: none !important;
        }
        [data-testid="stDecoration"] {
            display: none !important;
        }
        #MainMenu {
            display: none !important;
        }
        [data-testid="stStatusWidget"] {
            display: none !important;
        }
        footer {
            display: none !important;
        }
    </style>
""", unsafe_allow_html=True)

# --- URL 파라미터 읽기 ---
_base_url = ParamHandler.base_url
_mapi_key = ParamHandler.mapi_key
base_url = f"{_base_url}/gen"

DEFAULTS = {
    "stld_data": {},
    "unit_data": {},
    "story_df": pd.DataFrame(),
    "structure_height": 0.0,
    "api_connected": False,
    "calculated": False,
    "edc": "II",
    "t1_x": 1.0,
    "t1_y": 1.0,
    "period_calculated": False,
    "importance_level": "2",
    "soil_class": "Ae (Strong Rock)",
    "pa": "1/500",
    "Z": 0.08,
    "sp_x": 0.67,
    "sp_y": 0.67,
    "mu_x": 4.0,
    "mu_y": 4.0,
    "confirm_apply": False

}

for key, value in DEFAULTS.items():
    st.session_state.setdefault(key, value)

def get_session_param(param):
    """Get value from session state with error tracking."""
    return st.session_state.get(param)

def set_session_param(param, value):
    """Set value in session state with logging comment."""
    st.session_state[param] = value



def convert_to_meter(value, unit):
    unit = unit.upper()
    if unit == "MM":
        return value / 1000
    elif unit == "CM":
        return value / 100
    elif unit == "M":
        return value
    elif unit == "KM":
        return value * 1000
    else:
        raise ValueError(f"Unsupported distance unit: {unit}")

def calc_fundamental_period(hn, formula):
  if formula == "Moment Resisting Steel Frames (kt=0.11)":
      return round(0.1375 * (hn ** 0.75), 2)
  elif formula == "Moment Resisting Concrete Frames (kt=0.075)":
      return round(0.09375 * (hn ** 0.75), 2)
  elif formula == "Eccentrically Braced Frames (kt=0.06)":
      return round(0.075 * (hn ** 0.75), 2)
  elif formula == "All Other Structures (kt=0.05)":
      return round(0.0625 * (hn ** 0.75), 2)
  else:
      raise ValueError("Unknown formula selected.")


def calc_ch_t(t1, soil_class):
    """Calculate Ch(T) based on soil class and fundamental period."""

    if soil_class == "Ae (Strong Rock)":
        if t1 <= 0.1:
            return 2.35
        elif t1 <= 1.5:
            return min(0.704 / t1, 2.35)
        else:
            return 1.056 / (t1 ** 2)

    elif soil_class == "Be (Rock)":
        if t1 <= 0.1:
            return 2.94
        elif t1 <= 1.5:
            return min(0.88 / t1, 2.94)
        else:
            return 1.32 / (t1 ** 2)

    elif soil_class == "Ce (Shallow Soil)":
        if t1 <= 0.1:
            return 3.68
        elif t1 <= 1.5:
            return min(1.25 / t1, 3.68)
        else:
            return 1.874 / (t1 ** 2)

    elif soil_class == "De (Deep or Soft Soil)":
        if t1 <= 0.1:
            return 3.68
        elif t1 <= 1.5:
            return min(1.98 / t1, 3.68)
        else:
            return 1.874 / (t1 ** 2)

    else:  # 기타 (Default case)
        if t1 <= 0.1:
            return 3.68
        elif t1 <= 1.5:
            return min(3.08 / t1, 3.68)
        else:
            return 4.62 / (t1 ** 2)

def calc_kp(pa):
    return {
        "1/100": 0.5,
        "1/200": 0.7,
        "1/250": 0.8,
        "1/500": 1.0,
        "1/800": 1.2,
        "1/1000": 1.3,
        "1/1500": 1.4,
        "1/2000": 1.5,
        "1/2500": 1.6
    }.get(pa, 1.0)  # 기본값 1.0


##minimum value of kpZ
def calc_kpZ_2(kp, Z, pa):
    limit_values = {
        "1/500": 0.08,
        "1/1000": 0.10,
        "1/1500": 0.12,
        "1/2000": 0.14,
        "1/2500": 0.15,
    }
    return max(kp * Z, limit_values.get(pa, kp * Z))

def determine_kpZ(Z, pa):
    kp = calc_kp(pa)
    kpZ_1 = kp * Z
    kpZ_2 = calc_kpZ_2(kp, Z, pa)
    return max(kpZ_1, kpZ_2)

def calc_earthquake_design_category(importance_level, soil_class, Z, hn_meter, pa):
    kZ = determine_kpZ(Z, pa)
    categories = {
        "2": {
            "Ee (Very Soft Soil)": [(0.08, 12, "I"), (0.08, 50, "II"), (0.08, float("inf"), "III"),
                    (0.08, 50, "II"), (0.08, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "De (Deep or Soft Soil)": [(0.08, 12, "I"), (0.08, 50, "II"), (0.08, float("inf"), "III"),
                    (0.08, 50, "II"), (0.08, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "Ce (Shallow Soil)": [(0.08, 12, "I"), (0.08, 50, "II"), (0.08, float("inf"), "III"),
                    (0.12, 50, "II"), (0.12, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "Be (Rock)": [(0.11, 12, "I"), (0.11, 50, "II"), (0.11, float("inf"), "III"),
                    (0.17, 50, "II"), (0.17, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "Ae (Strong Rock)": [(0.14, 12, "I"), (0.14, 50, "II"), (0.14, float("inf"), "III"),
                    (0.21, 50, "II"), (0.21, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")]
        },
        "3": {
            "Ee (Very Soft Soil)": [(0.08, 50, "II"), (0.08, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "De (Deep or Soft Soil)": [(0.08, 50, "II"), (0.08, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "Ce (Shallow Soil)": [(0.12, 50, "II"), (0.12, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "Be (Rock)": [(0.17, 50, "II"), (0.17, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")],
            "Ae (Strong Rock)": [(0.21, 50, "II"), (0.21, float("inf"), "III"), (float("inf"), 25, "II"), (float("inf"), float("inf"), "III")]
        },
        "4": {
            "default": [(float("inf"), 12, "II"), (float("inf"), float("inf"), "III")]
        }
    }

    if importance_level == "4":
        for max_kz, max_hn, category in categories[importance_level]["default"]:
            if hn_meter <= max_hn:
                st.session_state.edc = category
                return category
        return "III"

    if importance_level not in categories or soil_class not in categories[importance_level]:
        return "Invalid input"

    for max_kz, max_hn, category in categories[importance_level][soil_class]:
        if kZ <= max_kz and hn_meter <= max_hn:
            st.session_state.edc = category
            return category
    st.session_state.edc = "III"
    return "III"

# 지진하중 계수 계산 함수
def calc_seismic_coefficient(ch_t, sp, mu, kpZ_value, Z, t1, edc="II"):
    edc = st.session_state.edc
    sp_mu = sp / mu
    if edc == "I":
        # st.toast("Earthquake Design Category is I, So, It calculated with Simple Static Check method.")
        return 0.1
    elif edc == "II":
        return kpZ_value * ch_t * sp_mu
    else:
        # st.toast("Earthquake Design Category is III, So, It can use Dynamic Analysis method. but, in this case, this plugin is out of range. but if you want,you can use Static analysis method.")
        return kpZ_value * ch_t * sp_mu

# k 값 계산 함수
def def_k_value(t1):
    if t1 <= 0.5:
        return 1
    elif t1 >= 2.5:
        return 2
    else:
        return 1 + (t1 - 0.5) * (2 - 1) / (2.5 - 0.5)

# EDC 설명 함수
def desc_edc(edc):
    if edc == "I":
        return "Simple Static"
    elif edc == "II":
        return "Static"
    else:
        return "Dynamic"

# 지진하중 계산 함수
def calculate_seismic_loads(df, t1, mu, sp, direction, soil_class, kpZ_value, Z):
    ch_t = calc_ch_t(t1, soil_class)
    cd = calc_seismic_coefficient(ch_t, sp, mu, kpZ_value, Z, t1)
    k = def_k_value(t1)
    df= df.copy() # 원본 데이터 보호
    df[f"Wi*Hi^k for {direction}"] = df["WEIGHT"] * df["ELEV"] ** k
    df[f"Story Force for {direction}"] = df[f"Wi*Hi^k for {direction}"] * df["WEIGHT"].sum() / df[f"Wi*Hi^k for {direction}"].sum() * cd
    df[f"Story Shear for {direction}"] = df[f"Story Force for {direction}"].cumsum().shift(1).fillna(0)
    df[f"Overturning Moment for {direction}"] = 0.0
    for i in range(1, len(df)):
        h_diff = df.loc[i - 1, "ELEV"] - df.loc[i, "ELEV"]
        shear_i = df.loc[i, f"Story Shear for {direction}"]
        moment_above = df.loc[i - 1, f"Overturning Moment for {direction}"]
        df.loc[i, f"Overturning Moment for {direction}"] = moment_above + shear_i * h_diff
    return df


# 그래프 그리기 함수
def plot_graph(df, column, direction, unit_force, unit_dist):
    formatted_text = df[column].apply(lambda x: f"{x:.2f}")

    if "Overturning Moment" in column:
        unit_label = f"{unit_force}·{unit_dist}"
    else:
        unit_label = f"{unit_force}"
    fig = px.bar(
        df,
        x=column,
        y="STORY",
        orientation="h",
        # title=f"{column} ({direction}-Direction)",
        labels={column: f"{column} [{unit_label}]", "STORY": "Story"},
        height=650,
        color_discrete_sequence=["#3498db"]  # 하늘색
    )
    # 막대 안쪽에 값 표시
    fig.update_traces(
        text=formatted_text,
        textposition='inside',
        textfont=dict(color='white', size=12)  # 대비를 위해 흰색 글씨
    )
    fig.update_layout(yaxis=dict(autorange="reversed"))
    st.plotly_chart(fig, use_container_width=True)

# 사용자 데이터 생성 함수
def build_user_data(df, direction):
    return [
        {
            "STORY_NAME": row["STORY"],
            "WEIGHT": row["WEIGHT"],
            "ELIV": row["ELEV"],
            "FORCE_X": row[f"Story Force for {direction}"] if direction == "X" else 0,
            "FORCE_Y": row[f"Story Force for {direction}"] if direction == "Y" else 0
        }
        for _, row in df.iterrows()
    ]

# 엑셀 템플릿 채우기 함수
def fill_template_excel(base_excel_file, metadata, cd_x, cd_y, df_x, df_y, unit_force, unit_dist):
    # 1. 템플릿 파일 로드
    wb = load_workbook(base_excel_file)
    # 2. 시트 선택
    ws = wb["SeismicCalcSheet"]

    # 3. H6부터 metadata value들을 입력
    start_row = 6
    for i, value in enumerate(metadata.values()):
        cell = f"G{start_row + i}"
        ws[cell] = value
    ws["G19"] = cd_x
    ws["G20"] = cd_y
    ws["G21"] = df_x["Story Shear for X"].max()
    ws["G22"] = df_y["Story Shear for Y"].max()
    ws["G23"] = df_x["Wi*Hi^k for X"].sum()
    ws["G24"] = df_y["Wi*Hi^k for Y"].sum()
    # 단위 힘, 단위 거리
    ws["C28"] = unit_force
    ws["D28"] = unit_dist

     # ▶ Seismic Loads X 시트에 데이터 작성 (A6부터)
    sheet_x = wb["Seismic Loads X"]
    columns_x = ["STORY", "WEIGHT", "ELEV", "Story Force for X", "Story Shear for X", "Overturning Moment for X"]

    for row_idx, row in enumerate(df_x[columns_x].itertuples(index=False), start=6):  # row 6부터 데이터
        for col_idx, value in enumerate(row, start=1):  # col A부터
            sheet_x.cell(row=row_idx, column=col_idx).value = value
    # ▶ Seismic Loads Y 시트에 데이터 작성 (A6부터)
    sheet_y = wb["Seismic Loads Y"]
    columns_y = ["STORY", "WEIGHT", "ELEV", "Story Force for Y", "Story Shear for Y", "Overturning Moment for Y"]

    for row_idx, row in enumerate(df_y[columns_y].itertuples(index=False), start=6):  # 시작 위치: A6
        for col_idx, value in enumerate(row, start=1):
            sheet_y.cell(row=row_idx, column=col_idx).value = value

    # 4. 메모리 버퍼에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# EDC 계산 (I 여부에 따라 불필요한 입력 비활성화)
def is_category_i(edc):
    if edc == "I":
        return True
    else:
        return False
# EDC 계산
def update_edc():
    """EDC 값을 재계산하고 세션 상태를 업데이트합니다 (변경된 경우만 알림)."""
    unit_dist = st.session_state.unit_data.get("DIST", "M")
    hn_meter = convert_to_meter(st.session_state.structure_height, unit_dist)

    previous_edc = st.session_state.get("edc", "II")  # 기본값은 "II"

    # 새 EDC 계산
    new_edc = calc_earthquake_design_category(
        st.session_state.importance_level,
        st.session_state.soil_class,
        st.session_state.Z,
        hn_meter,
        st.session_state.pa
    )

    # 변경되었을 경우만 toast 알림
    if previous_edc != new_edc:
        st.toast(f"EDC updated: {previous_edc} → {new_edc}", icon="🔁")

    # 세션에 저장
    st.session_state.edc = new_edc

    # 디버깅용 로그
    print(f"EDC updated to: {new_edc}")

# 주기 계산 Dialog
@st.dialog("Calculate T1 using Eq.6.2(7)")
def calc_t1_using_equation():
    hn_raw = float(st.session_state.structure_height)
    unit_dist = st.session_state.unit_data.get("DIST", "M")
    hn_meter = convert_to_meter(hn_raw, unit_dist)
    st.info(f"Structure Height,hn = {hn_meter} [m]")

    # st.markdown("**Fundamental Period - X**")
    st.markdown("""
        <div style='border: 1px solid #ccc; padding: 10px; border-radius: 0.25rem; margin: 10px 0;'>
            <p style='margin: 0;'>Fundamental Period - X</p>
        </div>
    """, unsafe_allow_html=True)
    x_formula = st.radio(
        "Fundamental Period - X",
        options=["Moment Resisting Steel Frames (kt=0.11)",
                "Moment Resisting Concrete Frames (kt=0.075)",
                "Eccentrically Braced Frames (kt=0.06)",
                "All Other Structures (kt=0.05)"],
        index=0,
        key="x_direction_period",
        label_visibility="collapsed"
    )
    # Y 방향 주기 선택
    # st.markdown("**Fundamental Period - Y**")
    st.markdown("""
        <div style='border: 1px solid #ccc; padding: 10px; border-radius: 0.25rem; margin: 10px 0;'>
            <p style='margin: 0;'>Fundamental Period - Y</p>
        </div>
    """, unsafe_allow_html=True)
    y_formula = st.radio(
    "Fundamental Period - Y",
        options=["Moment Resisting Steel Frames (kt=0.11)",
                "Moment Resisting Concrete Frames (kt=0.075)",
                "Eccentrically Braced Frames (kt=0.06)",
                "All Other Structures (kt=0.05)"],
        index=0,
        key="y_direction_period",
        label_visibility="collapsed"
    )

    if st.button("Calculate", key="btn_calc_t_period", use_container_width=True, type="primary"):
        try:
            calc_x = calc_fundamental_period(hn_meter, x_formula)
            calc_y = calc_fundamental_period(hn_meter, y_formula)
            st.session_state.t1_x = calc_x
            st.session_state.t1_y = calc_y
            st.rerun()

        except Exception as e:
            st.warning(f"❗ Exception: {e}")

# EDC III 확인 Dialog
@st.dialog("Proceed with Static Load Application?")
def confirm_apply_edc_iii():
    st.warning("""\
    **Earthquake Design Category: III**

    Dynamic Analysis is recommended instead of Static Analysis.

    Do you want to proceed with applying static seismic loads?""")
    if st.button("✅ OK, Proceed", use_container_width=True):
        st.session_state.confirm_apply = True
        st.rerun()
    if st.button("❌ Cancel", use_container_width=True):
        st.session_state.confirm_apply = False
        st.rerun()

# API 호출 함수
def fetch(base_url, mapi_key):
    headers = {"MAPI-Key": mapi_key}

    # 1. STLD 요청
    res_stld = requests.get(f"{base_url}/db/stld", headers=headers)
    if res_stld.status_code != 200:
        raise Exception(f"STLD fetch failed: {res_stld.status_code}")

    # 2. STORYPROP 요청
    res_story = requests.post(f"{base_url}/ope/storyprop", headers=headers, json={"Argument": {}})
    if res_story.status_code != 200:
        raise Exception(f"STORYPROP fetch failed: {res_story.status_code}")

    # 3. UNIT 요청
    res_unit = requests.get(f"{base_url}/db/unit", headers=headers)
    if res_unit.status_code != 200:
        raise Exception(f"UNIT fetch failed: {res_unit.status_code}")

    # 정상적으로 데이터를 처리 후 반환
    stld_data = res_stld.json().get("STLD", {})
    story_data = res_story.json().get("STORYPROP", {}).get("DATA", [])
    unit_data = res_unit.json().get("UNIT", {}).get("1", {})

    return stld_data, story_data, unit_data


if __name__ == "__main__":

    if _base_url and _mapi_key and not st.session_state.api_connected:
        headers = {"MAPI-Key": _mapi_key}
        try:
            with st.spinner("Connecting to API..."):
                stld_data, story_data, unit_data = fetch(base_url, _mapi_key)
                st.session_state.api_connected = True
                st.session_state.stld_data = stld_data
                # 원본 저장
                st.session_state.story_data = story_data
                st.session_state.unit_data = unit_data
                # 파싱 및 변환
                df = pd.DataFrame(story_data)
                df["WEIGHT"] = df["WEIGHT"].astype(float)
                df["ELEV"] = df["ELEV"].astype(float)
                st.session_state.story_df = df
                st.session_state.structure_height = df["ELEV"].max()

                update_edc()


        except Exception as e:
            st.error(f"Connection failed: {e}")
            st.stop()

    if st.session_state.api_connected:
        ## 화면 설정
        # st.title("[AS 1170.4:2024] Static Seismic Loads Generator")
        story_df = st.session_state.story_df.copy()
        unit_force = st.session_state.unit_data.get("FORCE", "KN")
        unit_dist = st.session_state.unit_data.get("DIST", "M")

        c1, c2 = st.columns([1,1.5])
        # 왼쪽 화면 설정
        with c1:
            with st.container():
                # st.markdown("**Seismic Design Parameters**")
                # st.divider()
                st.info("Seismic Design Parameters")

                col1_1, col1_2 = st.columns(2)
                with col1_1:
                    st.write("Sub Soil Class")
                with col1_2:
                    soil_class = st.selectbox("Soil Class", ["Ae (Strong Rock)", "Be (Rock)", "Ce (Shallow Soil)", "De (Deep or Soft Soil)", "Ee (Very Soft Soil)"], key="soil_class",label_visibility="collapsed", on_change=update_edc)

                colBB_1, colBB_2 = st.columns(2)
                with colBB_1:
                    st.write("Annual Probability of Exceedance")
                with colBB_2:
                    pa = st.selectbox("Annual Probability of Exceedance", ["1/100","1/200","1/250","1/500","1/800", "1/1000", "1/1500", "1/2000", "1/2500"], key="pa",label_visibility="collapsed",on_change=update_edc)

                col1_7, col1_8 = st.columns(2)
                with col1_7:
                    st.markdown(
                        """
                        <span>Hazard Factor (Z)</span>
                        <a href="https://tools.midasuser.com/en/seismic-hazard-map" target="_blank" style="text-decoration: none; margin-left: 12px;">[🗺️Map]</a>
                        """,
                        unsafe_allow_html=True
                    )
                with col1_8:
                    Z = st.number_input("Hazard Factor (Z)", min_value=0.08, max_value=1.0, step=0.01,key="Z",label_visibility="collapsed", on_change=update_edc)

                colAA_1, colAA_2 = st.columns([1,1])
                with colAA_1:
                    st.write("Value of kpZ")

                with colAA_2:
                    Z = st.session_state.Z
                    pa = st.session_state.pa
                    kpZ_value = determine_kpZ(Z, pa)
                    st.number_input("Value of kpZ", value=kpZ_value, disabled=True, label_visibility="collapsed",on_change=update_edc)
                    print(f"kpZ_value: {kpZ_value}")

                col1_3, col1_4 = st.columns(2)
                with col1_3:
                    st.write("Importance Level")
                with col1_4:
                    importance_level = st.selectbox("Importance Level", ["2", "3", "4"], key="importance_level",label_visibility="collapsed", on_change=update_edc)

                #Structure Height
                col1_9, col1_10 = st.columns([1,1])
                with col1_9:
                    st.write("Structure Height (hn)- Unit:", f"[{unit_dist}]")
                with col1_10:
                    st.number_input("Structure Height (hn)", value=st.session_state.structure_height, disabled=True,label_visibility="collapsed",step=None)


                col1_11, col1_12,col1_12_desc = st.columns([1,0.5,0.5])
                with col1_11:
                    st.write("Seismic Design Category")
                with col1_12:
                    st.markdown(f"**{st.session_state.edc}**")
                desc = desc_edc(st.session_state.edc)
                with col1_12_desc:
                    st.write(f"[{desc}]")


                #방향별 파라미터 입력
                col1_13, col1_14, col1_15 = st.columns([1,0.5,0.5])
                with col1_13:
                    st.write("")
                with col1_14:
                    st.write("X-Direction")
                with col1_15:
                    st.write("Y-Direction")

                # sp
                col1_16, col1_17, col1_18 = st.columns([1, 0.5,0.5])
                with col1_16:
                    st.write("Structural Performance Factor (Sp)")
                with col1_17:
                    sp_x = st.number_input("Sp (X)", min_value=0.01, key="sp_x",label_visibility="collapsed",disabled=is_category_i(st.session_state.edc))
                with col1_18:
                    sp_y = st.number_input("Sp (Y)", min_value=0.01, key="sp_y",label_visibility="collapsed",disabled=is_category_i(st.session_state.edc))

               # mu
                col1_30, col1_31, col1_32 = st.columns([1, 0.5,0.5])
                with col1_30:
                    st.write("Structural Ductility Factor (μ)")
                with col1_31:
                    mu_x = st.number_input("μ (X)", min_value=0.01, key="mu_x",label_visibility="collapsed",disabled=is_category_i(st.session_state.edc))
                with col1_32:
                    mu_y = st.number_input("μ (Y)", min_value=0.01, key="mu_y",label_visibility="collapsed",disabled=is_category_i(st.session_state.edc))


                #Fundamental Period
                col1_20, col1_21, col1_22 = st.columns([1, 0.5,0.5])
                with col1_20:
                    st.write("Fundamental Period (sec)")
                with col1_21:
                    t1_x = st.number_input("Fundamental Period - X", min_value=0.01,  key="t1_x", label_visibility="collapsed",disabled=is_category_i(st.session_state.edc))
                with col1_22:
                    t1_y = st.number_input("Fundamental Period - Y", min_value=0.01,  key="t1_y", label_visibility="collapsed",disabled=is_category_i(st.session_state.edc))


                col1_23, col1_24 = st.columns([1,1])
                with col1_23:
                    st.write("")
                with col1_24:
                    if st.button("Period Calculator", use_container_width=True, disabled=is_category_i(st.session_state.edc)):
                        calc_t1_using_equation()


                ## 하중 할당 화면 설정
                stld_names = [v["NAME"] for v in st.session_state.stld_data.values()]
                stld_key_map = {v["NAME"]: int(k) for k, v in st.session_state.stld_data.items()}

                st.info("Load Cases to Apply to Seismic Loads")

                colA_1, colA_2 = st.columns(2)
                with colA_1:
                    st.write("Load Case for X-Direction")
                with colA_2:
                    load_case_x = st.selectbox("Load Case for X-Direction", stld_names,label_visibility="collapsed")

                colB_1, colB_2 = st.columns(2)
                with colB_1:
                    st.write("Load Case for Y-Direction")
                with colB_2:
                    load_case_y = st.selectbox("Load Case for Y-Direction", stld_names,label_visibility="collapsed")

                hn = st.session_state.structure_height


                if story_df is not None and not story_df.empty:
                    df_x = calculate_seismic_loads(story_df.copy(), t1_x, mu_x, sp_x, "X", soil_class, kpZ_value, Z)
                    df_y = calculate_seismic_loads(story_df.copy(), t1_y, mu_y, sp_y, "Y", soil_class, kpZ_value, Z)
                    st.session_state.df_x = df_x
                    st.session_state.df_y = df_y
                    st.session_state.calculated = True

                st.info("Applied Seismic Loads & Excel Download")
                refresh_button_col, spacer, applied_button_col, excel_button_col = st.columns([1, 1, 1.5, 1.5])

                with refresh_button_col:
                    if st.button("🔄 Refresh", use_container_width=True, type="secondary"):
                        try:
                            with st.spinner("Updating Data..."):
                                stld_data, story_data, unit_data = fetch(base_url, _mapi_key)
                                st.session_state.stld_data = stld_data
                                st.session_state.story_data = story_data
                                st.session_state.unit_data = unit_data
                                df = pd.DataFrame(story_data)
                                df["WEIGHT"] = df["WEIGHT"].astype(float)
                                df["ELEV"] = df["ELEV"].astype(float)
                                st.session_state.story_df = df
                                previous_height = st.session_state.structure_height
                                st.session_state.structure_height = df["ELEV"].max()
                                if previous_height != st.session_state.structure_height:
                                    update_edc()
                                st.rerun()
                        except Exception as e:
                            st.error(f"🔁 Refresh failed: {e}")

                with applied_button_col:
                    apply_clicked = st.button("Apply Seismic Loads", use_container_width=True, type="primary")

                with excel_button_col:
                    metadata = {
                            "Sub Soil Class": soil_class,
                            "Importance Level": importance_level,
                            "Annual Probability of Exceedance (pa)": pa,
                            "Hazard Factor (Z)": Z,
                            "Value of kpZ": kpZ_value,
                            "Structure Height": st.session_state.structure_height,
                            "Seismic Design Category": st.session_state.edc,
                            "Sp X": sp_x,
                            "Sp Y": sp_y,
                            "μ X": mu_x,
                            "μ Y": mu_y,
                            "Fundamental Period X": t1_x,
                            "Fundamental Period Y": t1_y,
                        }
                    ch_t_x = calc_ch_t(t1_x, soil_class)
                    ch_t_y = calc_ch_t(t1_y, soil_class)
                    cd_x = calc_seismic_coefficient(ch_t_x, sp_x, mu_x, kpZ_value, Z, t1_x)
                    cd_y = calc_seismic_coefficient(ch_t_y, sp_y, mu_y, kpZ_value, Z, t1_y)


                    with open("./resources/as1170_4_2024/as1170_4_2024_2.xlsx", "rb") as base_excel:
                        filled_excel = fill_template_excel(base_excel, metadata, cd_x, cd_y, df_x, df_y, unit_force, unit_dist)
                    st.markdown(
                        f'''
                        <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{base64.b64encode(filled_excel.getvalue()).decode()}"
                           download="as1170_4_2024_SeismicLoads.xlsx"
                           target="_parent"
                           style="text-decoration: none;">
                            <button style="
                                background-color: #3498db;
                                color: white;
                                padding: 0.3rem 0.3rem;
                                border: none;
                                border-radius: 0.25rem;
                                cursor: pointer;
                                width: 100%;
                                font-size: 1rem;">
                                Download Excel Data
                            </button>
                        </a>
                        ''',
                        unsafe_allow_html=True
                    )

                # Apply Seismic Loads 버튼 클릭 시 실행되는 로직
                if apply_clicked:
                    if load_case_x == load_case_y:
                        st.toast("Load Case - X and Load Case - Y must be different.", icon="❗")
                    else:
                        if st.session_state.edc == "III" and not st.session_state.confirm_apply:
                            confirm_apply_edc_iii()
                        else:
                            try:
                                headers = {"MAPI-Key": _mapi_key, "Content-Type": "application/json"}
                                post_url = f"{base_url}/db/seis/"

                                # X방향 하중 전송
                                body_x = {
                                    "Assign": {
                                        stld_key_map[load_case_x]: {
                                            "SEIS_LOAD_CODE": "USER TYPE",
                                            "DESC": "",
                                            "SCALE_FACTOR_X": 1,
                                            "SCALE_FACTOR_Y": 0,
                                            "SCALE_FACTOR_R": 0,
                                            "bACCIDENTAL_ECCEN": False,
                                            "bINHERENT_ECCEN": False,
                                            "USER": build_user_data(df_x, "X")
                                        }
                                    }
                                }

                                res_x = requests.put(post_url, headers=headers, json=body_x)
                                if res_x.status_code == 200:
                                    st.toast("X-direction seismic loads successfully applied.", icon="✅")
                                else:
                                    try:
                                        error_detail = res_x.json().get("message") or res_x.text
                                    except Exception:
                                        error_detail = res_x.text
                                    st.toast(f"❌ Failed to apply X-direction loads.\n\n**Status Code:** {res_x.status_code}\n**Details:** {error_detail}")

                                # Y방향 하중 전송
                                body_y = {
                                    "Assign": {
                                        stld_key_map[load_case_y]: {
                                            "SEIS_LOAD_CODE": "USER TYPE",
                                            "DESC": "",
                                            "SCALE_FACTOR_X": 0,
                                            "SCALE_FACTOR_Y": 1,
                                            "SCALE_FACTOR_R": 0,
                                            "bACCIDENTAL_ECCEN": False,
                                            "bINHERENT_ECCEN": False,
                                            "USER": build_user_data(df_y, "Y")
                                        }
                                    }
                                }

                                res_y = requests.put(post_url, headers=headers, json=body_y)
                                if res_y.status_code == 200:
                                    st.toast("Y-direction seismic loads successfully applied.", icon="✅")
                                else:
                                    try:
                                        error_detail = res_y.json().get("message") or res_y.text
                                    except Exception:
                                        error_detail = res_y.text
                                    st.toast(f"❌ Failed to apply Y-direction loads.\n\n**Status Code:** {res_y.status_code}\n**Details:** {error_detail}")

                            except Exception as e:
                                st.toast(f"❗ Unexpected error occurred while sending API request.\n\n**Exception:** {e}")
                            finally:
                                st.session_state.confirm_apply = False

        # 오른쪽 화면 설정
        with c2:
            with st.container():
                # st.divider()
                st.info("Graphs and Tables")
                # st.markdown("*Graphs and Tables*")
                if st.session_state.calculated:
                    df_x = st.session_state.df_x
                    df_y = st.session_state.df_y

                    tab1, tab2 = st.tabs(["Graph", "Table"])
                    with tab1:
                        col1_1, col1_2 = st.columns(2)
                        with col1_1:
                            st.write("Select Graph Type")
                        with col1_2:
                            graph_type = st.selectbox("Graph Type", ["Story Force", "Story Shear", "Overturning Moment"], key="graph_type",label_visibility="collapsed")

                        col2_1, col2_2 = st.columns(2)
                        with col2_1:
                            plot_graph(df_x, f"{graph_type} for X", "X", unit_force, unit_dist)
                        with col2_2:
                            plot_graph(df_y, f"{graph_type} for Y", "Y", unit_force, unit_dist)

                    with tab2:
                        col2_1, col2_2 = st.columns(2)
                        with col2_1:
                            st.write("Select Table Direction")
                        with col2_2:
                            table_dir = st.selectbox(
                                "Select Table Direction",
                                options=["X-Direction", "Y-Direction"],
                                index=0,
                                key="table_dir",
                                label_visibility="collapsed"
                            )

                        # 테이블 숨길 열 목록
                        hidden_columns = ["LOADED_H", "LOADED_BX", "LOADED_BY", "Wi*Hi^k for X", "Wi*Hi^k for Y"]

                        def label(column, force_unit, dist_unit):
                            if "Overturning Moment" in column:
                                return f"{column} [{force_unit}·{dist_unit}]"
                            elif "Story Shear" in column or "Story Force" in column:
                                return f"{column} [{force_unit}]"
                            else:
                                return column

                        if table_dir == "X-Direction":
                            display_df = df_x.drop(columns=[col for col in hidden_columns if col in df_x.columns])
                            display_df = display_df.rename(columns={
                                col: label(col, unit_force, unit_dist)
                                for col in display_df.columns
                                if "Story Force" in col or "Story Shear" in col or "Overturning Moment" in col
                            })
                            # 소수점 2자리로 반올림
                            display_df = display_df.round(2)
                        else:
                            display_df = df_y.drop(columns=[col for col in hidden_columns if col in df_y.columns])
                            display_df = display_df.rename(columns={
                                col: label(col, unit_force, unit_dist)
                                for col in display_df.columns
                                if "Story Force" in col or "Story Shear" in col or "Overturning Moment" in col
                            })
                            # 소수점 2자리로 반올림
                            display_df = display_df.round(2)


                        st.dataframe(
                            display_df,
                            height=600,
                            use_container_width=True,
                            hide_index=True
                        )

                else:
                    st.info("Please fill in parameters to view results.")





